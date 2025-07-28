import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    Garante que valores inteiros permaneçam inteiros
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def get_pontos_config(df, linha_inicial, avanca_linha):
    """Identifica os pontos de calibração válidos na planilha"""
    pontos_validos = []
    inicio_linha = linha_inicial
    num_ponto = 1
    
    while True:
        valores_nulos = 0
        for i in range(3): 
            pulsos = get_numeric_value(df, inicio_linha + 3 + i, 2)
            if pulsos == 0 or pd.isna(pulsos):
                valores_nulos += 1
        
        if valores_nulos == 3:
            break
            
        ponto_config = {
            'inicio_linha': inicio_linha,
            'num_leituras': 3,
            'num_ponto': num_ponto
        }
        pontos_validos.append(ponto_config)
        inicio_linha += avanca_linha
        num_ponto += 1
    
    return pontos_validos

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    Fórmula: =SE(U54="";"";STDEV.S(U54:U56))
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos (equivalente ao SE(U54="";"";...))
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def extrair_vazoes_referencia_tendencia_desvio(arquivo_excel):
    """Extrai as vazões de referência, tendências e desvios padrão da aba 'Coleta de Dados'"""
    try:
        print(f"Lendo vazões de referência, tendências e desvios padrão do arquivo: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        pontos_config = get_pontos_config(coleta_df, 50, 9)
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        pontos_dados = []
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'vazoes_referencia': [],
                'erros': [],
                'tendencia': None,
                'desvio_padrao': None
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i  # +4 em vez de +3 para pular a linha do título
                
                # Lê a vazão de referência da coluna I (coluna 9)
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)
                
                # Lê o erro da coluna U (coluna 21) - fórmula: =SE(O54="";"";(O54-L54)/L54*100)
                erro = ler_valor_exato(coleta_sheet, linha, 21)
                
                ponto['vazoes_referencia'].append({
                    'linha': linha,
                    'vazao_referencia': vazao_referencia
                })
                
                ponto['erros'].append({
                    'linha': linha,
                    'erro': erro
                })
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Erro: {float(erro)} %")

            # Calcula a média das vazões de referência para este ponto
            if ponto['vazoes_referencia']:
                vazoes = [v['vazao_referencia'] for v in ponto['vazoes_referencia']]
                media_vazao = sum(vazoes) / Decimal(str(len(vazoes)))
                ponto['media_vazao_referencia'] = media_vazao
                
                print(f"   Média Vazão do Ponto {config['num_ponto']}: {float(media_vazao)} L/h")
            
            # Calcula a tendência (média dos erros) - fórmula: =SE(U54="";"";MÉDIA(U54:U56))
            if ponto['erros']:
                erros = [e['erro'] for e in ponto['erros']]
                # Filtra valores não nulos (equivalente ao SE(U54="";"";...))
                erros_validos = [e for e in erros if e != 0]
                if erros_validos:
                    tendencia = sum(erros_validos) / Decimal(str(len(erros_validos)))
                    ponto['tendencia'] = tendencia
                    print(f"   Tendência do Ponto {config['num_ponto']}: {float(tendencia)} %")
                else:
                    print(f"   Tendência do Ponto {config['num_ponto']}: N/A (todos os erros são zero)")
            
            # Calcula o desvio padrão amostral - fórmula: =SE(U54="";"";STDEV.S(U54:U56))
            if ponto['erros']:
                erros = [e['erro'] for e in ponto['erros']]
                desvio_padrao = calcular_desvio_padrao_amostral(erros)
                if desvio_padrao is not None:
                    ponto['desvio_padrao'] = desvio_padrao
                    print(f"   Desvio Padrão do Ponto {config['num_ponto']}: {float(desvio_padrao)} %")
                else:
                    print(f"   Desvio Padrão do Ponto {config['num_ponto']}: N/A (insuficientes dados válidos)")

            pontos_dados.append(ponto)
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['vazoes_referencia'])} leituras extraídas")
        
        return pontos_dados
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair vazões de referência: {e}")
        return None

def main():
    """Função principal para extrair vazões de referência, tendências e desvios padrão"""
    arquivo_excel = "SAN-038-25-09-1.xlsx"
    
    print("=== EXTRATOR DE VAZÕES DE REFERÊNCIA, TENDÊNCIAS E DESVIOS PADRÃO ===")
    print("Extrai vazões de referência (coluna I), tendências (coluna U) e desvios padrão da aba 'Coleta de Dados'")
    print("Usa precisão Decimal de 28 dígitos (igual ao script_final_precisao.py)")
    
    # Extrai as vazões de referência, tendências e desvios padrão
    pontos = extrair_vazoes_referencia_tendencia_desvio(arquivo_excel)
    
    if pontos:
        print(f"\nExtração concluída com sucesso!")
        print(f"Total de pontos extraídos: {len(pontos)}")
        
        # Mostra informações sobre precisão
        print(f"\nINFORMACOES DE PRECISAO:")
        print(f"   • Precisão Decimal: {getcontext().prec} dígitos")
        print(f"   • Método de leitura: openpyxl com data_only=True")
        print(f"   • Conversão: converter_para_decimal_padrao()")
        
        # Cria tabela formatada
        print("\n" + "="*170)
        print("TABELA DE VAZÕES DE REFERÊNCIA, ERROS E DESVIOS PADRÃO")
        print("="*170)
        
        # Cabeçalho da tabela
        header = f"{'Ponto':<6} {'Leitura':<8} {'Linha':<6} {'Vazão Ref (L/h)':<20} {'Erro (%)':<12} {'Média Vazão':<15} {'Tendência (%)':<15} {'Desvio Padrão (%)':<18} {'Erros para Cálculo':<35}"
        print(header)
        print("-"*170)
        
        # Dados da tabela
        for ponto in pontos:
            for i, (vazao, erro) in enumerate(zip(ponto['vazoes_referencia'], ponto['erros']), 1):
                # Formata os erros usados para calcular a tendência
                erros_formatados = []
                for e in ponto['erros']:
                    if e['erro'] != 0:  # Filtra valores não nulos
                        erros_formatados.append(f"{float(e['erro']):.2f}%")
                
                erros_str = ", ".join(erros_formatados) if erros_formatados else "N/A"
                
                # Formata a tendência
                tendencia_str = f"{float(ponto['tendencia']):.2f}%" if ponto['tendencia'] else "N/A"
                
                # Formata o desvio padrão
                desvio_str = f"{float(ponto['desvio_padrao']):.2f}%" if ponto['desvio_padrao'] else "N/A"
                
                linha = (f"{ponto['numero']:<6} "
                        f"{i:<8} "
                        f"{vazao['linha']:<6} "
                        f"{float(vazao['vazao_referencia']):<20.15f} "
                        f"{float(erro['erro']):<12.2f} "
                        f"{float(ponto['media_vazao_referencia']):<15.2f} "
                        f"{tendencia_str:<15} "
                        f"{desvio_str:<18} "
                        f"{erros_str:<35}")
                print(linha)
            
            # Linha de resumo do ponto
            print("-"*170)
        
        # Tabela resumida final
        print("\n" + "="*100)
        print("RESUMO GERAL - MÉDIAS, TENDÊNCIAS E DESVIOS PADRÃO")
        print("="*100)
        print(f"{'Ponto':<6} {'Média Vazão Ref (L/h)':<25} {'Tendência (%)':<15} {'Desvio Padrão (%)':<18} {'Nº Leituras':<12}")
        print("-"*100)
        
        for ponto in pontos:
            tendencia_str = f"{float(ponto['tendencia']):.2f}" if ponto['tendencia'] else "N/A"
            desvio_str = f"{float(ponto['desvio_padrao']):.2f}" if ponto['desvio_padrao'] else "N/A"
            resumo = (f"{ponto['numero']:<6} "
                     f"{float(ponto['media_vazao_referencia']):<25.2f} "
                     f"{tendencia_str:<15} "
                     f"{desvio_str:<18} "
                     f"{len(ponto['vazoes_referencia']):<12}")
            print(resumo)
        
        print("-"*100)
        
        # Salva em JSON para uso posterior
        output_data = {
            "metadata": {
                "arquivo": arquivo_excel,
                "data_extracao": datetime.now().isoformat(),
                "total_pontos": len(pontos),
                "descricao": "Vazões de referência, tendências e desvios padrão extraídos da aba 'Coleta de Dados'",
                "precisao": {
                    "decimal_precision": getcontext().prec,
                    "metodo_leitura": "openpyxl com data_only=True",
                    "colunas": {
                        "vazao_referencia": "I (coluna 9)",
                        "erro": "U (coluna 21)",
                        "formula_erro": "=SE(O54=\"\";\"\";(O54-L54)/L54*100)",
                        "formula_tendencia": "=SE(U54=\"\";\"\";MÉDIA(U54:U56))",
                        "formula_desvio_padrao": "=SE(U54=\"\";\"\";STDEV.S(U54:U56))"
                    }
                }
            },
            "pontos_dados": pontos
        }
        
        with open("vazoes_referencia_tendencia_desvio.json", "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\nDados salvos em 'vazoes_referencia_tendencia_desvio.json'")
        print(f"   • Inclui todas as vazões de referência por ponto")
        print(f"   • Inclui todos os erros por ponto")
        print(f"   • Inclui tendências calculadas (média dos erros)")
        print(f"   • Inclui desvios padrão calculados (STDEV.S)")
        print(f"   • Médias calculadas com precisão Decimal")
        print(f"   • Metadados de precisão incluídos")
        
    else:
        print("Falha na extração das vazões de referência, tendências e desvios padrão")

if __name__ == "__main__":
    main() 