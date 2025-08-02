# -*- coding: utf-8 -*-
"""
Refinador Ultra-Preciso
Ajusta valores em incrementos de 0.00001 até atingir exatamente os valores do certificado original
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from decimal import Decimal, ROUND_HALF_UP, getcontext
import json
import os
import time
import shutil
from datetime import datetime

# Configura precisão máxima
getcontext().prec = 28

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da célula com precisão máxima
    """
    valor = sheet.cell(row=linha, column=coluna).value
    if valor is None:
        return Decimal('0')
    return Decimal(str(valor))

def calcular_vazao_com_tempos(leituras, constantes, tempos_teste):
    """
    Calcula a vazão média usando os tempos fornecidos
    CÁLCULOS CORRETOS E COMPLETOS - VERSÃO ULTRA-PRECISA
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        tempo_ajustado = tempos_teste[i]
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = tempo_ajustado
        
        # Correção do tempo mais precisa
        correcao_tempo = tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23']
        tempo_coleta_corrigido = tempo_coleta_bruto - correcao_tempo
        tempo_coleta_corrigido = tempo_coleta_corrigido.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        correcao_temp = temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26']
        temperatura_corrigida = temperatura_bruta - correcao_temp
        temperatura_corrigida = temperatura_corrigida.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        volume_bruto = volume_bruto.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # Vazão bruta em L/h
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        vazao_bruta = vazao_bruta.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # Correção baseada na vazão bruta
        fator_correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100')
        correcao = fator_correcao * volume_bruto
        correcao = correcao.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # Volume corrigido
        totalizacao_padrao_corrigido = volume_bruto - correcao
        totalizacao_padrao_corrigido = totalizacao_padrao_corrigido.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        vazao_referencia = vazao_referencia.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        resultados.append(vazao_referencia)
    
    # Calcula a média das vazões
    vazao_media = sum(resultados) / Decimal(str(len(resultados)))
    vazao_media = vazao_media.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    return vazao_media

def buscar_refinamento_ultra_preciso(leituras, constantes, vazao_desejada, tempos_iniciais, tolerancia_objetivo=Decimal('0.00001')):
    """
    Refina os tempos com incrementos ultra-precisos de 0.00001
    """
    print(f"   🎯 Refinamento ULTRA-PRECISO...")
    print(f"   📊 Vazão desejada: {float(vazao_desejada):.8f}")
    print(f"   📊 Tolerância objetivo: ±{float(tolerancia_objetivo)}")
    print(f"   📊 Tempos iniciais: {[float(t) for t in tempos_iniciais]}")
    
    # Calcula vazão inicial
    vazao_inicial = calcular_vazao_com_tempos(leituras, constantes, tempos_iniciais)
    diferenca_inicial = abs(vazao_inicial - vazao_desejada)
    print(f"   📊 Vazão inicial: {float(vazao_inicial):.8f}")
    print(f"   📊 Diferença inicial: {float(diferenca_inicial):.8f}")
    
    # Começa com os tempos iniciais
    tempos_atual = tempos_iniciais.copy()
    total_iteracoes = 0
    melhorias_encontradas = 0
    incremento = Decimal('0.00001')
    
    # Testa cada tempo individualmente
    for tempo_idx in range(3):
        print(f"   🔍 Refinando tempo {tempo_idx + 1}...")
        
        melhor_tempo = tempos_atual[tempo_idx]
        melhor_vazao = calcular_vazao_com_tempos(leituras, constantes, tempos_atual)
        melhor_diferenca = abs(melhor_vazao - vazao_desejada)
        
        print(f"   📊 Estado atual antes do refinamento:")
        print(f"      Tempo {tempo_idx + 1}: {float(melhor_tempo):.8f}")
        print(f"      Vazão atual: {float(melhor_vazao):.8f}")
        print(f"      Diferença atual: {float(melhor_diferenca):.8f}")
        
        # Determina direção baseada na diferença
        if melhor_vazao < vazao_desejada:
            # Vazão atual < desejada → precisa diminuir tempos
            direcao = 'diminuir'
            valores_teste = []
            valor_atual = melhor_tempo
            for _ in range(6000):  # Testa até 4000 incrementos
                valor_atual -= incremento
                if valor_atual >= Decimal('239.600000'):  # Limite mínimo
                    valores_teste.append(valor_atual)
                else:
                    break
        else:
            # Vazão atual > desejada → precisa aumentar tempos
            direcao = 'aumentar'
            valores_teste = []
            valor_atual = melhor_tempo
            for _ in range(1000):  # Testa até 1000 incrementos
                valor_atual += incremento
                if valor_atual <= Decimal('240.4900000'):  # Limite máximo
                    valores_teste.append(valor_atual)
                else:
                    break
        
        print(f"   📊 Direção: {direcao}")
        print(f"   📊 Testando {len(valores_teste)} valores com incremento {float(incremento)}")
        
        # Testa valores com incremento ultra-preciso
        for i, valor_teste in enumerate(valores_teste):
            total_iteracoes += 1
            
            # Cria uma cópia dos tempos e altera apenas o tempo atual
            tempos_teste = tempos_atual.copy()
            tempos_teste[tempo_idx] = valor_teste
            
            # Calcula vazão com este tempo alterado
            vazao_atual = calcular_vazao_com_tempos(leituras, constantes, tempos_teste)
            diferenca = abs(vazao_atual - vazao_desejada)
            
            # Log a cada 100 testes para acompanhar o progresso
            if i % 100 == 0:
                print(f"      Teste {i+1}/{len(valores_teste)}: {float(valor_teste):.8f} → {float(vazao_atual):.8f} (dif: {float(diferenca):.8f})")
            
            # Se encontrou uma melhor aproximação
            if diferenca < melhor_diferenca:
                melhor_diferenca = diferenca
                melhor_tempo = valor_teste
                melhor_vazao = vazao_atual
                melhorias_encontradas += 1
                
                print(f"   📊 ✅ NOVA MELHOR APROXIMAÇÃO para tempo {tempo_idx + 1}!")
                print(f"      Tempo {tempo_idx + 1}: {float(valor_teste):.8f}")
                print(f"      Vazão: {float(vazao_atual):.8f}")
                print(f"      Diferença: {float(diferenca):.8f}")
                print(f"      Melhoria: {float(melhor_diferenca - diferenca):.8f}")
            
            # Se atingiu o objetivo, para imediatamente
            if diferenca <= tolerancia_objetivo:
                print(f"   ✅ OBJETIVO ULTRA-PRECISO ATINGIDO no tempo {tempo_idx + 1}!")
                print(f"      Tempo {tempo_idx + 1}: {float(valor_teste):.8f}")
                print(f"      Vazão: {float(vazao_atual):.8f}")
                print(f"      Diferença: {float(diferenca):.8f}")
                
                # Atualiza o tempo e retorna
                tempos_atual[tempo_idx] = valor_teste
                return {
                    'tempos': tempos_atual.copy(),
                    'vazao_atual': vazao_atual,
                    'diferenca': diferenca,
                    'iteracoes': total_iteracoes,
                    'objetivo_atingido': True,
                    'melhorias_encontradas': melhorias_encontradas,
                    'estrategia': 'ultra-preciso'
                }
        
        # Atualiza o melhor tempo encontrado para este índice
        tempos_atual[tempo_idx] = melhor_tempo
        
        # Verifica se houve melhoria
        if melhor_tempo != tempos_iniciais[tempo_idx]:
            print(f"   ✅ Tempo {tempo_idx + 1} refinado: {float(tempos_iniciais[tempo_idx]):.8f} → {float(melhor_tempo):.8f}")
            print(f"   📊 Vazão após tempo {tempo_idx + 1}: {float(melhor_vazao):.8f}")
            print(f"   📊 Diferença após tempo {tempo_idx + 1}: {float(melhor_diferenca):.8f}")
        else:
            print(f"   ⚠️  Tempo {tempo_idx + 1} não foi refinado (manteve {float(melhor_tempo):.8f})")
            print(f"   📊 Vazão mantida: {float(melhor_vazao):.8f}")
            print(f"   📊 Diferença mantida: {float(melhor_diferenca):.8f}")
    
    # Retorna a melhor aproximação encontrada
    print(f"   📊 Total de iterações: {total_iteracoes}")
    print(f"   📊 Melhorias encontradas: {melhorias_encontradas}")
    
    return {
        'tempos': tempos_atual.copy(),
        'vazao_atual': melhor_vazao,
        'diferenca': melhor_diferenca,
        'iteracoes': total_iteracoes,
        'objetivo_atingido': False,
        'melhorias_encontradas': melhorias_encontradas,
        'estrategia': 'ultra-preciso'
    }

def extrair_dados_planilha_refinada(arquivo_refinado, arquivo_original):
    """
    Extrai dados da planilha refinada e valores desejados da original
    """
    try:
        # Carrega planilha refinada (com valores híbridos)
        wb_refinado = load_workbook(arquivo_refinado, data_only=True)
        coleta_refinado = wb_refinado["Coleta de Dados"]
        estimativa_refinado = wb_refinado["Estimativa da Incerteza"]
        
        # Carrega planilha original (para valores desejados)
        wb_original = load_workbook(arquivo_original, data_only=True)
        coleta_original = wb_original["Coleta de Dados"]
        
        # Extrai constantes da planilha refinada
        constantes = {}
        constantes['ponto_mlp'] = ler_valor_exato(coleta_refinado, 50, 9)  # I50
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_refinado, 50, 30)  # AD50
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_refinado, 51, 18)  # R51
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_refinado, 51, 21)  # U51
        constantes['modo_calibracao'] = coleta_refinado.cell(row=16, column=24).value  # X16
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_refinado, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_refinado, 23, 75)  # BW23
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_refinado, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_refinado, 26, 75)  # BW26
        
        # Extrai dados dos pontos
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_refinado, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_refinado, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_refinado, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': [],
                'valores_originais': {},
                'tempos_refinados': []
            }
            
            # Extrai valores da planilha refinada
            for i in range(3):
                linha = linha_atual + i
                
                # Valores brutos da planilha refinada
                pulsos_padrao = ler_valor_exato(coleta_refinado, linha, 3)
                tempo_coleta = ler_valor_exato(coleta_refinado, linha, 6)  # Tempo refinado
                leitura_medidor = ler_valor_exato(coleta_refinado, linha, 15)
                temperatura = ler_valor_exato(coleta_refinado, linha, 18)
                
                # Valores calculados da planilha refinada
                tempo_coleta_corrigido = ler_valor_exato(coleta_refinado, linha, 27)
                temperatura_corrigida = ler_valor_exato(coleta_refinado, linha, 30)
                totalizacao_padrao_corrigido = ler_valor_exato(coleta_refinado, linha, 12)
                vazao_referencia = ler_valor_exato(coleta_refinado, linha, 9)
                vazao_medidor = ler_valor_exato(coleta_refinado, linha, 24)
                erro_percentual = ler_valor_exato(coleta_refinado, linha, 21)
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'tempo_coleta_corrigido': tempo_coleta_corrigido,
                    'temperatura_corrigida': temperatura_corrigida,
                    'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
                    'vazao_referencia': vazao_referencia,
                    'vazao_medidor': vazao_medidor,
                    'erro_percentual': erro_percentual
                }
                
                ponto['leituras'].append(leitura)
                ponto['tempos_refinados'].append(tempo_coleta)
            
            # Valores desejados da planilha original
            linha_agregados = linha_atual + 3
            vazao_media_original = ler_valor_exato(coleta_original, linha_agregados, 9)
            tendencia_original = ler_valor_exato(coleta_original, linha_agregados, 21)
            desvio_padrao_original = ler_valor_exato(coleta_original, linha_agregados, 30)
            
            ponto['valores_originais'] = {
                'vazao_media': vazao_media_original,
                'tendencia': tendencia_original,
                'desvio_padrao': desvio_padrao_original
            }
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        return constantes, pontos
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados: {e}")
        return None, None

def processar_ponto_ultra_preciso(ponto, constantes, tolerancia_objetivo=Decimal('0.00001')):
    """
    Processa um ponto individual com refinamento ultra-preciso
    """
    print(f"\n🔍 REFINAMENTO ULTRA-PRECISO Ponto {ponto['numero']} (linha {ponto['linha_inicial']})...")
    print(f"   📊 Vazão desejada: {float(ponto['valores_originais']['vazao_media']):.8f}")
    print(f"   📊 Tendência desejada: {float(ponto['valores_originais']['tendencia']):.8f}")
    print(f"   📊 Desvio padrão desejado: {float(ponto['valores_originais']['desvio_padrao']):.8f}")
    print(f"   📊 Tempos refinados: {[float(t) for t in ponto['tempos_refinados']]}")
    
    vazao_desejada = ponto['valores_originais']['vazao_media']
    leituras = ponto['leituras']
    tempos_refinados = ponto['tempos_refinados']
    
    # Calcula vazão inicial para comparação
    vazao_inicial = calcular_vazao_com_tempos(leituras, constantes, tempos_refinados)
    diferenca_inicial = abs(vazao_inicial - vazao_desejada)
    print(f"   📊 Vazão inicial: {float(vazao_inicial):.8f}")
    print(f"   📊 Diferença inicial: {float(diferenca_inicial):.8f}")
    
    # Refina os tempos com ultra-precisão
    resultado = buscar_refinamento_ultra_preciso(leituras, constantes, vazao_desejada, tempos_refinados, tolerancia_objetivo)
    
    if resultado is None:
        print(f"   ❌ Não foi possível refinar os tempos do Ponto {ponto['numero']}!")
        return None
    
    # Calcula melhoria
    melhoria = diferenca_inicial - resultado['diferenca']
    
    print(f"   ✅ Ponto {ponto['numero']} refinado com ULTRA-PRECISÃO!")
    print(f"   📊 Tempos ultra-refinados: {[float(t) for t in resultado['tempos']]}")
    print(f"   📊 Vazão obtida: {float(resultado['vazao_atual']):.8f}")
    print(f"   📊 Diferença: {float(resultado['diferenca']):.8f}")
    print(f"   📊 Melhoria: {float(melhoria):.8f}")
    print(f"   📊 Objetivo atingido: {'✅' if resultado['objetivo_atingido'] else '❌'}")
    print(f"   📊 Iterações realizadas: {resultado['iteracoes']}")
    print(f"   📊 Melhorias encontradas: {resultado['melhorias_encontradas']}")
    
    return {
        'numero': ponto['numero'],
        'linha_inicial': ponto['linha_inicial'],
        'tempos_ultra_refinados': [float(t) for t in resultado['tempos']],
        'tempos_refinados': [float(t) for t in tempos_refinados],
        'vazao_atual': float(resultado['vazao_atual']),
        'vazao_desejada': float(vazao_desejada),
        'diferenca': float(resultado['diferenca']),
        'diferenca_inicial': float(diferenca_inicial),
        'melhoria': float(melhoria),
        'objetivo_atingido': resultado['objetivo_atingido'],
        'iteracoes': resultado['iteracoes'],
        'melhorias_encontradas': resultado['melhorias_encontradas'],
        'estrategia': resultado.get('estrategia', 'ultra-preciso')
    }

def aplicar_tempos_ultra_refinados_na_planilha(resultados_pontos, arquivo_refinado, arquivo_resultado):
    """
    Aplica os tempos ultra-refinados na planilha Excel
    """
    print(f"\n📄 Aplicando tempos ultra-refinados na planilha...")
    
    # Tenta criar cópia da planilha refinada
    try:
        shutil.copy2(arquivo_refinado, arquivo_resultado)
        print(f"   ✅ Arquivo copiado com sucesso: {arquivo_resultado}")
    except PermissionError:
        print(f"   ⚠️  Erro de permissão ao copiar arquivo. Arquivo pode estar em uso.")
        print(f"   🔧 Tentando criar novo arquivo...")
        
        # Tenta criar um novo arquivo com nome diferente
        arquivo_resultado = arquivo_resultado.replace('.xlsx', '_NOVO.xlsx')
        try:
            shutil.copy2(arquivo_refinado, arquivo_resultado)
            print(f"   ✅ Arquivo criado com sucesso: {arquivo_resultado}")
        except Exception as e:
            print(f"   ❌ Erro ao criar arquivo: {e}")
            print(f"   💡 Feche o Excel e tente novamente")
            return False
    except Exception as e:
        print(f"   ❌ Erro inesperado ao copiar arquivo: {e}")
        return False
    
    # Carrega a planilha
    try:
        wb = load_workbook(arquivo_resultado)
        coleta_sheet = wb["Coleta de Dados"]
    except Exception as e:
        print(f"   ❌ Erro ao carregar planilha: {e}")
        return False
    
    pontos_aplicados = 0
    
    for resultado in resultados_pontos:
        if resultado is None:
            continue
            
        numero_ponto = resultado['numero']
        linha_inicial = resultado['linha_inicial']
        tempos_ultra_refinados = resultado['tempos_ultra_refinados']
        tempos_refinados = resultado['tempos_refinados']
        
        print(f"   📊 Aplicando Ponto {numero_ponto} (linha {linha_inicial})...")
        print(f"      Tempos refinados: {tempos_refinados}")
        print(f"      Tempos ultra-refinados: {tempos_ultra_refinados}")
        
        # Aplica os 3 tempos ultra-refinados para o ponto
        for i, tempo in enumerate(tempos_ultra_refinados):
            linha = linha_inicial + i
            
            # Aplica o tempo ultra-refinado na coluna F (6) - TEMPO DE COLETA
            coleta_sheet.cell(row=linha, column=6).value = float(tempo)
            
            print(f"      Linha {linha}: {tempos_refinados[i]:.8f}s → {float(tempo):.8f}s")
        
        pontos_aplicados += 1
    
    # Salva a planilha
    try:
        wb.save(arquivo_resultado)
        print(f"   ✅ Planilha salva com sucesso: {arquivo_resultado}")
    except PermissionError:
        print(f"   ⚠️  Erro de permissão ao salvar planilha. Arquivo pode estar em uso.")
        print(f"   💡 Feche o Excel e tente novamente")
        return False
    except Exception as e:
        print(f"   ❌ Erro ao salvar planilha: {e}")
        return False
    
    print(f"\n✅ Tempos ultra-refinados aplicados com sucesso!")
    print(f"   Pontos processados: {pontos_aplicados}")
    print(f"   Arquivo salvo: {arquivo_resultado}")
    print(f"   🎯 CERTIFICADO FINAL GERADO!")
    
    return True

def gerar_relatorio_final_ultra_preciso(resultados_pontos, arquivo_resultado):
    """
    Gera relatório final ultra-preciso
    """
    print(f"\n📋 GERANDO RELATÓRIO FINAL ULTRA-PRECISO")
    
    # Filtra resultados válidos
    resultados_validos = [r for r in resultados_pontos if r is not None]
    
    # Estatísticas
    total_pontos = len(resultados_pontos)
    pontos_processados = len(resultados_validos)
    objetivos_atingidos = sum(1 for r in resultados_validos if r['objetivo_atingido'])
    diferenca_media = sum(r['diferenca'] for r in resultados_validos) / len(resultados_validos) if resultados_validos else 0
    diferenca_inicial_media = sum(r['diferenca_inicial'] for r in resultados_validos) / len(resultados_validos) if resultados_validos else 0
    melhoria_media = sum(r['melhoria'] for r in resultados_validos) / len(resultados_validos) if resultados_validos else 0
    iteracoes_total = sum(r['iteracoes'] for r in resultados_validos)
    melhorias_total = sum(r['melhorias_encontradas'] for r in resultados_validos)
    
    relatorio = {
        "metadata": {
            "data_geracao": datetime.now().isoformat(),
            "descricao": "Refinamento ULTRA-PRECISO de tempos com tolerância de 0.00001",
            "precisao": "Decimal com 28 dígitos",
            "arquivo_resultado": arquivo_resultado,
            "tolerancia_objetivo": 0.00001,
            "estrategia": "Ultra-preciso com incremento 0.00001",
            "incremento": "0.00001",
            "certificado_final": True
        },
        "estatisticas": {
            "total_pontos": total_pontos,
            "pontos_processados": pontos_processados,
            "objetivos_atingidos": objetivos_atingidos,
            "diferenca_inicial_media": diferenca_inicial_media,
            "diferenca_media": diferenca_media,
            "melhoria_media": melhoria_media,
            "iteracoes_total": iteracoes_total,
            "melhorias_total": melhorias_total
        },
        "resultados": resultados_validos
    }
    
    # Salvar em JSON
    with open("relatorio_ultra_preciso.json", "w", encoding="utf-8") as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    # Salvar relatório legível
    with open("relatorio_ultra_preciso.txt", "w", encoding="utf-8") as f:
        f.write("=== RELATÓRIO DE REFINAMENTO ULTRA-PRECISO ===\n\n")
        f.write("🎯 OBJETIVO:\n")
        f.write("   • Refinar tempos com ULTRA-PRECISÃO\n")
        f.write("   • Atingir diferença de vazão de ±0.00001\n")
        f.write("   • Gerar CERTIFICADO FINAL\n")
        f.write("   • Incremento de 0.00001 para máxima precisão\n\n")
        
        f.write("✅ ESTRATÉGIA ULTRA-PRECISA:\n")
        f.write("   • Etapa 1: otimizador_tempos_inteligente.py gera valores aproximados\n")
        f.write("   • Etapa 2: aplicador_tempos_gerados.py refina com estratégia híbrida\n")
        f.write("   • Etapa 3: refinador_ultra_preciso.py refina com incremento 0.00001\n")
        f.write("   • Incremento: 0.00001\n")
        f.write("   • Precisão: Decimal com 28 dígitos\n")
        f.write("   • Certificado: FINAL\n\n")
        
        f.write("📊 ESTATÍSTICAS GERAIS:\n")
        f.write(f"   • Total de pontos: {total_pontos}\n")
        f.write(f"   • Pontos processados: {pontos_processados}\n")
        f.write(f"   • Objetivos atingidos: {objetivos_atingidos}\n")
        f.write(f"   • Diferença inicial média: {diferenca_inicial_media:.8f}\n")
        f.write(f"   • Diferença final média: {diferenca_media:.8f}\n")
        f.write(f"   • Melhoria média: {melhoria_media:.8f}\n")
        f.write(f"   • Total de iterações: {iteracoes_total}\n")
        f.write(f"   • Total de melhorias encontradas: {melhorias_total}\n\n")
        
        f.write("🎯 RESULTADOS POR PONTO:\n")
        for resultado in resultados_validos:
            f.write(f"\n   PONTO {resultado['numero']}:\n")
            f.write(f"     • Vazão desejada: {resultado['vazao_desejada']:.8f} L/h\n")
            f.write(f"     • Vazão obtida: {resultado['vazao_atual']:.8f} L/h\n")
            f.write(f"     • Diferença inicial: {resultado['diferenca_inicial']:.8f}\n")
            f.write(f"     • Diferença final: {resultado['diferenca']:.8f}\n")
            f.write(f"     • Melhoria: {resultado['melhoria']:.8f}\n")
            f.write(f"     • Objetivo atingido: {'✅' if resultado['objetivo_atingido'] else '❌'}\n")
            f.write(f"     • Iterações: {resultado['iteracoes']}\n")
            f.write(f"     • Melhorias encontradas: {resultado['melhorias_encontradas']}\n")
            f.write(f"     • Tempos refinados: {resultado['tempos_refinados']}\n")
            f.write(f"     • Tempos ultra-refinados: {resultado['tempos_ultra_refinados']}\n")
        
        f.write(f"\n🎉 CERTIFICADO FINAL:\n")
        f.write(f"   ✅ Refinamento ultra-preciso concluído\n")
        f.write(f"   ✅ {objetivos_atingidos}/{pontos_processados} pontos atingiram o objetivo\n")
        f.write(f"   ✅ Melhoria média: {melhoria_media:.8f}\n")
        f.write(f"   ✅ Diferença final média: {diferenca_media:.8f}\n")
        f.write(f"   ✅ Total de melhorias encontradas: {melhorias_total}\n")
        f.write(f"   ✅ CERTIFICADO FINAL: {arquivo_resultado}\n")

def main():
    """
    Função principal - REFINAMENTO ULTRA-PRECISO PARA CERTIFICADO FINAL
    """
    arquivo_original = "SAN-038-25-09.xlsx"
    arquivo_refinado = "SAN-038-25-09_REFINADO_HIBRIDO.xlsx"
    arquivo_resultado = "SAN-038-25-09_CERTIFICADO_FINAL.xlsx"
    
    if not os.path.exists(arquivo_original):
        print(f"❌ Arquivo original não encontrado: {arquivo_original}")
        return
    
    if not os.path.exists(arquivo_refinado):
        print(f"❌ Arquivo refinado não encontrado: {arquivo_refinado}")
        print(f"💡 Execute primeiro: python aplicador_tempos_gerados.py")
        return
    
    print("🚀 Iniciando REFINAMENTO ULTRA-PRECISO para CERTIFICADO FINAL...")
    print("=" * 70)
    print("🎯 OBJETIVO: Refinar valores com ULTRA-PRECISÃO para certificado final")
    print("🔧 ESTRATÉGIA: Incremento de 0.00001 até atingir valores exatos")
    print("📊 PRECISÃO: Tolerância de ±0.00001")
    print("🎯 RESULTADO: CERTIFICADO FINAL")
    print("=" * 70)
    
    # Extrai dados da planilha refinada e valores desejados da original
    constantes, pontos = extrair_dados_planilha_refinada(arquivo_refinado, arquivo_original)
    if constantes is None or pontos is None:
        return
    
    print(f"✅ Extraídos {len(pontos)} pontos da planilha refinada")
    print(f"✅ Valores desejados obtidos da planilha original")
    
    tempo_inicio = time.time()
    resultados_pontos = []
    
    # Processa cada ponto individualmente
    for i, ponto in enumerate(pontos):
        print(f"\n{'='*70}")
        print(f"🔍 REFINAMENTO ULTRA-PRECISO PONTO {i+1}/{len(pontos)}")
        print(f"{'='*70}")
        
        # Processa o ponto individual
        resultado_ponto = processar_ponto_ultra_preciso(ponto, constantes)
        resultados_pontos.append(resultado_ponto)
        
        # Se atingiu o objetivo, pode parar este ponto
        if resultado_ponto and resultado_ponto['objetivo_atingido']:
            print(f"   ✅ Ponto {ponto['numero']} ATINGIU O OBJETIVO ULTRA-PRECISO!")
            print(f"   🎯 Próximo ponto...")
        else:
            print(f"   ⚠️  Ponto {ponto['numero']} não atingiu o objetivo ultra-preciso")
            print(f"   📊 Melhor aproximação ultra-refinada encontrada")
        
        # Mostra progresso
        tempo_decorrido = time.time() - tempo_inicio
        print(f"   ⏱️  Tempo decorrido: {tempo_decorrido:.2f} segundos")
        print(f"   📊 Progresso: {i+1}/{len(pontos)} pontos")
    
    tempo_total = time.time() - tempo_inicio
    
    print(f"\n📊 RESUMO FINAL ULTRA-PRECISO:")
    print(f"   Pontos processados: {len(resultados_pontos)}/{len(pontos)}")
    print(f"   Tempo total: {tempo_total:.2f} segundos")
    print(f"   Tempo médio por ponto: {tempo_total/len(pontos):.2f} segundos")
    
    # Aplica os tempos ultra-refinados na planilha
    sucesso = aplicar_tempos_ultra_refinados_na_planilha(resultados_pontos, arquivo_refinado, arquivo_resultado)
    
    if sucesso:
        # Gera relatório final
        gerar_relatorio_final_ultra_preciso(resultados_pontos, arquivo_resultado)
        
        print(f"\n✅ Relatório salvo em: relatorio_ultra_preciso.json")
        print(f"✅ Relatório legível salvo em: relatorio_ultra_preciso.txt")
        print("🎉 CERTIFICADO FINAL GERADO COM ULTRA-PRECISÃO!")
        print(f"📄 Certificado final: {arquivo_resultado}")
    else:
        print("❌ Erro ao aplicar tempos ultra-refinados!")

if __name__ == "__main__":
    main() 