# simulador_ajuste_certificado.py

import pandas as pd
from decimal import Decimal, getcontext
from openpyxl import load_workbook

getcontext().prec = 28  # precisão máxima

ARQUIVO = "SAN-038-25-09-1.xlsx"
SAIDA = "relatorio_simulacao.txt"

# Simula a função DEF.NÚM.DEC(valor, casas)
def def_num_dec(valor: Decimal, casas: int) -> Decimal:
    fator = Decimal('1e-' + str(casas))
    return valor.quantize(fator)

def converter_para_decimal(valor):
    """Converte valor para Decimal, lidando com strings que usam vírgula como separador"""
    if valor is None:
        return Decimal('0')
    if isinstance(valor, str):
        # Remove espaços, remove pontos de milhares e substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    return Decimal(str(valor))

def ler_valores_reais(sheet, linhas, colunas):
    """Lê os valores reais (não formatados) da planilha Excel"""
    valores = {}
    for col in colunas:
        col_letra = chr(65 + col)  # A = 0
        for i, linha in enumerate(linhas):
            key = f"{col_letra}{linha}"
            valor = sheet.cell(row=linha, column=col + 1).value
            valores[key] = converter_para_decimal(valor)
    return valores

def calcular_media(valores, col_letra, linhas):
    return sum(valores[f"{col_letra}{l}"] for l in linhas) / len(linhas)

def main():
    wb = load_workbook(ARQUIVO, data_only=True)
    coleta = wb["Coleta de Dados"]
    cert = wb["Emissão do Certificado"]
    incerteza = wb["Estimativa da Incerteza"]

    linhas = [54, 55, 56]
    # Incluindo a coluna I (8) que estava faltando
    colunas = [2, 5, 8, 14, 17]  # C, F, I, O, R

    dados_orig = ler_valores_reais(coleta, linhas, colunas)

    # Extrai casas decimais da célula BQ10 (formato do arredondamento no certificado)
    casas_cert = int(incerteza["BQ10"].value)

    # Cálculo das médias originais (como no certificado)
    media_I = calcular_media(dados_orig, "I", linhas)  # Volume Corrigido
    media_O = calcular_media(dados_orig, "O", linhas)  # Leitura Medidor
    media_F = calcular_media(dados_orig, "F", linhas)  # Tempo de coleta

    # Arredondar como Excel
    media_vazao_padrao = def_num_dec(media_I, casas_cert)
    media_vazao_medidor = def_num_dec(media_O, casas_cert)
    tendencia = def_num_dec(media_I, casas_cert)

    # Novo tempo base (mesma parte inteira)
    tempo_base = int(media_F)
    tempo_novo = Decimal(str(tempo_base))

    # Correção proporcional
    dados_novos = {}
    for linha in linhas:
        fator = tempo_novo / dados_orig[f"F{linha}"]
        dados_novos[f"F{linha}"] = tempo_novo
        dados_novos[f"C{linha}"] = dados_orig[f"C{linha}"] * fator
        dados_novos[f"I{linha}"] = dados_orig[f"I{linha}"] * fator  # Corrigindo para usar I
        dados_novos[f"O{linha}"] = dados_orig[f"O{linha}"] * fator
        dados_novos[f"R{linha}"] = dados_orig[f"R{linha}"] * fator

    # Recalcular médias com dados novos
    nova_media_I = calcular_media(dados_novos, "I", linhas)  # Corrigindo para usar I
    nova_media_O = calcular_media(dados_novos, "O", linhas)
    nova_media_F = calcular_media(dados_novos, "F", linhas)

    nova_vazao_padrao = def_num_dec(nova_media_I, casas_cert)
    nova_vazao_medidor = def_num_dec(nova_media_O, casas_cert)
    nova_tendencia = def_num_dec(nova_media_I, casas_cert)

    # Dados do certificado real (linha 74 = index 73)
    cert_row = 74
    cert_valores = {
        "vazao_medidor": converter_para_decimal(cert.cell(row=cert_row, column=15).value),
        "vazao_padrao": converter_para_decimal(cert.cell(row=cert_row, column=6).value),
        "tendencia": converter_para_decimal(cert.cell(row=cert_row, column=9).value)
    }

    # Gera relatório
    with open(SAIDA, "w", encoding="utf-8") as f:
        f.write("=== SIMULAÇÃO DE AJUSTE DE TEMPOS E PARÂMETROS ===\n\n")

        f.write("📌 Médias originais:\n")
        f.write(f"Vazão Medidor:    {media_vazao_medidor}\n")
        f.write(f"Vazão Padrão:     {media_vazao_padrao}\n")
        f.write(f"Tendência:        {tendencia}\n\n")

        f.write("📌 Médias simuladas após ajustes:\n")
        f.write(f"Nova Vazão Medidor: {nova_vazao_medidor}\n")
        f.write(f"Nova Vazão Padrão:  {nova_vazao_padrao}\n")
        f.write(f"Nova Tendência:     {nova_tendencia}\n\n")

        f.write("📌 Valores do Certificado (linha 74):\n")
        f.write(f"Vazão Medidor:    {cert_valores['vazao_medidor']}\n")
        f.write(f"Vazão Padrão:     {cert_valores['vazao_padrao']}\n")
        f.write(f"Tendência:        {cert_valores['tendencia']}\n\n")

        f.write("📌 Diferenças entre simulado e certificado:\n")
        f.write(f"Δ Vazão Medidor:  {nova_vazao_medidor - cert_valores['vazao_medidor']}\n")
        f.write(f"Δ Vazão Padrão:   {nova_vazao_padrao - cert_valores['vazao_padrao']}\n")
        f.write(f"Δ Tendência:      {nova_tendencia - cert_valores['tendencia']}\n\n")

        f.write("⚠️  Conclusão:\n")
        if (nova_vazao_medidor == cert_valores['vazao_medidor'] and
            nova_vazao_padrao == cert_valores['vazao_padrao'] and
            nova_tendencia == cert_valores['tendencia']):
            f.write("✅ As médias simuladas são EXATAMENTE IGUAIS às do certificado.\n")
        else:
            f.write("❌ As médias simuladas NÃO são iguais às do certificado. Verificar ajustes.\n")

    print(f"✅ Simulação concluída! Verifique o relatório em '{SAIDA}'")

if __name__ == "__main__":
    main()
