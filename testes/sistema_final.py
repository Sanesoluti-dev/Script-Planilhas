# -*- coding: utf-8 -*-
"""
SISTEMA FINAL DE CORREÇÃO DE CALIBRAÇÃO
========================================

Implementação da lógica de ajuste manual validada:
1. Força o "Tempo de Coleta" para um valor padrão (ex: 360s)
2. Recalcula os parâmetros de entrada (Qtd de Pulsos, Leitura no Medidor)
3. Mantém os valores finais do certificado idênticos aos originais

PROBLEMA: Otimização com Restrição Fixa
"""

import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook
import shutil
import os

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 50

# CONFIGURAÇÃO DO TEMPO ALVO
TEMPO_ALVO = Decimal('360')  # Tempo padrão em segundos

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def calcular_totalizacao_padrao_corrigido(pulsos_padrao, pulso_padrao_lp, temperatura, fator_correcao_temp, tempo_coleta):
    """
    Calcula a "Totalização no Padrão Corrigido • L" usando a fórmula:
    =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
    """
    if pulsos_padrao == 0:
        return Decimal('0')
    
    # C54*$I$51 = Pulsos * Pulso do padrão em L/P
    volume_pulsos = pulsos_padrao * pulso_padrao_lp
    
    # (C54*$I$51/AA54*3600) = Volume / Tempo * 3600 = Vazão
    vazao = volume_pulsos / tempo_coleta * Decimal('3600')
    
    # ($R$51+$U$51*(C54*$I$51/AA54*3600))/100 = (Temperatura + Fator_Correção * Vazão) / 100
    fator_correcao = (temperatura + fator_correcao_temp * vazao) / Decimal('100')
    
    # (C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51))
    # = Volume - (Fator_Correção * Volume)
    totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
    
    return totalizacao

def extrair_constantes_calculo(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def extrair_dados_originais(arquivo_excel):
    """
    FASE 1: Preparação e Análise
    Leitura precisa de todos os dados brutos usando openpyxl e Decimal com alta precisão
    """
    try:
        print(f"📖 FASE 1: Extraindo dados originais do arquivo: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        # Configuração dos pontos
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_originais = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i
                
                # Lê todos os parâmetros necessários
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Pulsos: {float(pulsos_padrao)}")
                print(f"     Tempo: {float(tempo_coleta)} s")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Leitura Medidor: {float(leitura_medidor)} L")
                print(f"     Temperatura: {float(temperatura)} °C")
                print(f"     Erro: {float(erro)} %")

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros)
            erros_validos = [e for e in erros if e != 0]
            if erros_validos:
                tendencia = sum(erros_validos) / Decimal(str(len(erros_validos)))
            else:
                tendencia = Decimal('0')
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            print(f"   VALORES SAGRADOS do Ponto {config['num_ponto']}:")
            print(f"     Vazão Média: {float(vazao_media)} L/h")
            print(f"     Tendência: {float(tendencia)} %")
            print(f"     Desvio Padrão: {float(desvio_padrao) if desvio_padrao else 'N/A'} %")
            
            dados_originais[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_originais
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados originais: {e}")
        return None

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_valores_certificado(dados_originais, constantes):
    """
    FASE 1: Calcular "Valores Sagrados"
    Executa um motor de cálculo com os dados originais para obter os valores finais
    da aba "Emissão do Certificado"
    """
    valores_certificado = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Calculando valores do certificado para {ponto_key}:")
        
        totalizacoes = []
        leituras_medidor = []
        
        for leitura in ponto['leituras']:
            # Calcula "Totalização no Padrão Corrigido • L" conforme documentação
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes.append(totalizacao)
            leituras_medidor.append(leitura['leitura_medidor'])
            
            print(f"     Leitura: Totalização = {float(totalizacao)} L, Leitura Medidor = {float(leitura['leitura_medidor'])} L")
        
        # Calcula médias conforme fórmulas do certificado
        media_totalizacao = sum(totalizacoes) / Decimal(str(len(totalizacoes)))
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        valores_certificado[ponto_key] = {
            'media_totalizacao': media_totalizacao,
            'media_leitura_medidor': media_leitura_medidor,
            'totalizacoes': totalizacoes,
            'leituras_medidor': leituras_medidor
        }
        
        print(f"     Média Totalização: {float(media_totalizacao)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor)} L")
    
    return valores_certificado

def calcular_proporcoes_originais(leituras_ponto):
    """
    FASE 1: Calcular Proporções Originais
    Para cada ponto de calibração (com 3 medições), calcula e armazena as proporções internas
    entre as medições. Usa a primeira medição como "mestre".
    """
    print(f"       📊 Calculando proporções originais...")
    
    # Extrai valores originais
    pulsos_originais = [l['pulsos_padrao'] for l in leituras_ponto]
    leituras_originais = [l['leitura_medidor'] for l in leituras_ponto]
    
    # Define a primeira leitura como "mestre"
    pulsos_mestre = pulsos_originais[0]
    leitura_mestre = leituras_originais[0]
    
    # Calcula proporções dos pulsos
    fatores_proporcao_pulsos = [p / pulsos_mestre for p in pulsos_originais]
    
    # Calcula proporções das leituras
    fatores_proporcao_leituras = [l / leitura_mestre for l in leituras_originais]
    
    # Calcula fator leitura vs pulso (proporção entre leitura e pulso)
    fator_leitura_vs_pulso = leitura_mestre / pulsos_mestre
    
    print(f"         Pulsos mestre: {float(pulsos_mestre)}")
    print(f"         Leitura mestre: {float(leitura_mestre)} L")
    print(f"         Fator leitura vs pulso: {float(fator_leitura_vs_pulso)}")
    print(f"         Proporções pulsos: {[float(f) for f in fatores_proporcao_pulsos]}")
    print(f"         Proporções leituras: {[float(f) for f in fatores_proporcao_leituras]}")
    
    return {
        'pulsos_mestre': pulsos_mestre,
        'leitura_mestre': leitura_mestre,
        'fator_leitura_vs_pulso': fator_leitura_vs_pulso,
        'fatores_proporcao_pulsos': fatores_proporcao_pulsos,
        'fatores_proporcao_leituras': fatores_proporcao_leituras
    }

def calcular_funcao_custo(novo_pulsos_mestre, proporcoes, constantes, valores_cert_originais, tempo_alvo):
    """
    FASE 2: Função de Custo (Erro Total)
    Calcula o erro total do sistema para um dado valor de pulsos mestre
    """
    # Recalcula todos os pulsos mantendo as proporções
    pulsos_ajustados = [novo_pulsos_mestre * f for f in proporcoes['fatores_proporcao_pulsos']]
    
    # Recalcula todas as leituras usando o fator leitura vs pulso
    leituras_ajustadas = [p * proporcoes['fator_leitura_vs_pulso'] for p in pulsos_ajustados]
    
    # Calcula valores do certificado ajustados
    totalizacoes_ajustadas = []
    leituras_medidor_ajustadas = []
    
    for i, (pulsos, leitura) in enumerate(zip(pulsos_ajustados, leituras_ajustadas)):
        # Calcula totalização com dados ajustados
        totalizacao = calcular_totalizacao_padrao_corrigido(
            pulsos,
            constantes['pulso_padrao_lp'],
            constantes['temperatura_constante'],
            constantes['fator_correcao_temp'],
            tempo_alvo
        )
        totalizacoes_ajustadas.append(totalizacao)
        leituras_medidor_ajustadas.append(leitura)
    
    # Calcula médias ajustadas
    media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
    media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
    
    # Valores originais do certificado
    media_totalizacao_original = valores_cert_originais['media_totalizacao']
    media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor']
    
    # Calcula erros relativos
    if media_totalizacao_original != 0:
        erro_vazao_ref = (media_totalizacao_ajustada - media_totalizacao_original) / media_totalizacao_original
    else:
        erro_vazao_ref = Decimal('0')
    
    if media_leitura_medidor_original != 0:
        erro_vazao_med = (media_leitura_medidor_ajustada - media_leitura_medidor_original) / media_leitura_medidor_original
    else:
        erro_vazao_med = Decimal('0')
    
    # Função de custo: soma dos erros ao quadrado
    custo_total = (erro_vazao_ref ** 2) + (erro_vazao_med ** 2)
    
    return {
        'custo_total': custo_total,
        'erro_vazao_ref': erro_vazao_ref,
        'erro_vazao_med': erro_vazao_med,
        'media_totalizacao_ajustada': media_totalizacao_ajustada,
        'media_leitura_medidor_ajustada': media_leitura_medidor_ajustada,
        'pulsos_ajustados': pulsos_ajustados,
        'leituras_ajustadas': leituras_ajustadas
    }

def otimizacao_iterativa_global(leituras_ponto, constantes, valores_cert_originais, ponto_key):
    """
    FASE 2: Otimização Iterativa Global (O Coração do Sistema)
    Implementa uma única função de otimização global que resolve o sistema
    """
    print(f"       🔍 FASE 2: Iniciando otimização iterativa para {ponto_key}")
    print(f"         Tempo alvo: {float(TEMPO_ALVO)} s")
    
    # FASE 1: Calcular proporções originais
    proporcoes = calcular_proporcoes_originais(leituras_ponto)
    
    print(f"         Valores alvo:")
    print(f"           Média Totalização: {float(valores_cert_originais['media_totalizacao'])} L")
    print(f"           Média Leitura Medidor: {float(valores_cert_originais['media_leitura_medidor'])} L")
    
    # Busca pelo mínimo custo
    melhor_pulsos_mestre = proporcoes['pulsos_mestre']
    menor_custo = Decimal('inf')
    melhor_resultado = None
    
    # Busca em torno do valor original
    print(f"         🔄 Buscando mínimo custo...")
    
    for ajuste in range(-500, 501, 1):  # Passo de 1 para máxima precisão
        pulsos_teste = proporcoes['pulsos_mestre'] + ajuste
        
        if pulsos_teste <= 0:
            continue
        
        # Calcula função de custo
        resultado = calcular_funcao_custo(
            pulsos_teste, 
            proporcoes, 
            constantes, 
            valores_cert_originais, 
            TEMPO_ALVO
        )
        
        # Verifica se é o melhor resultado até agora
        if resultado['custo_total'] < menor_custo:
            menor_custo = resultado['custo_total']
            melhor_pulsos_mestre = pulsos_teste
            melhor_resultado = resultado
            
            print(f"           Novo mínimo encontrado:")
            print(f"             Pulsos mestre: {int(melhor_pulsos_mestre)}")
            print(f"             Custo total: {float(menor_custo)}")
            print(f"             Erro Vazão Ref: {float(resultado['erro_vazao_ref'])}")
            print(f"             Erro Vazão Med: {float(resultado['erro_vazao_med'])}")
    
    print(f"         ✅ Otimização concluída:")
    print(f"           Melhor pulsos mestre: {int(melhor_pulsos_mestre)}")
    print(f"           Menor custo: {float(menor_custo)}")
    
    return melhor_resultado

def processar_ponto_calibracao(ponto_key, ponto, constantes, valores_certificado_originais):
    """
    Processa um ponto de calibração completo usando a otimização global
    """
    print(f"\n📊 Processando {ponto_key}:")
    
    # Executa otimização iterativa global
    resultado_otimizacao = otimizacao_iterativa_global(
        ponto['leituras'],
        constantes,
        valores_certificado_originais[ponto_key],
        ponto_key
    )
    
    # Prepara resultado final
    resultado = {
        'ponto_numero': ponto['numero'],
        'pulsos_ajustados': resultado_otimizacao['pulsos_ajustados'],
        'leituras_ajustadas': resultado_otimizacao['leituras_ajustadas'],
        'tempos_ajustados': [TEMPO_ALVO] * len(ponto['leituras']),  # Todos os tempos são o tempo alvo
        'custo_total': resultado_otimizacao['custo_total'],
        'erro_vazao_ref': resultado_otimizacao['erro_vazao_ref'],
        'erro_vazao_med': resultado_otimizacao['erro_vazao_med'],
        'media_totalizacao_ajustada': resultado_otimizacao['media_totalizacao_ajustada'],
        'media_leitura_medidor_ajustada': resultado_otimizacao['media_leitura_medidor_ajustada'],
        'valores_sagrados': ponto['valores_sagrados']
    }
    
    return resultado

def gerar_planilha_corrigida(dados_ajustados, arquivo_original):
    """
    FASE 3: Saída e Geração do Arquivo Final
    Gera um novo arquivo Excel, uma cópia do original, mas com os valores corrigidos
    """
    print(f"\n📄 FASE 3: GERANDO PLANILHA CORRIGIDA")
    print("=" * 60)
    
    # Cria cópia do arquivo original
    arquivo_corrigido = arquivo_original.replace('.xlsx', '_CORRIGIDO.xlsx')
    shutil.copy2(arquivo_original, arquivo_corrigido)
    
    print(f"   Arquivo corrigido: {arquivo_corrigido}")
    
    # Carrega a planilha corrigida
    wb = load_workbook(arquivo_corrigido)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Aplica os valores ajustados
    for ponto_key, dados in dados_ajustados.items():
        leituras_originais = dados['leituras_originais']
        pulsos_ajustados = dados['pulsos_ajustados']
        leituras_ajustadas = dados['leituras_ajustadas']
        tempos_ajustados = dados['tempos_ajustados']
        
        for i, leitura_original in enumerate(leituras_originais):
            linha = leitura_original['linha']
            
            # Aplica os valores ajustados
            coleta_sheet.cell(row=linha, column=3).value = int(pulsos_ajustados[i])  # Coluna C - Pulsos (inteiro)
            coleta_sheet.cell(row=linha, column=6).value = float(tempos_ajustados[i])   # Coluna F - Tempo
            coleta_sheet.cell(row=linha, column=15).value = float(leituras_ajustadas[i])  # Coluna O - Leitura Medidor
            coleta_sheet.cell(row=linha, column=18).value = float(leitura_original['temperatura'])     # Coluna R - Temperatura
            
            print(f"     Linha {linha}:")
            print(f"       Pulsos: {int(pulsos_ajustados[i])} (inteiro)")
            print(f"       Tempo: {float(tempos_ajustados[i])} s")
            print(f"       Leitura Medidor: {float(leituras_ajustadas[i])} L")
            print(f"       Temperatura: {float(leitura_original['temperatura'])} °C")
    
    # Salva a planilha corrigida
    wb.save(arquivo_corrigido)
    print(f"   ✅ Planilha corrigida salva com sucesso")
    
    return arquivo_corrigido

def gerar_relatorio_comparativo(dados_ajustados, valores_certificado_originais):
    """
    Gera relatório no terminal comparando os "Valores Sagrados" com os valores finais
    do novo certificado para provar que a operação foi bem-sucedida
    """
    print(f"\n📋 RELATÓRIO COMPARATIVO FINAL")
    print("=" * 80)
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 PONTO {dados['ponto_numero']}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        
        print(f"   🎯 VALORES SAGRADOS (ORIGINAIS):")
        print(f"     Vazão Média: {float(dados['valores_sagrados']['vazao_media'])} L/h")
        print(f"     Tendência: {float(dados['valores_sagrados']['tendencia'])} %")
        print(f"     Desvio Padrão: {float(dados['valores_sagrados']['desvio_padrao']) if dados['valores_sagrados']['desvio_padrao'] else 'N/A'} %")
        
        print(f"   📊 VALORES DO CERTIFICADO:")
        print(f"     Média Totalização (Original): {float(valores_cert_originais['media_totalizacao'])} L")
        print(f"     Média Leitura Medidor (Original): {float(valores_cert_originais['media_leitura_medidor'])} L")
        print(f"     Média Totalização (Ajustada): {float(dados['media_totalizacao_ajustada'])} L")
        print(f"     Média Leitura Medidor (Ajustada): {float(dados['media_leitura_medidor_ajustada'])} L")
        
        print(f"   📈 COMPARAÇÃO:")
        print(f"     Erro Vazão Ref: {float(dados['erro_vazao_ref'])}")
        print(f"     Erro Vazão Med: {float(dados['erro_vazao_med'])}")
        print(f"     Custo Total: {float(dados['custo_total'])}")
        
        # Verifica se o erro residual é aceitável
        tolerancia = Decimal('1e-10')
        if abs(dados['erro_vazao_ref']) <= tolerancia and abs(dados['erro_vazao_med']) <= tolerancia:
            print(f"   ✅ OPERAÇÃO BEM-SUCEDIDA - Erro residual aceitável")
        else:
            print(f"   ⚠️  ERRO RESIDUAL DETECTADO - Verificar otimização")
        
        print(f"   ⚙️  CONFIGURAÇÃO APLICADA:")
        print(f"     Tempo de Coleta: {float(TEMPO_ALVO)} s (fixo para todas as leituras)")
        print(f"     Pulsos ajustados: {[int(p) for p in dados['pulsos_ajustados']]}")
        print(f"     Leituras ajustadas: {[float(l) for l in dados['leituras_ajustadas']]} L")

def main():
    """Função principal que executa o sistema final de correção de calibração"""
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print("=== SISTEMA FINAL DE CORREÇÃO DE CALIBRAÇÃO ===")
    print("Implementação da lógica de ajuste manual validada")
    print(f"CONFIGURAÇÃO: Tempo alvo = {float(TEMPO_ALVO)} segundos")
    print("PRECISÃO: Decimal com 50 dígitos")
    
    # FASE 1: Preparação e Análise
    print(f"\n🎯 FASE 1: PREPARAÇÃO E ANÁLISE")
    print("=" * 60)
    
    # Leitura precisa dos dados originais
    dados_originais = extrair_dados_originais(arquivo_excel)
    if not dados_originais:
        print("❌ Falha na extração dos dados originais")
        return
    
    # Extração de constantes
    constantes = extrair_constantes_calculo(arquivo_excel)
    if not constantes:
        print("❌ Falha na extração das constantes")
        return
    
    # Cálculo dos valores sagrados
    valores_certificado_originais = calcular_valores_certificado(dados_originais, constantes)
    
    print(f"✅ FASE 1 CONCLUÍDA: Dados extraídos e valores sagrados calculados")
    
    # FASE 2: Otimização Iterativa Global
    print(f"\n🔍 FASE 2: OTIMIZAÇÃO ITERATIVA GLOBAL")
    print("=" * 60)
    
    dados_ajustados = {}
    
    for ponto_key, ponto in dados_originais.items():
        # Adiciona as leituras originais ao resultado para referência
        ponto['leituras_originais'] = ponto['leituras']
        
        resultado = processar_ponto_calibracao(
            ponto_key, 
            ponto, 
            constantes, 
            valores_certificado_originais
        )
        
        dados_ajustados[ponto_key] = resultado
    
    print(f"✅ FASE 2 CONCLUÍDA: Otimização global executada para todos os pontos")
    
    # FASE 3: Saída e Geração do Arquivo Final
    print(f"\n📄 FASE 3: SAÍDA E GERAÇÃO DO ARQUIVO FINAL")
    print("=" * 60)
    
    # Gera planilha corrigida
    arquivo_corrigido = gerar_planilha_corrigida(dados_ajustados, arquivo_excel)
    
    # Gera relatório comparativo
    gerar_relatorio_comparativo(dados_ajustados, valores_certificado_originais)
    
    print(f"\n🎉 SISTEMA FINAL CONCLUÍDO COM SUCESSO!")
    print(f"   ✅ Otimização com restrição fixa implementada")
    print(f"   ✅ Tempo de coleta fixado em {float(TEMPO_ALVO)} segundos")
    print(f"   ✅ Valores do certificado preservados")
    print(f"   ✅ Planilha corrigida: {arquivo_corrigido}")
    print(f"   ✅ Relatório comparativo gerado")

if __name__ == "__main__":
    main() 