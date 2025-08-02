#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para ler os valores reais do certificado da planilha corrigida
Lê diretamente os valores calculados pela planilha Excel
"""

import os
import time
from decimal import Decimal
from openpyxl import load_workbook
import json
from datetime import datetime

def ler_valores_certificado_planilha():
    """
    Lê os valores reais do certificado da planilha corrigida
    """
    arquivo_corrigido = "correto/SAN-038-25-09_CORRIGIDO.xlsx"
    
    print("📊 LEITOR DE VALORES DO CERTIFICADO")
    print("=" * 60)
    
    # Verifica se o arquivo existe
    if not os.path.exists(arquivo_corrigido):
        print(f"❌ Arquivo não encontrado: {arquivo_corrigido}")
        return None
    
    print(f"✅ Arquivo encontrado: {arquivo_corrigido}")
    
    try:
        # Carrega a planilha com data_only=True para ler valores calculados
        wb = load_workbook(arquivo_corrigido, data_only=True)
        
        # Lista todas as planilhas disponíveis
        print(f"\n📋 PLANILHAS DISPONÍVEIS:")
        for sheet_name in wb.sheetnames:
            print(f"   • {sheet_name}")
        
        # Procura pela planilha que contém os dados do certificado
        sheet_certificado = None
        for sheet_name in wb.sheetnames:
            if "certificado" in sheet_name.lower() or "resultado" in sheet_name.lower():
                sheet_certificado = wb[sheet_name]
                print(f"\n✅ Planilha do certificado encontrada: {sheet_name}")
                break
        
        if not sheet_certificado:
            # Se não encontrar planilha específica, usa a primeira
            sheet_certificado = wb.active
            print(f"\n⚠️  Usando planilha ativa: {sheet_certificado.title}")
        
        print(f"   Dimensões: {sheet_certificado.dimensions}")
        
        # Busca pelos valores do certificado
        valores_certificado = {}
        
        print(f"\n🔍 BUSCANDO VALORES DO CERTIFICADO:")
        
        # Para cada ponto de calibração (1 a 8)
        for ponto_num in range(1, 9):
            print(f"\n   📊 PONTO {ponto_num}:")
            
            # Determina as linhas base para cada ponto
            # Ponto 1: linhas 54-56, Ponto 2: linhas 59-61, etc.
            linha_base = 54 + (ponto_num - 1) * 5
            
            # Busca por valores em diferentes colunas e linhas
            valores_encontrados = {}
            
            # Busca por vazão média (valores tipicamente > 1000 L/h)
            vazao_encontrada = None
            for col in range(1, 30):  # Busca em todas as colunas
                for linha_offset in range(-10, 20):  # Busca em range maior
                    try:
                        cell = sheet_certificado.cell(row=linha_base + linha_offset, column=col)
                        if cell.value and isinstance(cell.value, (int, float)) and cell.value > 1000 and cell.value < 1000000:
                            # Verifica se é um valor de vazão (entre 1000 e 1000000 L/h)
                            vazao_encontrada = {
                                'valor': cell.value,
                                'coordenada': cell.coordinate
                            }
                            print(f"     ✅ Vazão média: {cell.value} L/h em {cell.coordinate}")
                            break
                    except:
                        continue
                if vazao_encontrada:
                    break
            
            # Busca por tendência (valores tipicamente pequenos, entre -10 e +10)
            tendencia_encontrada = None
            for col in range(1, 30):
                for linha_offset in range(-10, 20):
                    try:
                        cell = sheet_certificado.cell(row=linha_base + linha_offset, column=col)
                        if cell.value and isinstance(cell.value, (int, float)) and abs(cell.value) < 10 and cell.value != 0:
                            # Verifica se é um valor de tendência
                            tendencia_encontrada = {
                                'valor': cell.value,
                                'coordenada': cell.coordinate
                            }
                            print(f"     ✅ Tendência: {cell.value} % em {cell.coordinate}")
                            break
                    except:
                        continue
                if tendencia_encontrada:
                    break
            
            # Busca por desvio padrão (valores tipicamente pequenos e positivos)
            desvio_encontrado = None
            for col in range(1, 30):
                for linha_offset in range(-10, 20):
                    try:
                        cell = sheet_certificado.cell(row=linha_base + linha_offset, column=col)
                        if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0 and cell.value < 5:
                            # Verifica se é um valor de desvio padrão
                            desvio_encontrado = {
                                'valor': cell.value,
                                'coordenada': cell.coordinate
                            }
                            print(f"     ✅ Desvio padrão: {cell.value} % em {cell.coordinate}")
                            break
                    except:
                        continue
                if desvio_encontrado:
                    break
            
            # Armazena os valores encontrados
            valores_certificado[f"ponto_{ponto_num}"] = {
                'numero_ponto': ponto_num,
                'vazao_media': vazao_encontrada,
                'tendencia': tendencia_encontrada,
                'desvio_padrao': desvio_encontrado
            }
            
            # Mostra resumo do ponto
            valores_validos = sum(1 for v in [vazao_encontrada, tendencia_encontrada, desvio_encontrado] if v is not None)
            print(f"     📊 Resumo: {valores_validos}/3 valores encontrados")
        
        wb.close()
        
        # Gera relatório dos valores encontrados
        print(f"\n📋 RELATÓRIO DOS VALORES ENCONTRADOS:")
        print("=" * 60)
        
        for ponto_key, dados in valores_certificado.items():
            print(f"\n   📊 {ponto_key.upper()}:")
            if dados['vazao_media']:
                print(f"     Vazão Média: {dados['vazao_media']['valor']} L/h ({dados['vazao_media']['coordenada']})")
            else:
                print(f"     Vazão Média: ❌ Não encontrada")
            
            if dados['tendencia']:
                print(f"     Tendência: {dados['tendencia']['valor']} % ({dados['tendencia']['coordenada']})")
            else:
                print(f"     Tendência: ❌ Não encontrada")
            
            if dados['desvio_padrao']:
                print(f"     Desvio Padrão: {dados['desvio_padrao']['valor']} % ({dados['desvio_padrao']['coordenada']})")
            else:
                print(f"     Desvio Padrão: ❌ Não encontrado")
        
        # Salva os valores em JSON
        resultado = {
            "metadata": {
                "data_geracao": datetime.now().isoformat(),
                "descricao": "Valores reais do certificado lidos da planilha corrigida",
                "arquivo_planilha": arquivo_corrigido,
                "total_pontos": len(valores_certificado)
            },
            "valores_certificado": valores_certificado
        }
        
        nome_arquivo = "valores_certificado_reais.json"
        with open(nome_arquivo, "w", encoding="utf-8") as f:
            json.dump(resultado, f, indent=2, ensure_ascii=False)
        
        print(f"\n📄 VALORES SALVOS EM: {nome_arquivo}")
        print(f"✅ Leitura concluída com sucesso")
        
        return valores_certificado
        
    except Exception as e:
        print(f"❌ Erro ao ler planilha: {e}")
        return None

if __name__ == "__main__":
    ler_valores_certificado_planilha() 