# -*- coding: utf-8 -*-
"""
Teste de leitura dos valores sagrados da planilha original
"""

from openpyxl import load_workbook
from decimal import Decimal, getcontext

# Configura precisão máxima
getcontext().prec = 28

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da célula com precisão máxima
    """
    valor = sheet.cell(row=linha, column=coluna).value
    if valor is None:
        return Decimal('0')
    return Decimal(str(valor))

def testar_leitura_valores_sagrados():
    """
    Testa a leitura dos valores sagrados da planilha original
    """
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print(f"📖 Testando leitura de valores sagrados do arquivo: {arquivo_excel}")
    
    wb = load_workbook(arquivo_excel, data_only=True)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Testa os primeiros 3 pontos
    for ponto in range(1, 4):
        linha_inicial = 50 + (ponto - 1) * 9
        linha_agregados = linha_inicial + 3
        
        print(f"\n📊 Ponto {ponto}:")
        print(f"   Linha inicial: {linha_inicial}")
        print(f"   Linha agregados: {linha_agregados}")
        
        # Lê os valores sagrados
        vazao_media = ler_valor_exato(coleta_sheet, linha_agregados, 9)   # Coluna I
        tendencia = ler_valor_exato(coleta_sheet, linha_agregados, 21)     # Coluna U
        desvio_padrao = ler_valor_exato(coleta_sheet, linha_agregados, 30) # Coluna AD
        
        print(f"   Vazão Média (I{linha_agregados}): {float(vazao_media)} L/h")
        print(f"   Tendência (U{linha_agregados}): {float(tendencia)} %")
        print(f"   Desvio Padrão (AD{linha_agregados}): {float(desvio_padrao)} %")
        
        # Testa também as leituras individuais
        for i in range(3):
            linha = linha_inicial + 4 + i
            vazao_ref = ler_valor_exato(coleta_sheet, linha, 9)  # Coluna I
            erro = ler_valor_exato(coleta_sheet, linha, 21)       # Coluna U
            
            print(f"     Leitura {i+1} (linha {linha}):")
            print(f"       Vazão Ref (I{linha}): {float(vazao_ref)} L/h")
            print(f"       Erro (U{linha}): {float(erro)} %")

if __name__ == "__main__":
    testar_leitura_valores_sagrados() 