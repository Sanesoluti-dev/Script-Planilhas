# -*- coding: utf-8 -*-
"""
Otimizador de Tempos de Coleta - Versão Precisão Máxima
Testa milhares de combinações para encontrar tempos que gerem valores exatos
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from decimal import Decimal, ROUND_HALF_UP, getcontext
import json
import os
import itertools
import time

# Configura precisão máxima
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal máxima
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def extrair_dados_planilha_original(arquivo_excel):
    """
    Extrai todos os dados necessários da planilha original
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        estimativa_sheet = wb["Estimativa da Incerteza"]
        
        # Extrai constantes
        constantes = {}
        constantes['ponto_mlp'] = ler_valor_exato(coleta_sheet, 50, 9)  # I50
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_sheet, 50, 30)  # AD50
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_sheet, 51, 18)  # R51
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_sheet, 51, 21)  # U51
        constantes['modo_calibracao'] = coleta_sheet.cell(row=16, column=24).value  # X16
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_sheet, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_sheet, 23, 75)  # BW23
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_sheet, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_sheet, 26, 75)  # BW26
        
        # Extrai dados dos pontos
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': [],
                'valores_originais': {}
            }
            
            # Extrai valores brutos e calculados
            for i in range(3):
                linha = linha_atual + i
                
                # Valores brutos
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)
                
                # Valores calculados (para manter iguais)
                tempo_coleta_corrigido = ler_valor_exato(coleta_sheet, linha, 27)
                temperatura_corrigida = ler_valor_exato(coleta_sheet, linha, 30)
                totalizacao_padrao_corrigido = ler_valor_exato(coleta_sheet, linha, 12)
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)
                vazao_medidor = ler_valor_exato(coleta_sheet, linha, 24)
                erro_percentual = ler_valor_exato(coleta_sheet, linha, 21)
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'tempo_coleta_corrigido': tempo_coleta_corrigido,
                    'temperatura_corrigida': temperatura_corrigida,
                    'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
                    'vazao_referencia': vazao_referencia,
                    'vazao_medidor': vazao_medidor,
                    'erro_percentual': erro_percentual
                }
                
                ponto['leituras'].append(leitura)
            
            # Valores agregados originais
            linha_agregados = linha_atual + 3
            vazao_media_original = ler_valor_exato(coleta_sheet, linha_agregados, 9)
            tendencia_original = ler_valor_exato(coleta_sheet, linha_agregados, 21)
            desvio_padrao_original = ler_valor_exato(coleta_sheet, linha_agregados, 30)
            
            ponto['valores_originais'] = {
                'vazao_media': vazao_media_original,
                'tendencia': tendencia_original,
                'desvio_padrao': desvio_padrao_original
            }
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        return constantes, pontos
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados: {e}")
        return None, None

def calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_ajustados):
    """
    Calcula todas as fórmulas com tempos ajustados usando precisão máxima
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        linha = leitura['linha']
        tempo_ajustado = tempos_ajustados[i]
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = tempo_ajustado
        tempo_coleta_corrigido = tempo_coleta_bruto - (tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23'])
        tempo_coleta_corrigido = tempo_coleta_corrigido.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        temperatura_corrigida = temperatura_bruta - (temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26'])
        temperatura_corrigida = temperatura_corrigida.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        volume_bruto = volume_bruto.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        vazao_bruta = vazao_bruta.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100') * volume_bruto
        correcao = correcao.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        totalizacao_padrao_corrigido = volume_bruto - correcao
        totalizacao_padrao_corrigido = totalizacao_padrao_corrigido.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        vazao_referencia = vazao_referencia.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 7: Vazão do Medidor • L/h (X54)
        leitura_medidor = leitura['leitura_medidor']
        modo_calibracao = constantes['modo_calibracao']
        
        if modo_calibracao in ["Visual com início dinâmico", "Visual com início estática"]:
            vazao_medidor = leitura_medidor
        else:
            vazao_medidor = (leitura_medidor / tempo_coleta_corrigido) * Decimal('3600')
            vazao_medidor = vazao_medidor.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 8: Erro % (U54)
        erro_percentual = (leitura_medidor - totalizacao_padrao_corrigido) / totalizacao_padrao_corrigido * Decimal('100')
        erro_percentual = erro_percentual.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
        
        resultado = {
            'linha': linha,
            'tempo_coleta_ajustado': tempo_ajustado,
            'tempo_coleta_corrigido': tempo_coleta_corrigido,
            'temperatura_corrigida': temperatura_corrigida,
            'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
            'vazao_referencia': vazao_referencia,
            'vazao_medidor': vazao_medidor,
            'erro_percentual': erro_percentual
        }
        
        resultados.append(resultado)
    
    return resultados

def calcular_agregados_com_tempo_ajustado(resultados):
    """
    Calcula os valores agregados com tempos ajustados usando precisão máxima
    """
    vazoes_referencia = [r['vazao_referencia'] for r in resultados]
    erros_percentuais = [r['erro_percentual'] for r in resultados]
    
    # FÓRMULA 9: Vazão Média • L/h (I57)
    vazao_media = sum(vazoes_referencia) / Decimal(str(len(vazoes_referencia)))
    vazao_media = vazao_media.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    
    # FÓRMULA 10: Tendência (U57)
    tendencia = sum(erros_percentuais) / Decimal(str(len(erros_percentuais)))
    tendencia = tendencia.quantize(Decimal('0.000000000000000000000000000'), rounding=ROUND_HALF_UP)
    
    # FÓRMULA 11: DESVIO PADRÃO AMOSTRAL (AD57)
    desvio_padrao = calcular_desvio_padrao_amostral(erros_percentuais)
    
    return {
        'vazao_media': vazao_media,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def gerar_combinacoes_tempos():
    """
    Gera combinações de tempos entre 239.5999 e 240.499
    """
    # Gera valores com incrementos pequenos para maior precisão
    valores_base = []
    
    # De 239.5999 a 240.499 com incrementos de 0.0001
    for i in range(9000):  # 239.5999 a 240.4999
        valor = 239.5999 + (i * 0.0001)
        if valor <= 240.499:
            valores_base.append(Decimal(str(valor)))
    
    return valores_base

def otimizar_tempos_ponto_preciso(leituras, constantes, valores_originais):
    """
    Otimiza os tempos de coleta usando busca exaustiva com precisão máxima
    """
    print(f"   🔍 Iniciando busca exaustiva para Ponto {leituras[0]['linha']}...")
    
    # Gera combinações de tempos
    valores_tempo = gerar_combinacoes_tempos()
    print(f"   📊 Testando {len(valores_tempo)} valores de tempo...")
    
    melhor_combinacao = None
    melhor_diferenca = Decimal('inf')
    contador = 0
    
    # Testa todas as combinações possíveis (3 leituras)
    for tempo1 in valores_tempo:
        for tempo2 in valores_tempo:
            for tempo3 in valores_tempo:
                contador += 1
                
                if contador % 100000 == 0:
                    print(f"   ⏳ Testadas {contador} combinações...")
                
                tempos_teste = [tempo1, tempo2, tempo3]
                
                # Calcula resultados com estes tempos
                resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_teste)
                agregados = calcular_agregados_com_tempo_ajustado(resultados)
                
                # Calcula diferença total
                diff_vazao = abs(agregados['vazao_media'] - valores_originais['vazao_media'])
                diff_tendencia = abs(agregados['tendencia'] - valores_originais['tendencia'])
                
                # Se tem desvio padrão, inclui na comparação
                diff_desvio = Decimal('0')
                if agregados['desvio_padrao'] and valores_originais['desvio_padrao']:
                    diff_desvio = abs(agregados['desvio_padrao'] - valores_originais['desvio_padrao'])
                
                diferenca_total = diff_vazao + diff_tendencia + diff_desvio
                
                # Se encontrou uma combinação melhor
                if diferenca_total < melhor_diferenca:
                    melhor_diferenca = diferenca_total
                    melhor_combinacao = {
                        'tempos': tempos_teste,
                        'agregados': agregados,
                        'diferenca_total': diferenca_total,
                        'diff_vazao': diff_vazao,
                        'diff_tendencia': diff_tendencia,
                        'diff_desvio': diff_desvio
                    }
                    
                    # Se a diferença é muito pequena, considera encontrado
                    if diferenca_total < Decimal('0.0001'):
                        print(f"   ✅ Encontrada combinação com diferença mínima: {float(diferenca_total):.8f}")
                        return melhor_combinacao
    
    print(f"   ✅ Melhor combinação encontrada com diferença: {float(melhor_diferenca):.8f}")
    return melhor_combinacao

def gerar_planilha_otimizada(constantes, pontos_otimizados, arquivo_saida):
    """
    Gera uma nova planilha com os tempos otimizados
    """
    try:
        # Carrega planilha original como template
        wb_original = load_workbook("SAN-038-25-09.xlsx")
        wb_novo = Workbook()
        
        # Copia todas as abas
        for sheet_name in wb_original.sheetnames:
            if sheet_name in wb_novo.sheetnames:
                wb_novo.remove(wb_novo[sheet_name])
            wb_novo.create_sheet(sheet_name)
        
        # Copia dados da planilha original
        for sheet_name in wb_original.sheetnames:
            sheet_original = wb_original[sheet_name]
            sheet_novo = wb_novo[sheet_name]
            
            for row in sheet_original.iter_rows():
                for cell in row:
                    sheet_novo[cell.coordinate] = cell.value
        
        # Atualiza tempos de coleta na aba "Coleta de Dados"
        coleta_sheet = wb_novo["Coleta de Dados"]
        
        for ponto in pontos_otimizados:
            for i, leitura in enumerate(ponto['leituras']):
                linha = leitura['linha']
                tempo_ajustado = ponto['melhor_combinacao']['tempos'][i]
                
                # Atualiza tempo de coleta (coluna F)
                coleta_sheet.cell(row=linha, column=6, value=float(tempo_ajustado))
        
        # Salva nova planilha
        wb_novo.save(arquivo_saida)
        print(f"✅ Planilha otimizada salva como: {arquivo_saida}")
        
    except Exception as e:
        print(f"ERRO ao gerar planilha: {e}")

def main():
    """
    Função principal
    """
    arquivo_original = "SAN-038-25-09.xlsx"
    
    if not os.path.exists(arquivo_original):
        print(f"❌ Arquivo não encontrado: {arquivo_original}")
        return
    
    print("🚀 Iniciando otimização de tempos de coleta - VERSÃO PRECISÃO MÁXIMA...")
    print("=" * 60)
    print("⚠️  ATENÇÃO: Este processo pode demorar várias horas!")
    print("=" * 60)
    
    # Extrai dados da planilha original
    constantes, pontos = extrair_dados_planilha_original(arquivo_original)
    if constantes is None or pontos is None:
        return
    
    print(f"✅ Extraídos {len(pontos)} pontos da planilha original")
    
    # Otimiza cada ponto
    pontos_otimizados = []
    tempo_inicio = time.time()
    
    for i, ponto in enumerate(pontos):
        print(f"\n🔍 Otimizando Ponto {ponto['numero']} ({i+1}/{len(pontos)})...")
        
        # Otimiza tempos
        melhor_combinacao = otimizar_tempos_ponto_preciso(
            ponto['leituras'], 
            constantes, 
            ponto['valores_originais']
        )
        
        # Calcula resultados com tempos otimizados
        resultados_otimizados = calcular_formulas_com_tempo_ajustado(
            ponto['leituras'], 
            constantes, 
            melhor_combinacao['tempos']
        )
        
        # Verifica se os valores estão corretos
        vazao_diff = abs(float(melhor_combinacao['agregados']['vazao_media'] - ponto['valores_originais']['vazao_media']))
        tendencia_diff = abs(float(melhor_combinacao['agregados']['tendencia'] - ponto['valores_originais']['tendencia']))
        
        print(f"   Vazão Média Original: {float(ponto['valores_originais']['vazao_media']):.6f}")
        print(f"   Vazão Média Otimizada: {float(melhor_combinacao['agregados']['vazao_media']):.6f}")
        print(f"   Diferença: {vazao_diff:.8f}")
        print(f"   Tendência Original: {float(ponto['valores_originais']['tendencia']):.6f}")
        print(f"   Tendência Otimizada: {float(melhor_combinacao['agregados']['tendencia']):.6f}")
        print(f"   Diferença: {tendencia_diff:.8f}")
        print(f"   Tempos Otimizados: {[float(t) for t in melhor_combinacao['tempos']]}")
        
        ponto_otimizado = {
            'numero': ponto['numero'],
            'leituras': resultados_otimizados,
            'melhor_combinacao': melhor_combinacao,
            'valores_originais': ponto['valores_originais']
        }
        
        pontos_otimizados.append(ponto_otimizado)
        
        # Mostra tempo decorrido
        tempo_decorrido = time.time() - tempo_inicio
        print(f"   ⏱️  Tempo decorrido: {tempo_decorrido/60:.1f} minutos")
    
    # Gera planilha otimizada
    arquivo_otimizado = "SAN-038-25-09_TEMPOS_OTIMIZADOS_PRECISOS.xlsx"
    gerar_planilha_otimizada(constantes, pontos_otimizados, arquivo_otimizado)
    
    # Salva resultados da otimização
    resultado_otimizacao = {
        'constantes': {k: float(v) if isinstance(v, Decimal) else v for k, v in constantes.items()},
        'pontos_otimizados': []
    }
    
    for ponto in pontos_otimizados:
        ponto_resultado = {
            'numero': ponto['numero'],
            'tempos_otimizados': [float(t) for t in ponto['melhor_combinacao']['tempos']],
            'agregados_otimizados': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto['melhor_combinacao']['agregados'].items()},
            'valores_originais': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto['valores_originais'].items()},
            'diferencas': {
                'total': float(ponto['melhor_combinacao']['diferenca_total']),
                'vazao': float(ponto['melhor_combinacao']['diff_vazao']),
                'tendencia': float(ponto['melhor_combinacao']['diff_tendencia']),
                'desvio': float(ponto['melhor_combinacao']['diff_desvio'])
            }
        }
        resultado_otimizacao['pontos_otimizados'].append(ponto_resultado)
    
    with open('resultados_otimizacao_tempos_precisos.json', 'w', encoding='utf-8') as f:
        json.dump(resultado_otimizacao, f, indent=2, ensure_ascii=False)
    
    tempo_total = time.time() - tempo_inicio
    print(f"\n✅ Resultados salvos em: resultados_otimizacao_tempos_precisos.json")
    print(f"⏱️  Tempo total de processamento: {tempo_total/3600:.2f} horas")
    print("🎉 Otimização de precisão máxima concluída!")

if __name__ == "__main__":
    main() 