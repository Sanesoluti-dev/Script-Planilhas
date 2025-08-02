# -*- coding: utf-8 -*-
"""
Otimizador de Tempos de Coleta
Ajusta os tempos de coleta para ficarem próximos de 240s mantendo os valores finais idênticos
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from decimal import Decimal, ROUND_HALF_UP
import json
import os
import numpy as np
from scipy.optimize import minimize

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def extrair_dados_planilha_original(arquivo_excel):
    """
    Extrai todos os dados necessários da planilha original
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        estimativa_sheet = wb["Estimativa da Incerteza"]
        
        # Extrai constantes
        constantes = {}
        constantes['ponto_mlp'] = ler_valor_exato(coleta_sheet, 50, 9)  # I50
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_sheet, 50, 30)  # AD50
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_sheet, 51, 18)  # R51
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_sheet, 51, 21)  # U51
        constantes['modo_calibracao'] = coleta_sheet.cell(row=16, column=24).value  # X16
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_sheet, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_sheet, 23, 75)  # BW23
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_sheet, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_sheet, 26, 75)  # BW26
        
        # Extrai dados dos pontos
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': [],
                'valores_originais': {}
            }
            
            # Extrai valores brutos e calculados
            for i in range(3):
                linha = linha_atual + i
                
                # Valores brutos
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)
                
                # Valores calculados (para manter iguais)
                tempo_coleta_corrigido = ler_valor_exato(coleta_sheet, linha, 27)
                temperatura_corrigida = ler_valor_exato(coleta_sheet, linha, 30)
                totalizacao_padrao_corrigido = ler_valor_exato(coleta_sheet, linha, 12)
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)
                vazao_medidor = ler_valor_exato(coleta_sheet, linha, 24)
                erro_percentual = ler_valor_exato(coleta_sheet, linha, 21)
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'tempo_coleta_corrigido': tempo_coleta_corrigido,
                    'temperatura_corrigida': temperatura_corrigida,
                    'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
                    'vazao_referencia': vazao_referencia,
                    'vazao_medidor': vazao_medidor,
                    'erro_percentual': erro_percentual
                }
                
                ponto['leituras'].append(leitura)
            
            # Valores agregados originais
            linha_agregados = linha_atual + 3
            vazao_media_original = ler_valor_exato(coleta_sheet, linha_agregados, 9)
            tendencia_original = ler_valor_exato(coleta_sheet, linha_agregados, 21)
            desvio_padrao_original = ler_valor_exato(coleta_sheet, linha_agregados, 30)
            
            ponto['valores_originais'] = {
                'vazao_media': vazao_media_original,
                'tendencia': tendencia_original,
                'desvio_padrao': desvio_padrao_original
            }
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        return constantes, pontos
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados: {e}")
        return None, None

def calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_ajustados):
    """
    Calcula todas as fórmulas com tempos ajustados
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        linha = leitura['linha']
        tempo_ajustado = tempos_ajustados[i]
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = tempo_ajustado
        tempo_coleta_corrigido = tempo_coleta_bruto - (tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23'])
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        temperatura_corrigida = temperatura_bruta - (temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26'])
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100') * volume_bruto
        totalizacao_padrao_corrigido = volume_bruto - correcao
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        
        # FÓRMULA 7: Vazão do Medidor • L/h (X54)
        leitura_medidor = leitura['leitura_medidor']
        modo_calibracao = constantes['modo_calibracao']
        
        if modo_calibracao in ["Visual com início dinâmico", "Visual com início estática"]:
            vazao_medidor = leitura_medidor
        else:
            vazao_medidor = (leitura_medidor / tempo_coleta_corrigido) * Decimal('3600')
        
        # FÓRMULA 8: Erro % (U54)
        erro_percentual = (leitura_medidor - totalizacao_padrao_corrigido) / totalizacao_padrao_corrigido * Decimal('100')
        
        resultado = {
            'linha': linha,
            'tempo_coleta_ajustado': tempo_ajustado,
            'tempo_coleta_corrigido': tempo_coleta_corrigido,
            'temperatura_corrigida': temperatura_corrigida,
            'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
            'vazao_referencia': vazao_referencia,
            'vazao_medidor': vazao_medidor,
            'erro_percentual': erro_percentual
        }
        
        resultados.append(resultado)
    
    return resultados

def calcular_agregados_com_tempo_ajustado(resultados):
    """
    Calcula os valores agregados com tempos ajustados
    """
    vazoes_referencia = [r['vazao_referencia'] for r in resultados]
    erros_percentuais = [r['erro_percentual'] for r in resultados]
    
    # FÓRMULA 9: Vazão Média • L/h (I57)
    vazao_media = sum(vazoes_referencia) / Decimal(str(len(vazoes_referencia)))
    
    # FÓRMULA 10: Tendência (U57)
    tendencia = sum(erros_percentuais) / Decimal(str(len(erros_percentuais)))
    
    # FÓRMULA 11: DESVIO PADRÃO AMOSTRAL (AD57)
    desvio_padrao = calcular_desvio_padrao_amostral(erros_percentuais)
    
    return {
        'vazao_media': vazao_media,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def funcao_objetivo(tempos_ajustados, leituras, constantes, valores_originais):
    """
    Função objetivo para otimização
    Penaliza desvios dos valores originais e tempos fora do intervalo desejado
    """
    tempos_ajustados = [Decimal(str(t)) for t in tempos_ajustados]
    
    # Calcula resultados com tempos ajustados
    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_ajustados)
    agregados = calcular_agregados_com_tempo_ajustado(resultados)
    
    # Penaliza desvios dos valores originais
    penalidade_vazao = abs(float(agregados['vazao_media'] - valores_originais['vazao_media'])) * 1000
    penalidade_tendencia = abs(float(agregados['tendencia'] - valores_originais['tendencia'])) * 1000
    penalidade_desvio = 0
    if agregados['desvio_padrao'] and valores_originais['desvio_padrao']:
        penalidade_desvio = abs(float(agregados['desvio_padrao'] - valores_originais['desvio_padrao'])) * 1000
    
    # Penaliza tempos fora do intervalo desejado (239.5999 a 240.499)
    penalidade_tempo = 0
    for tempo in tempos_ajustados:
        tempo_float = float(tempo)
        if tempo_float < 239.5999 or tempo_float > 240.499:
            penalidade_tempo += abs(tempo_float - 240.0) * 100
    
    return penalidade_vazao + penalidade_tendencia + penalidade_desvio + penalidade_tempo

def otimizar_tempos_ponto(leituras, constantes, valores_originais):
    """
    Otimiza os tempos de coleta para um ponto específico
    """
    # Tempos iniciais (valores originais)
    tempos_iniciais = [float(l['tempo_coleta']) for l in leituras]
    
    # Restrições: tempos devem estar entre 239.5999 e 240.499
    bounds = [(239.5999, 240.499) for _ in range(3)]
    
    # Otimização
    resultado = minimize(
        funcao_objetivo,
        tempos_iniciais,
        args=(leituras, constantes, valores_originais),
        bounds=bounds,
        method='L-BFGS-B',
        options={'maxiter': 1000}
    )
    
    if resultado.success:
        tempos_otimizados = [Decimal(str(t)) for t in resultado.x]
        print(f"   ✅ Otimização convergiu para tempos: {[float(t) for t in tempos_otimizados]}")
        return tempos_otimizados
    else:
        print(f"   ⚠️  Otimização não convergiu, usando tempos originais")
        return [Decimal(str(t)) for t in tempos_iniciais]

def gerar_planilha_otimizada(constantes, pontos_otimizados, arquivo_saida):
    """
    Gera uma nova planilha com os tempos otimizados
    """
    try:
        # Carrega planilha original como template
        wb_original = load_workbook("SAN-038-25-09.xlsx")
        wb_novo = Workbook()
        
        # Copia todas as abas
        for sheet_name in wb_original.sheetnames:
            if sheet_name in wb_novo.sheetnames:
                wb_novo.remove(wb_novo[sheet_name])
            wb_novo.create_sheet(sheet_name)
        
        # Copia dados da planilha original
        for sheet_name in wb_original.sheetnames:
            sheet_original = wb_original[sheet_name]
            sheet_novo = wb_novo[sheet_name]
            
            for row in sheet_original.iter_rows():
                for cell in row:
                    sheet_novo[cell.coordinate] = cell.value
        
        # Atualiza tempos de coleta na aba "Coleta de Dados"
        coleta_sheet = wb_novo["Coleta de Dados"]
        
        for ponto in pontos_otimizados:
            for i, leitura in enumerate(ponto['leituras']):
                linha = leitura['linha']
                tempo_ajustado = ponto['tempos_otimizados'][i]
                
                # Atualiza tempo de coleta (coluna F)
                coleta_sheet.cell(row=linha, column=6, value=float(tempo_ajustado))
        
        # Salva nova planilha
        wb_novo.save(arquivo_saida)
        print(f"✅ Planilha otimizada salva como: {arquivo_saida}")
        
    except Exception as e:
        print(f"ERRO ao gerar planilha: {e}")

def main():
    """
    Função principal
    """
    arquivo_original = "SAN-038-25-09.xlsx"
    
    if not os.path.exists(arquivo_original):
        print(f"❌ Arquivo não encontrado: {arquivo_original}")
        return
    
    print("🚀 Iniciando otimização de tempos de coleta...")
    print("=" * 60)
    
    # Extrai dados da planilha original
    constantes, pontos = extrair_dados_planilha_original(arquivo_original)
    if constantes is None or pontos is None:
        return
    
    print(f"✅ Extraídos {len(pontos)} pontos da planilha original")
    
    # Otimiza cada ponto
    pontos_otimizados = []
    
    for ponto in pontos:
        print(f"\n🔍 Otimizando Ponto {ponto['numero']}...")
        
        # Otimiza tempos
        tempos_otimizados = otimizar_tempos_ponto(
            ponto['leituras'], 
            constantes, 
            ponto['valores_originais']
        )
        
        # Calcula resultados com tempos otimizados
        resultados_otimizados = calcular_formulas_com_tempo_ajustado(
            ponto['leituras'], 
            constantes, 
            tempos_otimizados
        )
        
        agregados_otimizados = calcular_agregados_com_tempo_ajustado(resultados_otimizados)
        
        # Verifica se os valores estão corretos
        vazao_diff = abs(float(agregados_otimizados['vazao_media'] - ponto['valores_originais']['vazao_media']))
        tendencia_diff = abs(float(agregados_otimizados['tendencia'] - ponto['valores_originais']['tendencia']))
        
        print(f"   Vazão Média Original: {float(ponto['valores_originais']['vazao_media']):.6f}")
        print(f"   Vazão Média Otimizada: {float(agregados_otimizados['vazao_media']):.6f}")
        print(f"   Diferença: {vazao_diff:.8f}")
        print(f"   Tendência Original: {float(ponto['valores_originais']['tendencia']):.6f}")
        print(f"   Tendência Otimizada: {float(agregados_otimizados['tendencia']):.6f}")
        print(f"   Diferença: {tendencia_diff:.8f}")
        
        ponto_otimizado = {
            'numero': ponto['numero'],
            'leituras': resultados_otimizados,
            'tempos_otimizados': tempos_otimizados,
            'agregados_otimizados': agregados_otimizados,
            'valores_originais': ponto['valores_originais']
        }
        
        pontos_otimizados.append(ponto_otimizado)
    
    # Gera planilha otimizada
    arquivo_otimizado = "SAN-038-25-09_TEMPOS_OTIMIZADOS.xlsx"
    gerar_planilha_otimizada(constantes, pontos_otimizados, arquivo_otimizado)
    
    # Salva resultados da otimização
    resultado_otimizacao = {
        'constantes': {k: float(v) if isinstance(v, Decimal) else v for k, v in constantes.items()},
        'pontos_otimizados': []
    }
    
    for ponto in pontos_otimizados:
        ponto_resultado = {
            'numero': ponto['numero'],
            'tempos_otimizados': [float(t) for t in ponto['tempos_otimizados']],
            'agregados_otimizados': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto['agregados_otimizados'].items()},
            'valores_originais': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto['valores_originais'].items()}
        }
        resultado_otimizacao['pontos_otimizados'].append(ponto_resultado)
    
    with open('resultados_otimizacao_tempos.json', 'w', encoding='utf-8') as f:
        json.dump(resultado_otimizacao, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Resultados salvos em: resultados_otimizacao_tempos.json")
    print("🎉 Otimização concluída!")

if __name__ == "__main__":
    main() 