# -*- coding: utf-8 -*-
"""
Verificador de Cálculos da Planilha Original
Lê os valores brutos da planilha original e refaz todos os cálculos
para verificar se estão corretos
"""

import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import json
import os

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def extrair_constantes_planilha_original(arquivo_excel):
    """
    Extrai todas as constantes necessárias para os cálculos da planilha original
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        estimativa_sheet = wb["Estimativa da Incerteza"]
        
        constantes = {}
        
        # Constantes do ponto (linha 50-51)
        constantes['ponto_mlp'] = ler_valor_exato(coleta_sheet, 50, 9)  # I50: 1° Ponto • mL/P
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_sheet, 50, 30)  # AD50: Pulso do Equipamento • mL/P
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_sheet, 51, 18)  # R51: Constante de Correção (Temperatura)
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_sheet, 51, 21)  # U51: Constante de Correção (Inclinação/Fator)
        constantes['modo_calibracao'] = coleta_sheet.cell(row=16, column=24).value  # X16: Modo de Calibração
        
        # Constantes de correção do Tempo (Estimativa da Incerteza)
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_sheet, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_sheet, 23, 75)  # BW23
        
        # Constantes de correção da Temperatura (Estimativa da Incerteza)
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_sheet, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_sheet, 26, 75)  # BW26
        
        print("✅ Constantes extraídas da planilha original:")
        print(f"   Ponto mL/P: {float(constantes['ponto_mlp'])}")
        print(f"   Pulso Equipamento mL/P: {float(constantes['pulso_equipamento_mlp'])}")
        print(f"   Constante Correção Temp: {float(constantes['constante_correcao_temp'])}")
        print(f"   Constante Correção Inclinação: {float(constantes['constante_correcao_inclinacao'])}")
        print(f"   Modo Calibração: {constantes['modo_calibracao']}")
        print(f"   Correção Tempo BU23: {float(constantes['correcao_tempo_bu23'])}")
        print(f"   Correção Tempo BW23: {float(constantes['correcao_tempo_bw23'])}")
        print(f"   Correção Temp BU26: {float(constantes['correcao_temp_bu26'])}")
        print(f"   Correção Temp BW26: {float(constantes['correcao_temp_bw26'])}")
        
        return constantes
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def extrair_valores_brutos_planilha_original(arquivo_excel):
    """
    Extrai os valores brutos (entradas) da planilha original
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        valores_brutos = {}
        
        # Identifica os pontos (mesma lógica)
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': []
            }
            
            # Extrai valores brutos (entradas) da planilha original
            for i in range(3):
                linha = linha_atual + i
                
                # Valores brutos (entradas)
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # C54: Qtd de Pulsos do Padrão
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # F54: Tempo de Coleta • (s)
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # O54: Leitura no Medidor • L
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # R54: Temperatura da Água • °C
                
                leitura = {
                    'linha': linha,
                    # Valores brutos (entradas)
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura
                }
                
                ponto['leituras'].append(leitura)
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        for ponto in pontos:
            valores_brutos[f"ponto_{ponto['numero']}"] = ponto
        
        return valores_brutos
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair valores brutos: {e}")
        return None

def calcular_formulas_ponto(leituras, constantes):
    """
    Calcula todas as fórmulas para um ponto específico
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        linha = leitura['linha']
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = leitura['tempo_coleta']
        tempo_coleta_corrigido = tempo_coleta_bruto - (tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23'])
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        temperatura_corrigida = temperatura_bruta - (temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26'])
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100') * volume_bruto
        totalizacao_padrao_corrigido = volume_bruto - correcao
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        
        # FÓRMULA 7: Vazão do Medidor • L/h (X54)
        leitura_medidor = leitura['leitura_medidor']
        modo_calibracao = constantes['modo_calibracao']
        
        if modo_calibracao in ["Visual com início dinâmico", "Visual com início estática"]:
            vazao_medidor = leitura_medidor
        else:
            vazao_medidor = (leitura_medidor / tempo_coleta_corrigido) * Decimal('3600')
        
        # FÓRMULA 8: Erro % (U54)
        erro_percentual = (leitura_medidor - totalizacao_padrao_corrigido) / totalizacao_padrao_corrigido * Decimal('100')
        
        resultado = {
            'linha': linha,
            'pulsos_padrao': pulsos_padrao,
            'tempo_coleta_bruto': tempo_coleta_bruto,
            'tempo_coleta_corrigido': tempo_coleta_corrigido,
            'temperatura_bruta': temperatura_bruta,
            'temperatura_corrigida': temperatura_corrigida,
            'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
            'vazao_referencia': vazao_referencia,
            'leitura_medidor': leitura_medidor,
            'vazao_medidor': vazao_medidor,
            'erro_percentual': erro_percentual,
            'constantes_usadas': {
                'pulso_padrao_lp': pulso_padrao_lp,
                'pulso_equipamento_lp': pulso_equipamento_lp
            }
        }
        
        resultados.append(resultado)
        
        print(f"   Leitura {i+1} (Linha {linha}):")
        print(f"     Tempo Corrigido: {float(tempo_coleta_corrigido):.6f} s")
        print(f"     Temp Corrigida: {float(temperatura_corrigida):.6f} °C")
        print(f"     Totalização Padrão: {float(totalizacao_padrao_corrigido):.6f} L")
        print(f"     Vazão Referência: {float(vazao_referencia):.6f} L/h")
        print(f"     Vazão Medidor: {float(vazao_medidor):.6f} L/h")
        print(f"     Erro: {float(erro_percentual):.6f} %")
    
    return resultados

def calcular_agregados_ponto(resultados):
    """
    Calcula os valores agregados do ponto (Fórmulas 9, 10, 11)
    """
    vazoes_referencia = [r['vazao_referencia'] for r in resultados]
    erros_percentuais = [r['erro_percentual'] for r in resultados]
    
    # FÓRMULA 9: Vazão Média • L/h (I57)
    vazao_media = sum(vazoes_referencia) / Decimal(str(len(vazoes_referencia)))
    
    # FÓRMULA 10: Tendência (U57)
    tendencia = sum(erros_percentuais) / Decimal(str(len(erros_percentuais)))
    
    # FÓRMULA 11: DESVIO PADRÃO AMOSTRAL (AD57)
    desvio_padrao = calcular_desvio_padrao_amostral(erros_percentuais)
    
    return {
        'vazao_media': vazao_media,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def processar_ponto_original(ponto_brutos, constantes):
    """
    Processa um ponto específico da planilha original
    """
    print(f"\n🔍 Processando Ponto {ponto_brutos['numero']}...")
    
    # Calcula todas as fórmulas usando os valores brutos
    resultados_calculados = calcular_formulas_ponto(ponto_brutos['leituras'], constantes)
    
    # Calcula agregados
    agregados_calculados = calcular_agregados_ponto(resultados_calculados)
    
    print(f"   ✅ Ponto {ponto_brutos['numero']} processado:")
    print(f"     Vazão Média: {float(agregados_calculados['vazao_media']):.6f} L/h")
    print(f"     Tendência: {float(agregados_calculados['tendencia']):.6f} %")
    print(f"     Desvio Padrão: {float(agregados_calculados['desvio_padrao']) if agregados_calculados['desvio_padrao'] else 'N/A':.6f} %")
    
    return {
        'numero': ponto_brutos['numero'],
        'leituras': resultados_calculados,
        'agregados': agregados_calculados
    }

def main():
    """
    Função principal
    """
    # Arquivo da planilha original
    arquivo_planilha = "SAN-038-25-09.xlsx"
    
    if not os.path.exists(arquivo_planilha):
        print(f"❌ Arquivo não encontrado: {arquivo_planilha}")
        return
    
    print("🚀 Iniciando processamento da planilha original...")
    print("=" * 60)
    
    # Extrai constantes
    constantes = extrair_constantes_planilha_original(arquivo_planilha)
    if constantes is None:
        return
    
    # Extrai valores brutos
    valores_brutos = extrair_valores_brutos_planilha_original(arquivo_planilha)
    if valores_brutos is None:
        return
    
    print(f"\n✅ Extraídos {len(valores_brutos)} pontos com valores brutos")
    
    # Processa cada ponto
    resultados_completos = {}
    
    for ponto_key, ponto_brutos in valores_brutos.items():
        resultado_ponto = processar_ponto_original(ponto_brutos, constantes)
        resultados_completos[ponto_key] = resultado_ponto
    
    # Salva resultados
    resultado_final = {
        'constantes': {k: float(v) if isinstance(v, Decimal) else v for k, v in constantes.items()},
        'pontos': resultados_completos,
        'total_pontos': len(resultados_completos)
    }
    
    # Converte Decimal para float para serialização JSON
    def converter_decimais(obj):
        if isinstance(obj, dict):
            return {k: converter_decimais(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [converter_decimais(v) for v in obj]
        elif isinstance(obj, Decimal):
            return float(obj)
        else:
            return obj
    
    resultado_final = converter_decimais(resultado_final)
    
    with open('resultados_planilha_original.json', 'w', encoding='utf-8') as f:
        json.dump(resultado_final, f, indent=2, ensure_ascii=False)
    
    print("\n" + "=" * 60)
    print("✅ Processamento da planilha original concluído!")
    print(f"📊 Total de pontos processados: {len(resultados_completos)}")
    print(f"✅ Resultados salvos em: resultados_planilha_original.json")
    
    # Resumo final
    for ponto_key, ponto in resultados_completos.items():
        agregados = ponto['agregados']
        print(f"   {ponto_key}:")
        print(f"     Vazão Média: {float(agregados['vazao_media']):.6f} L/h")
        print(f"     Tendência: {float(agregados['tendencia']):.6f} %")
        if agregados['desvio_padrao']:
            print(f"     Desvio Padrão: {float(agregados['desvio_padrao']):.6f} %")

if __name__ == "__main__":
    main() 