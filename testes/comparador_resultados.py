# -*- coding: utf-8 -*-
"""
Comparador de Resultados
Compara os valores calculados com os valores da planilha corrigida
"""

import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import json
import os

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def comparar_valores(calculado, original, tolerancia=Decimal('0.000001')):
    """
    Compara dois valores com uma tolerância
    """
    # Converte ambos para Decimal para comparação
    if isinstance(calculado, float):
        calculado = Decimal(str(calculado))
    if isinstance(original, float):
        original = Decimal(str(original))
    
    diferenca = abs(calculado - original)
    return diferenca <= tolerancia

def extrair_valores_planilha_corrigida(arquivo_excel):
    """
    Extrai os valores calculados da planilha corrigida para comparação
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        valores_planilha = {}
        
        # Identifica os pontos (mesma lógica)
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': []
            }
            
            # Extrai valores calculados da planilha corrigida
            for i in range(3):
                linha = linha_atual + i
                
                # Valores calculados da planilha
                tempo_coleta_corrigido = ler_valor_exato(coleta_sheet, linha, 27)  # AA54
                temperatura_corrigida = ler_valor_exato(coleta_sheet, linha, 30)    # AD54
                totalizacao_padrao_corrigido = ler_valor_exato(coleta_sheet, linha, 12)  # L54
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)         # I54
                vazao_medidor = ler_valor_exato(coleta_sheet, linha, 24)           # X54
                erro_percentual = ler_valor_exato(coleta_sheet, linha, 21)         # U54
                
                leitura = {
                    'linha': linha,
                    'tempo_coleta_corrigido': tempo_coleta_corrigido,
                    'temperatura_corrigida': temperatura_corrigida,
                    'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
                    'vazao_referencia': vazao_referencia,
                    'vazao_medidor': vazao_medidor,
                    'erro_percentual': erro_percentual
                }
                
                ponto['leituras'].append(leitura)
            
            # Valores agregados (linha 57)
            linha_agregados = linha_atual + 3
            vazao_media = ler_valor_exato(coleta_sheet, linha_agregados, 9)      # I57
            tendencia = ler_valor_exato(coleta_sheet, linha_agregados, 21)        # U57
            desvio_padrao = ler_valor_exato(coleta_sheet, linha_agregados, 30)    # AD57
            
            ponto['agregados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        for ponto in pontos:
            valores_planilha[f"ponto_{ponto['numero']}"] = ponto
        
        return valores_planilha
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair valores da planilha corrigida: {e}")
        return None

def comparar_resultados(resultados_calculados, valores_planilha):
    """
    Compara os resultados calculados com os valores da planilha
    """
    print("🔍 Comparando resultados...")
    print("=" * 60)
    
    total_comparacoes = 0
    total_iguais = 0
    diferencas_encontradas = []
    
    for ponto_key in resultados_calculados.keys():
        if ponto_key not in valores_planilha:
            print(f"❌ Ponto {ponto_key} não encontrado na planilha corrigida")
            continue
        
        ponto_calculado = resultados_calculados[ponto_key]
        ponto_planilha = valores_planilha[ponto_key]
        
        print(f"\n📊 Comparando {ponto_key}:")
        
        # Compara cada leitura
        for i, (leitura_calc, leitura_plan) in enumerate(zip(ponto_calculado['leituras'], ponto_planilha['leituras'])):
            linha = leitura_calc['linha']
            print(f"   Leitura {i+1} (Linha {linha}):")
            
            # Compara cada valor calculado
            campos_comparacao = [
                ('tempo_coleta_corrigido', 'Tempo Corrigido'),
                ('temperatura_corrigida', 'Temp Corrigida'),
                ('totalizacao_padrao_corrigido', 'Totalização Padrão'),
                ('vazao_referencia', 'Vazão Referência'),
                ('vazao_medidor', 'Vazão Medidor'),
                ('erro_percentual', 'Erro %')
            ]
            
            for campo, nome in campos_comparacao:
                valor_calc = leitura_calc[campo]
                valor_plan = leitura_plan[campo]
                
                total_comparacoes += 1
                if comparar_valores(valor_calc, valor_plan):
                    total_iguais += 1
                    print(f"     ✅ {nome}: {float(valor_calc):.6f} = {float(valor_plan):.6f}")
                else:
                    diferenca = abs(valor_calc - valor_plan)
                    print(f"     ❌ {nome}: {float(valor_calc):.6f} ≠ {float(valor_plan):.6f} (dif: {float(diferenca):.8f})")
                    diferencas_encontradas.append({
                        'ponto': ponto_key,
                        'leitura': i+1,
                        'linha': linha,
                        'campo': nome,
                        'calculado': valor_calc,
                        'planilha': valor_plan,
                        'diferenca': diferenca
                    })
        
        # Compara valores agregados
        agregados_calc = ponto_calculado['agregados']
        agregados_plan = ponto_planilha['agregados']
        
        print(f"   Agregados:")
        campos_agregados = [
            ('vazao_media', 'Vazão Média'),
            ('tendencia', 'Tendência'),
            ('desvio_padrao', 'Desvio Padrão')
        ]
        
        for campo, nome in campos_agregados:
            valor_calc = agregados_calc[campo]
            valor_plan = agregados_plan[campo]
            
            total_comparacoes += 1
            if comparar_valores(valor_calc, valor_plan):
                total_iguais += 1
                print(f"     ✅ {nome}: {float(valor_calc):.6f} = {float(valor_plan):.6f}")
            else:
                diferenca = abs(valor_calc - valor_plan)
                print(f"     ❌ {nome}: {float(valor_calc):.6f} ≠ {float(valor_plan):.6f} (dif: {float(diferenca):.8f})")
                diferencas_encontradas.append({
                    'ponto': ponto_key,
                    'leitura': 'agregado',
                    'linha': 'agregado',
                    'campo': nome,
                    'calculado': valor_calc,
                    'planilha': valor_plan,
                    'diferenca': diferenca
                })
    
    # Resumo final
    print("\n" + "=" * 60)
    print("📊 RESUMO DA COMPARAÇÃO:")
    print(f"   Total de comparações: {total_comparacoes}")
    print(f"   Valores iguais: {total_iguais}")
    print(f"   Valores diferentes: {total_comparacoes - total_iguais}")
    print(f"   Taxa de acerto: {(total_iguais/total_comparacoes)*100:.2f}%")
    
    if diferencas_encontradas:
        print(f"\n❌ DIFERENÇAS ENCONTRADAS ({len(diferencas_encontradas)}):")
        for diff in diferencas_encontradas:
            print(f"   {diff['ponto']}, {diff['campo']}: {float(diff['calculado']):.6f} ≠ {float(diff['planilha']):.6f} (dif: {float(diff['diferenca']):.8f})")
    else:
        print("\n✅ TODOS OS VALORES ESTÃO IDÊNTICOS!")
    
    return total_iguais == total_comparacoes

def main():
    """
    Função principal
    """
    # Arquivos
    arquivo_planilha_corrigida = "SAN-038-25-09_CORRIGIDO.xlsx"
    arquivo_resultados = "resultados_completos_planilha.json"
    
    if not os.path.exists(arquivo_planilha_corrigida):
        print(f"❌ Arquivo não encontrado: {arquivo_planilha_corrigida}")
        return
    
    if not os.path.exists(arquivo_resultados):
        print(f"❌ Arquivo não encontrado: {arquivo_resultados}")
        return
    
    print("🚀 Iniciando comparação de resultados...")
    print("=" * 60)
    
    # Carrega resultados calculados
    with open(arquivo_resultados, 'r', encoding='utf-8') as f:
        resultados_calculados = json.load(f)
    
    # Extrai valores da planilha corrigida
    valores_planilha = extrair_valores_planilha_corrigida(arquivo_planilha_corrigida)
    
    if valores_planilha is None:
        print("❌ Falha ao extrair valores da planilha corrigida")
        return
    
    # Compara os resultados
    sucesso = comparar_resultados(resultados_calculados['pontos'], valores_planilha)
    
    if sucesso:
        print("\n🎉 SUCESSO: Todos os cálculos estão corretos!")
    else:
        print("\n⚠️  ATENÇÃO: Foram encontradas diferenças nos cálculos")

if __name__ == "__main__":
    main() 