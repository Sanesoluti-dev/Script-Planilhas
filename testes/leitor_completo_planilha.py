# -*- coding: utf-8 -*-
"""
Leitor Completo da Planilha Original
Implementa todas as fórmulas mapeadas com máxima precisão
Baseado no mapeamento.md e formulas.md
"""

import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import json
import os

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            # Remove espaços e converte vírgula para ponto
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def extrair_constantes_planilha(arquivo_excel):
    """
    Extrai todas as constantes necessárias para os cálculos
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        estimativa_sheet = wb["Estimativa da Incerteza"]
        
        constantes = {}
        
        # Constantes do ponto (linha 50-51)
        constantes['ponto_mlp'] = ler_valor_exato(coleta_sheet, 50, 9)  # I50: 1° Ponto • mL/P
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_sheet, 50, 30)  # AD50: Pulso do Equipamento • mL/P
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_sheet, 51, 18)  # R51: Constante de Correção (Temperatura)
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_sheet, 51, 21)  # U51: Constante de Correção (Inclinação/Fator)
        constantes['modo_calibracao'] = coleta_sheet.cell(row=16, column=24).value  # X16: Modo de Calibração
        
        # Constantes de correção do Tempo (Estimativa da Incerteza)
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_sheet, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_sheet, 23, 75)  # BW23
        
        # Constantes de correção da Temperatura (Estimativa da Incerteza)
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_sheet, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_sheet, 26, 75)  # BW26
        
        print("✅ Constantes extraídas:")
        print(f"   Ponto mL/P: {float(constantes['ponto_mlp'])}")
        print(f"   Pulso Equipamento mL/P: {float(constantes['pulso_equipamento_mlp'])}")
        print(f"   Constante Correção Temp: {float(constantes['constante_correcao_temp'])}")
        print(f"   Constante Correção Inclinação: {float(constantes['constante_correcao_inclinacao'])}")
        print(f"   Modo Calibração: {constantes['modo_calibracao']}")
        print(f"   Correção Tempo BU23: {float(constantes['correcao_tempo_bu23'])}")
        print(f"   Correção Tempo BW23: {float(constantes['correcao_tempo_bw23'])}")
        print(f"   Correção Temp BU26: {float(constantes['correcao_temp_bu26'])}")
        print(f"   Correção Temp BW26: {float(constantes['correcao_temp_bw26'])}")
        
        return constantes
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def calcular_formulas_ponto(leituras, constantes):
    """
    Calcula todas as fórmulas para um ponto específico
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        linha = leitura['linha']
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = leitura['tempo_coleta']
        tempo_coleta_corrigido = tempo_coleta_bruto - (tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23'])
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        temperatura_corrigida = temperatura_bruta - (temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26'])
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100') * volume_bruto
        totalizacao_padrao_corrigido = volume_bruto - correcao
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        
        # FÓRMULA 7: Vazão do Medidor • L/h (X54)
        leitura_medidor = leitura['leitura_medidor']
        modo_calibracao = constantes['modo_calibracao']
        
        if modo_calibracao in ["Visual com início dinâmico", "Visual com início estática"]:
            vazao_medidor = leitura_medidor
        else:
            vazao_medidor = (leitura_medidor / tempo_coleta_corrigido) * Decimal('3600')
        
        # FÓRMULA 8: Erro % (U54)
        erro_percentual = (leitura_medidor - totalizacao_padrao_corrigido) / totalizacao_padrao_corrigido * Decimal('100')
        
        resultado = {
            'linha': linha,
            'pulsos_padrao': pulsos_padrao,
            'tempo_coleta_bruto': tempo_coleta_bruto,
            'tempo_coleta_corrigido': tempo_coleta_corrigido,
            'temperatura_bruta': temperatura_bruta,
            'temperatura_corrigida': temperatura_corrigida,
            'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
            'vazao_referencia': vazao_referencia,
            'leitura_medidor': leitura_medidor,
            'vazao_medidor': vazao_medidor,
            'erro_percentual': erro_percentual,
            'constantes_usadas': {
                'pulso_padrao_lp': pulso_padrao_lp,
                'pulso_equipamento_lp': pulso_equipamento_lp
            }
        }
        
        resultados.append(resultado)
        
        print(f"   Leitura {i+1} (Linha {linha}):")
        print(f"     Tempo Corrigido: {float(tempo_coleta_corrigido):.6f} s")
        print(f"     Temp Corrigida: {float(temperatura_corrigida):.6f} °C")
        print(f"     Totalização Padrão: {float(totalizacao_padrao_corrigido):.6f} L")
        print(f"     Vazão Referência: {float(vazao_referencia):.6f} L/h")
        print(f"     Vazão Medidor: {float(vazao_medidor):.6f} L/h")
        print(f"     Erro: {float(erro_percentual):.6f} %")
    
    return resultados

def calcular_agregados_ponto(resultados):
    """
    Calcula os valores agregados do ponto (Fórmulas 9, 10, 11)
    """
    vazoes_referencia = [r['vazao_referencia'] for r in resultados]
    erros_percentuais = [r['erro_percentual'] for r in resultados]
    
    # FÓRMULA 9: Vazão Média • L/h (I57)
    vazao_media = sum(vazoes_referencia) / Decimal(str(len(vazoes_referencia)))
    
    # FÓRMULA 10: Tendência (U57)
    tendencia = sum(erros_percentuais) / Decimal(str(len(erros_percentuais)))
    
    # FÓRMULA 11: DESVIO PADRÃO AMOSTRAL (AD57)
    desvio_padrao = calcular_desvio_padrao_amostral(erros_percentuais)
    
    return {
        'vazao_media': vazao_media,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def identificar_pontos_planilha(arquivo_excel):
    """
    Identifica todos os pontos de calibração na planilha
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        pontos = []
        linha_atual = 54  # Primeira linha de dados
        num_ponto = 1
        
        while True:
            # Verifica se há dados na linha atual
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)  # Coluna C
            
            if pulsos == 0:
                # Verifica as próximas 2 linhas também
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break  # Fim dos pontos
            
            # Encontrou um ponto
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': []
            }
            
            # Extrai as 3 leituras do ponto
            for i in range(3):
                linha = linha_atual + i
                
                # Lê todos os valores brutos
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # C54: Qtd de Pulsos do Padrão
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # F54: Tempo de Coleta • (s)
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # O54: Leitura no Medidor • L
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # R54: Temperatura da Água • °C
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura
                }
                
                ponto['leituras'].append(leitura)
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9  # Avança para o próximo ponto
        
        print(f"✅ Encontrados {len(pontos)} pontos de calibração")
        for ponto in pontos:
            print(f"   Ponto {ponto['numero']}: linhas {ponto['linha_inicial']}-{ponto['linha_inicial']+2}")
        
        return pontos
        
    except Exception as e:
        print(f"ERRO: Erro ao identificar pontos: {e}")
        return None

def processar_planilha_completa(arquivo_excel):
    """
    Processa a planilha completa aplicando todas as fórmulas
    """
    try:
        print(f"📖 Processando planilha completa: {arquivo_excel}")
        
        # Extrai constantes
        constantes = extrair_constantes_planilha(arquivo_excel)
        if constantes is None:
            return None
        
        # Identifica todos os pontos
        pontos = identificar_pontos_planilha(arquivo_excel)
        if pontos is None:
            return None
        
        resultados_completos = {}
        
        for ponto in pontos:
            print(f"\n🔍 Processando Ponto {ponto['numero']}...")
            
            # Calcula todas as fórmulas para o ponto
            resultados_calculos = calcular_formulas_ponto(ponto['leituras'], constantes)
            
            # Calcula agregados
            agregados = calcular_agregados_ponto(resultados_calculos)
            
            resultado_ponto = {
                'numero': ponto['numero'],
                'leituras': resultados_calculos,
                'agregados': agregados
            }
            
            resultados_completos[f"ponto_{ponto['numero']}"] = resultado_ponto
            
            print(f"   ✅ Ponto {ponto['numero']} processado:")
            print(f"     Vazão Média: {float(agregados['vazao_media']):.6f} L/h")
            print(f"     Tendência: {float(agregados['tendencia']):.6f} %")
            print(f"     Desvio Padrão: {float(agregados['desvio_padrao']) if agregados['desvio_padrao'] else 'N/A':.6f} %")
        
        return {
            'constantes': constantes,
            'pontos': resultados_completos,
            'total_pontos': len(pontos)
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao processar planilha: {e}")
        return None

def salvar_resultados_json(resultados, nome_arquivo):
    """
    Salva os resultados em formato JSON
    """
    try:
        # Converte Decimal para float para serialização JSON
        def converter_decimais(obj):
            if isinstance(obj, dict):
                return {k: converter_decimais(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [converter_decimais(v) for v in obj]
            elif isinstance(obj, Decimal):
                return float(obj)
            else:
                return obj
        
        resultados_convertidos = converter_decimais(resultados)
        
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            json.dump(resultados_convertidos, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Resultados salvos em: {nome_arquivo}")
        
    except Exception as e:
        print(f"ERRO ao salvar resultados: {e}")

def main():
    """
    Função principal
    """
    # Arquivo da planilha original
    arquivo_planilha = "SAN-038-25-09.xlsx"
    
    if not os.path.exists(arquivo_planilha):
        print(f"❌ Arquivo não encontrado: {arquivo_planilha}")
        return
    
    print("🚀 Iniciando processamento completo da planilha...")
    print("=" * 60)
    
    # Processa a planilha completa
    resultados = processar_planilha_completa(arquivo_planilha)
    
    if resultados is None:
        print("❌ Falha ao processar planilha")
        return
    
    # Salva os resultados
    salvar_resultados_json(resultados, "resultados_completos_planilha.json")
    
    print("=" * 60)
    print("✅ Processamento completo concluído!")
    print(f"📊 Total de pontos processados: {resultados['total_pontos']}")
    
    # Resumo final
    for ponto_key, ponto in resultados['pontos'].items():
        agregados = ponto['agregados']
        print(f"   {ponto_key}:")
        print(f"     Vazão Média: {float(agregados['vazao_media']):.6f} L/h")
        print(f"     Tendência: {float(agregados['tendencia']):.6f} %")
        if agregados['desvio_padrao']:
            print(f"     Desvio Padrão: {float(agregados['desvio_padrao']):.6f} %")

if __name__ == "__main__":
    main() 