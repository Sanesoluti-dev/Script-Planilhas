# -*- coding: utf-8 -*-
"""
Leitor da Nova Planilha
Aplica a mesma lógica de leitura do ajustador_tempo_coleta.py para a nova planilha
"""

import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import json
import os

def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            # Remove espaços e converte vírgula para ponto
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal com 15 casas decimais
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média com 15 casas decimais
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def get_numeric_value(df, row, col):
    """
    Obtém valor numérico do DataFrame pandas
    """
    try:
        valor = df.iloc[row, col]
        if pd.isna(valor):
            return 0
        return float(valor)
    except:
        return 0

def extrair_dados_nova_planilha(arquivo_excel):
    """
    PASSO 1: Extração de Dados da Nova Planilha
    Extrai todos os parâmetros de entrada brutos das abas "Coleta de Dados"
    Usa a mesma lógica do ajustador_tempo_coleta.py
    """
    try:
        print(f"📖 PASSO 1: Extraindo dados da nova planilha: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        # Configuração dos pontos (mesma lógica do original)
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_originais = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto (mesma estrutura de colunas)
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i  # +4 em vez de +3 para pular a linha do título
                
                # Lê todos os parâmetros necessários (mesmas colunas do original)
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Pulsos: {float(pulsos_padrao)}")
                print(f"     Tempo: {float(tempo_coleta)} s")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Leitura Medidor: {float(leitura_medidor)} L")
                print(f"     Temperatura: {float(temperatura)} °C")
                print(f"     Erro: {float(erro)} %")

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros) - usa todos os erros, não filtra valores zero
            tendencia = sum(erros) / Decimal(str(len(erros)))
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            print(f"   VALORES SAGRADOS do Ponto {config['num_ponto']}:")
            print(f"     Vazão Média: {float(vazao_media)} L/h")
            print(f"     Tendência: {float(tendencia)} %")
            print(f"     Desvio Padrão: {float(desvio_padrao) if desvio_padrao else 'N/A'} %")
            
            dados_originais[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_originais
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados da nova planilha: {e}")
        return None

def extrair_constantes_nova_planilha(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas (mesmas posições do original)
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas da nova planilha:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes da nova planilha: {e}")
        return None

def salvar_dados_json(dados, nome_arquivo):
    """
    Salva os dados em formato JSON para análise posterior
    """
    try:
        # Converte Decimal para float para serialização JSON
        def converter_decimais(obj):
            if isinstance(obj, dict):
                return {k: converter_decimais(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [converter_decimais(v) for v in obj]
            elif isinstance(obj, Decimal):
                return float(obj)
            else:
                return obj
        
        dados_convertidos = converter_decimais(dados)
        
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            json.dump(dados_convertidos, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Dados salvos em: {nome_arquivo}")
        
    except Exception as e:
        print(f"ERRO ao salvar dados: {e}")

def main():
    """
    Função principal para ler a nova planilha
    """
    # Arquivo da nova planilha
    arquivo_nova_planilha = "SAN-038-25-09_CORRIGIDO.xlsx"
    
    if not os.path.exists(arquivo_nova_planilha):
        print(f"❌ Arquivo não encontrado: {arquivo_nova_planilha}")
        return
    
    print("🚀 Iniciando leitura da nova planilha...")
    print("=" * 60)
    
    # Extrai dados originais
    dados_originais = extrair_dados_nova_planilha(arquivo_nova_planilha)
    
    if dados_originais is None:
        print("❌ Falha ao extrair dados originais")
        return
    
    # Extrai constantes
    constantes = extrair_constantes_nova_planilha(arquivo_nova_planilha)
    
    if constantes is None:
        print("❌ Falha ao extrair constantes")
        return
    
    # Salva os dados extraídos
    resultado = {
        'dados_originais': dados_originais,
        'constantes': constantes,
        'arquivo_origem': arquivo_nova_planilha
    }
    
    salvar_dados_json(resultado, "dados_nova_planilha.json")
    
    print("=" * 60)
    print("✅ Leitura da nova planilha concluída com sucesso!")
    print(f"📊 Total de pontos extraídos: {len(dados_originais)}")
    
    # Resumo dos dados
    for ponto_key, ponto in dados_originais.items():
        print(f"   {ponto_key}: {len(ponto['leituras'])} leituras")
        if ponto['valores_sagrados']['desvio_padrao']:
            print(f"     Desvio Padrão: {float(ponto['valores_sagrados']['desvio_padrao']):.6f} %")

if __name__ == "__main__":
    main() 