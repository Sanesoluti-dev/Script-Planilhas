# -*- coding: utf-8 -*-
"""
AJUSTADOR DE VAZÃO MÉDIA - ANÁLISE E AJUSTE DE TEMPOS DE COLETA
==================================================================

Este script analisa o tempo de coleta e ajusta os valores para aproximar ao máximo
a vazão média original, usando incrementos de 0.001 segundos e mantendo a base de 240.

PRINCÍPIO FUNDAMENTAL:
- Lê dados originais do arquivo SAN-038-25-09.xlsx
- Lê dados corrigidos do arquivo SAN-038-25-09_CORRIGIDO.xlsx
- Ajusta tempos de coleta com incrementos de 0.001s
- Mantém base de 240 (pode ter casas decimais)
- Aplica para CADA LEITURA de CADA PONTO
- Usa fórmulas críticas da planilha conforme documentação

FÓRMULAS UTILIZADAS:
- Vazão de Referência: =SE(C54="";"";L54/AA54*3600)
- Vazão Média: =SE(I54="";"";MÉDIA(I54:I56))
- Totalização: =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
"""

import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook
import shutil
import os

# Configurar precisão alta
getcontext().prec = 15

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_totalizacao_padrao_corrigido(pulsos_padrao, pulso_padrao_lp, temperatura, fator_correcao_temp, tempo_coleta):
    """
    Calcula a "Totalização no Padrão Corrigido • L" usando a fórmula:
    =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
    """
    if pulsos_padrao == 0:
        return Decimal('0')
    
    # C54*$I$51 = Pulsos * Pulso do padrão em L/P
    volume_pulsos = pulsos_padrao * pulso_padrao_lp
    
    # (C54*$I$51/AA54*3600) = Volume / Tempo * 3600 = Vazão
    vazao = volume_pulsos / tempo_coleta * Decimal('3600')
    
    # ($R$51+$U$51*(C54*$I$51/AA54*3600))/100 = (Temperatura + Fator_Correção * Vazão) / 100
    fator_correcao = (temperatura + fator_correcao_temp * vazao) / Decimal('100')
    
    # (C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51))
    # = Volume - (Fator_Correção * Volume)
    totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
    
    return totalizacao

def calcular_vazao_referencia(totalizacao, tempo_coleta):
    """
    Calcula Vazão de Referência • L/h usando a fórmula: =SE(C54="";"";L54/AA54*3600)
    """
    if tempo_coleta == 0:
        return Decimal('0')
    
    # L54/AA54*3600 = Totalização / Tempo * 3600
    vazao_ref = (totalizacao / tempo_coleta) * Decimal('3600')
    return vazao_ref

def calcular_vazao_media(vazoes_ref):
    """
    Calcula Vazão Média • L/h usando a fórmula: =SE(I54="";"";MÉDIA(I54:I56))
    """
    if not vazoes_ref:
        return Decimal('0')
    
    # MÉDIA(I54:I56) = média das vazões de referência
    vazao_media = sum(vazoes_ref) / Decimal(str(len(vazoes_ref)))
    return vazao_media

def extrair_constantes_calculo(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def extrair_dados_planilha(arquivo_excel, descricao):
    """
    Extrai dados de uma planilha (original ou corrigida)
    """
    try:
        print(f"📖 Extraindo dados {descricao} do arquivo: {arquivo_excel}")
        
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Identifica os pontos de calibração
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_planilha = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i
                
                # Lê todos os parâmetros necessários
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros)
            tendencia = sum(erros) / Decimal(str(len(erros)))
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            dados_planilha[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_planilha
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados {descricao}: {e}")
        return None

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def ajustar_tempos_para_vazao_media(dados_originais, dados_corrigidos, constantes):
    """
    Ajusta tempos de coleta para aproximar ao máximo a vazão média original
    """
    print(f"\n🎯 AJUSTANDO TEMPOS DE COLETA PARA VAZÃO MÉDIA")
    print("=" * 80)
    
    resultados_ajuste = {}
    
    for ponto_key in dados_originais.keys():
        print(f"\n📊 Processando {ponto_key}:")
        
        # Extrai dados originais
        dados_orig = dados_originais[ponto_key]
        vazao_media_original = dados_orig['valores_sagrados']['vazao_media']
        
        # Extrai dados corrigidos
        dados_corr = dados_corrigidos[ponto_key]
        vazao_media_corrigida = dados_corr['valores_sagrados']['vazao_media']
        
        print(f"   📊 VALORES DE VAZÃO MÉDIA:")
        print(f"     Original: {float(vazao_media_original)} L/h")
        print(f"     Corrigida: {float(vazao_media_corrigida)} L/h")
        print(f"     Diferença: {float(vazao_media_corrigida - vazao_media_original)} L/h")
        
        # Calcula o erro percentual
        erro_percentual = ((vazao_media_corrigida - vazao_media_original) / vazao_media_original) * 100
        print(f"     Erro Percentual: {float(erro_percentual)} %")
        
        # Ajusta tempos para cada leitura
        leituras_ajustadas = []
        
        for i, leitura_corrigida in enumerate(dados_corr['leituras']):
            print(f"\n   🔧 Ajustando Leitura {i+1} (Linha {leitura_corrigida['linha']}):")
            
            # Tempo inicial baseado no arquivo corrigido
            tempo_inicial = leitura_corrigida['tempo_coleta']
            tempo_atual = tempo_inicial
            melhor_tempo = tempo_atual
            melhor_erro = abs(vazao_media_corrigida - vazao_media_original)
            
            # Parâmetros do loop
            incremento = Decimal('0.001')  # Incremento de 0.001 segundos
            max_iteracoes = 1000
            tolerancia = Decimal('1e-6')
            
            print(f"     Tempo inicial: {float(tempo_inicial)} s")
            print(f"     Incremento: {float(incremento)} s")
            print(f"     Máximo de iterações: {max_iteracoes}")
            
            # Loop de ajuste
            for iteracao in range(max_iteracoes):
                # Calcula totalização com tempo atual
                totalizacao = calcular_totalizacao_padrao_corrigido(
                    leitura_corrigida['pulsos_padrao'],
                    constantes['pulso_padrao_lp'],
                    constantes['temperatura_constante'],
                    constantes['fator_correcao_temp'],
                    tempo_atual
                )
                
                # Calcula vazão de referência
                vazao_ref = calcular_vazao_referencia(totalizacao, tempo_atual)
                
                # Calcula vazão média (simula as 3 leituras)
                # Para simplificar, usa a mesma vazão para as 3 leituras
                vazoes_ref = [vazao_ref, vazao_ref, vazao_ref]
                vazao_media_calculada = calcular_vazao_media(vazoes_ref)
                
                # Calcula erro
                erro_atual = abs(vazao_media_calculada - vazao_media_original)
                
                # Verifica se encontrou melhor solução
                if erro_atual < melhor_erro:
                    melhor_erro = erro_atual
                    melhor_tempo = tempo_atual
                    
                    if iteracao % 100 == 0:
                        print(f"       Iteração {iteracao + 1}: Nova melhor solução!")
                        print(f"         Tempo: {float(tempo_atual)} s")
                        print(f"         Vazão Média: {float(vazao_media_calculada)} L/h")
                        print(f"         Erro: {float(erro_atual)} L/h")
                
                # Verifica convergência
                if erro_atual < float(tolerancia):
                    print(f"       ✅ CONVERGÊNCIA ATINGIDA na iteração {iteracao + 1}!")
                    break
                
                # Incrementa tempo
                tempo_atual += incremento
                
                # Verifica se passou do limite (mantém base de 240)
                if tempo_atual > Decimal('240.499'):
                    tempo_atual = Decimal('240.0')
            
            else:
                print(f"       ⚠️  MÁXIMO DE ITERAÇÕES ATINGIDO")
            
            # Resultado final para esta leitura
            leitura_ajustada = {
                'linha': leitura_corrigida['linha'],
                'pulsos_padrao': leitura_corrigida['pulsos_padrao'],
                'tempo_coleta_original': leitura_corrigida['tempo_coleta'],
                'tempo_coleta_ajustado': melhor_tempo,
                'leitura_medidor': leitura_corrigida['leitura_medidor'],
                'temperatura': leitura_corrigida['temperatura'],
                'erro': leitura_corrigida['erro'],
                'vazao_referencia_original': leitura_corrigida['vazao_referencia'],
                'vazao_referencia_ajustada': float(calcular_vazao_referencia(
                    calcular_totalizacao_padrao_corrigido(
                        leitura_corrigida['pulsos_padrao'],
                        constantes['pulso_padrao_lp'],
                        constantes['temperatura_constante'],
                        constantes['fator_correcao_temp'],
                        melhor_tempo
                    ),
                    melhor_tempo
                )),
                'melhor_erro': float(melhor_erro),
                'iteracoes_realizadas': min(iteracao + 1, max_iteracoes)
            }
            
            leituras_ajustadas.append(leitura_ajustada)
            
            print(f"     ✅ RESULTADO FINAL:")
            print(f"       Tempo Original: {float(leitura_corrigida['tempo_coleta'])} s")
            print(f"       Tempo Ajustado: {float(melhor_tempo)} s")
            print(f"       Vazão Ref Original: {float(leitura_corrigida['vazao_referencia'])} L/h")
            print(f"       Vazão Ref Ajustada: {leitura_ajustada['vazao_referencia_ajustada']} L/h")
            print(f"       Melhor Erro: {leitura_ajustada['melhor_erro']} L/h")
            print(f"       Iterações: {leitura_ajustada['iteracoes_realizadas']}")
        
        # Calcula vazão média final com tempos ajustados
        vazoes_ref_finais = []
        for leitura in leituras_ajustadas:
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta_ajustado']
            )
            vazao_ref = calcular_vazao_referencia(totalizacao, leitura['tempo_coleta_ajustado'])
            vazoes_ref_finais.append(vazao_ref)
        
        vazao_media_final = calcular_vazao_media(vazoes_ref_finais)
        
        print(f"\n   📊 RESULTADO FINAL DO PONTO:")
        print(f"     Vazão Média Original: {float(vazao_media_original)} L/h")
        print(f"     Vazão Média Final: {float(vazao_media_final)} L/h")
        print(f"     Diferença Final: {float(vazao_media_final - vazao_media_original)} L/h")
        print(f"     Erro Percentual Final: {float(((vazao_media_final - vazao_media_original) / vazao_media_original) * 100)} %")
        
        resultados_ajuste[ponto_key] = {
            'numero_ponto': dados_orig['numero'],
            'vazao_media_original': float(vazao_media_original),
            'vazao_media_final': float(vazao_media_final),
            'diferenca_vazao': float(vazao_media_final - vazao_media_original),
            'erro_percentual': float(((vazao_media_final - vazao_media_original) / vazao_media_original) * 100),
            'leituras_ajustadas': leituras_ajustadas
        }
    
    return resultados_ajuste

def converter_decimal_para_float(obj):
    """
    Converte recursivamente todos os valores Decimal para float
    para permitir serialização JSON
    """
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {key: converter_decimal_para_float(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [converter_decimal_para_float(item) for item in obj]
    else:
        return obj

def gerar_json_ajuste_vazao_media(resultados_ajuste):
    """
    Gera JSON com os resultados do ajuste de vazão média
    """
    print(f"\n📄 GERANDO JSON DE AJUSTE DE VAZÃO MÉDIA")
    print("=" * 60)
    
    resultado_final = {
        "metadata": {
            "data_geracao": datetime.now().isoformat(),
            "descricao": "Ajuste de tempos de coleta para aproximar vazão média original",
            "arquivo_original": "SAN-038-25-09.xlsx",
            "arquivo_corrigido": "SAN-038-25-09_CORRIGIDO.xlsx",
            "incremento_utilizado": "0.001 segundos",
            "base_tempo": "240 segundos",
            "total_pontos": len(resultados_ajuste)
        },
        "resultados_por_ponto": resultados_ajuste
    }
    
    # Converte todos os valores Decimal para float
    resultado_final = converter_decimal_para_float(resultado_final)
    
    # Salva o JSON
    nome_arquivo = "ajuste_vazao_media.json"
    with open(nome_arquivo, "w", encoding="utf-8") as f:
        json.dump(resultado_final, f, indent=2, ensure_ascii=False)
    
    print(f"   ✅ JSON gerado: {nome_arquivo}")
    print(f"   📊 Total de pontos processados: {len(resultados_ajuste)}")
    
    return nome_arquivo

def main():
    """Função principal"""
    arquivo_original = "SAN-038-25-09.xlsx"
    arquivo_corrigido = "SAN-038-25-09_CORRIGIDO.xlsx"
    
    print("=== AJUSTADOR DE VAZÃO MÉDIA - ANÁLISE E AJUSTE DE TEMPOS DE COLETA ===")
    print("Ajusta tempos de coleta para aproximar ao máximo a vazão média original")
    print("Usa incrementos de 0.001 segundos mantendo base de 240")
    print("Aplica para CADA LEITURA de CADA PONTO")
    print(f"Arquivo Original: {arquivo_original}")
    print(f"Arquivo Corrigido: {arquivo_corrigido}")
    
    # Carrega dados originais
    dados_originais = extrair_dados_planilha(arquivo_original, "originais")
    if not dados_originais:
        print("❌ Falha na extração dos dados originais")
        return
    
    # Carrega dados corrigidos
    dados_corrigidos = extrair_dados_planilha(arquivo_corrigido, "corrigidos")
    if not dados_corrigidos:
        print("❌ Falha na extração dos dados corrigidos")
        return
    
    # Carrega constantes do arquivo corrigido
    constantes = extrair_constantes_calculo(arquivo_corrigido)
    if not constantes:
        print("❌ Falha na extração das constantes")
        return
    
    print(f"\n✅ Dados carregados com sucesso:")
    print(f"   Pontos originais: {len(dados_originais)}")
    print(f"   Pontos corrigidos: {len(dados_corrigidos)}")
    
    # Ajusta tempos para vazão média
    resultados_ajuste = ajustar_tempos_para_vazao_media(dados_originais, dados_corrigidos, constantes)
    
    # Gera JSON com resultados
    nome_arquivo_json = gerar_json_ajuste_vazao_media(resultados_ajuste)
    
    print(f"\n🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
    print(f"   ✅ Ajuste de tempos realizado para todos os pontos")
    print(f"   ✅ JSON gerado: {nome_arquivo_json}")
    print(f"   ✅ Incremento utilizado: 0.001 segundos")
    print(f"   ✅ Base de tempo mantida: 240 segundos")

if __name__ == "__main__":
    main() 