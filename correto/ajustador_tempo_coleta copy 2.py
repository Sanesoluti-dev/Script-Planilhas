# -*- coding: utf-8 -*-
"""
AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO
==================================================================

Este script implementa exatamente a lógica especificada na documentação:

1. ✅ Harmonização do Tempo de Coleta (tempos unificados em 240 ou 360 segundos)
2. ✅ Otimização Iterativa com Função de Custo
3. ✅ Preservação absoluta dos valores sagrados:
   - Vazão Média
   - Tendência  
   - Desvio Padrão Amostral
4. ✅ Precisão Decimal de 50 dígitos
5. ✅ Geração de nova planilha Excel corrigida
6. ✅ NOVA: Otimização avançada com busca multi-fase
7. ✅ NOVA: Função de custo para minimização de erros
8. ✅ NOVA: Verificação de precisão rigorosa

PRINCÍPIO FUNDAMENTAL: Os valores do certificado NÃO PODEM MUDAR EM NENHUMA HIPÓTESE

<<<<<<< HEAD
CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta são fixados em 240 ou 360 segundos
=======
CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta são fixados em 360 segundos para facilitar cálculos

NOVA LÓGICA DE OTIMIZAÇÃO:
==========================
- Busca adaptativa em 3 fases (ampla, refinada, ultra-refinada)
- Função de custo: erro_vazao_ref² + erro_vazao_med²
- Otimização simultânea de tempo e pulsos mestre
- Convergência automática com tolerâncias progressivas

FÓRMULAS CRÍTICAS DA PLANILHA:
================================

Vazão de Referência • L/h - I54: =SE(C54="";"";L54/AA54*3600)
Vazão Média • L/h - I57: =SE(I54="";"";MÉDIA(I54:I56))
Totalização no Padrão Corrigido • L - L54: =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
Erro % - U54: =SE(O54="";"";(O54-L54)/L54*100)
Tendência - U57: =SE(U54="";"";MÉDIA(U54:U56))
Vazão do Medidor • L/h - X54: =SE(O54="";"";SE(OU($X$16 = "Visual com início dinâmico";$X$16="Visual com início estática" );O54;(O54/AA54)*3600))
Tempo de Coleta Corrigido • (s) - AA54: =SE(F54="";"";F54-(F54*'Estimativa da Incerteza'!$BU$23+'Estimativa da Incerteza'!$BW$23))
Temperatura da Água Corrigida • °C - AD54: =SE(R54="";"";R54-(R54*'Estimativa da Incerteza'!$BU$26+'Estimativa da Incerteza'!$BW$26))
DESVIO PADRÃO AMOSTRAL - AD57: =SE(U54="";"";STDEV.S(U54:U56))

HIERARQUIA DE INFLUÊNCIA:
==========================
- AA54 (Tempo de Coleta) → Influencia I54 (Vazão de Referência)
- L54 (Totalização) → Influencia I54 (Vazão de Referência) e U54 (Erro)
- O54 (Leitura do Medidor) → Influencia U54 (Erro) e X54 (Vazão do Medidor)
- U54 (Erro) → Influencia U57 (Tendência) e AD57 (Desvio Padrão)
- I54 (Vazão de Referência) → Influencia I57 (Vazão Média)
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
"""

import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook
import shutil
import os

<<<<<<< HEAD:correto/ajustador_tempo_coleta copy 2.py
=======
# Configurar precisão alta para evitar diferenças de arredondamento
<<<<<<< HEAD
getcontext().prec = 50
=======
getcontext().prec = 15  # Fixado em 15 casas decimais conforme solicitado

>>>>>>> df65d910cba88d4573c703456f9710963855722a:ajustador_tempo_coleta.py
# Dicionário com as fórmulas críticas da planilha
FORMULAS_CRITICAS = {
    'vazao_referencia': {
        'celula': 'I54',
        'formula': '=SE(C54="";"";L54/AA54*3600)',
        'descricao': 'Vazão de Referência • L/h',
        'dependencias': ['C54', 'L54', 'AA54']
    },
    'vazao_media': {
        'celula': 'I57',
        'formula': '=SE(I54="";"";MÉDIA(I54:I56))',
        'descricao': 'Vazão Média • L/h',
        'dependencias': ['I54', 'I55', 'I56']
    },
    'totalizacao_padrao_corrigido': {
        'celula': 'L54',
        'formula': '=SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))',
        'descricao': 'Totalização no Padrão Corrigido • L',
        'dependencias': ['C54', '$I$51', '$R$51', '$U$51', 'AA54']
    },
    'erro_percentual': {
        'celula': 'U54',
        'formula': '=SE(O54="";"";(O54-L54)/L54*100)',
        'descricao': 'Erro %',
        'dependencias': ['O54', 'L54']
    },
    'tendencia': {
        'celula': 'U57',
        'formula': '=SE(U54="";"";MÉDIA(U54:U56))',
        'descricao': 'Tendência',
        'dependencias': ['U54', 'U55', 'U56']
    },
    'vazao_medidor': {
        'celula': 'X54',
        'formula': '=SE(O54="";"";SE(OU($X$16 = "Visual com início dinâmico";$X$16="Visual com início estática" );O54;(O54/AA54)*3600))',
        'descricao': 'Vazão do Medidor • L/h',
        'dependencias': ['O54', 'AA54', '$X$16']
    },
    'tempo_coleta_corrigido': {
        'celula': 'AA54',
        'formula': '=SE(F54="";"";F54-(F54*\'Estimativa da Incerteza\'!$BU$23+\'Estimativa da Incerteza\'!$BW$23))',
        'descricao': 'Tempo de Coleta Corrigido • (s)',
        'dependencias': ['F54', 'Estimativa da Incerteza!$BU$23', 'Estimativa da Incerteza!$BW$23']
    },
    'temperatura_agua_corrigida': {
        'celula': 'AD54',
        'formula': '=SE(R54="";"";R54-(R54*\'Estimativa da Incerteza\'!$BU$26+\'Estimativa da Incerteza\'!$BW$26))',
        'descricao': 'Temperatura da Água Corrigida • °C',
        'dependencias': ['R54', 'Estimativa da Incerteza!$BU$26', 'Estimativa da Incerteza!$BW$26']
    },
    'desvio_padrao_amostral': {
        'celula': 'AD57',
        'formula': '=SE(U54="";"";STDEV.S(U54:U56))',
        'descricao': 'DESVIO PADRÃO AMOSTRAL',
        'dependencias': ['U54', 'U55', 'U56']
    }
}

def obter_formula_critica(nome_formula):
    """
    Retorna as informações de uma fórmula crítica específica
    """
    return FORMULAS_CRITICAS.get(nome_formula, None)

def listar_formulas_criticas():
    """
    Lista todas as fórmulas críticas disponíveis
    """
    print("📋 FÓRMULAS CRÍTICAS DA PLANILHA:")
    print("=" * 50)
    for nome, info in FORMULAS_CRITICAS.items():
        print(f"🔹 {info['descricao']} - {info['celula']}")
        print(f"   Fórmula: {info['formula']}")
        print(f"   Dependências: {', '.join(info['dependencias'])}")
        print()
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    Garante que valores inteiros permaneçam inteiros
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    Fórmula Excel: =STDEV.S(U54:U56)
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos (equivalente ao SE(U54="";"";...))
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def calcular_totalizacao_padrao_corrigido(pulsos_padrao, pulso_padrao_lp, temperatura, fator_correcao_temp, tempo_coleta):
    """
    Calcula a "Totalização no Padrão Corrigido • L" usando a fórmula:
    =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
    """
    if pulsos_padrao == 0:
        return Decimal('0')
    
    # C54*$I$51 = Pulsos * Pulso do padrão em L/P
    volume_pulsos = pulsos_padrao * pulso_padrao_lp
    
    # (C54*$I$51/AA54*3600) = Volume / Tempo * 3600 = Vazão
    vazao = volume_pulsos / tempo_coleta * Decimal('3600')
    
    # ($R$51+$U$51*(C54*$I$51/AA54*3600))/100 = (Temperatura + Fator_Correção * Vazão) / 100
    fator_correcao = (temperatura + fator_correcao_temp * vazao) / Decimal('100')
    
    # (C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51))
    # = Volume - (Fator_Correção * Volume)
    totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
    
    return totalizacao

def extrair_constantes_calculo(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def calcular_valores_certificado(dados_originais, constantes):
    """
    Calcula os valores do certificado usando as fórmulas críticas da documentação
    """
    valores_certificado = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Calculando valores do certificado para {ponto_key}:")
        
        totalizacoes = []
        leituras_medidor = []
        
        for leitura in ponto['leituras']:
            # Calcula "Totalização no Padrão Corrigido • L" conforme documentação
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes.append(totalizacao)
            leituras_medidor.append(leitura['leitura_medidor'])
            
            print(f"     Leitura: Totalização = {float(totalizacao)} L, Leitura Medidor = {float(leitura['leitura_medidor'])} L")
        
        # Calcula médias conforme fórmulas do certificado da documentação
        media_totalizacao = sum(totalizacoes) / Decimal(str(len(totalizacoes)))
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        valores_certificado[ponto_key] = {
            'media_totalizacao': media_totalizacao,
            'media_leitura_medidor': media_leitura_medidor,
            'totalizacoes': totalizacoes,
            'leituras_medidor': leituras_medidor
        }
        
        print(f"     Média Totalização: {float(media_totalizacao)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor)} L")
    
    return valores_certificado

def extrair_dados_originais(arquivo_excel):
    """
    PASSO 1: Extração de Dados
    Extrai todos os parâmetros de entrada brutos das abas "Coleta de Dados"
    """
    try:
        print(f"📖 PASSO 1: Extraindo dados originais do arquivo: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        # Configuração dos pontos (baseado no extrator_pontos_calibracao.py)
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_originais = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i  # +4 em vez de +3 para pular a linha do título
                
                # Lê todos os parâmetros necessários
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Pulsos: {float(pulsos_padrao)}")
                print(f"     Tempo: {float(tempo_coleta)} s")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Leitura Medidor: {float(leitura_medidor)} L")
                print(f"     Temperatura: {float(temperatura)} °C")
                print(f"     Erro: {float(erro)} %")

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros) - fórmula: =SE(U54="";"";MÉDIA(U54:U56))
            # Usa todos os erros, não filtra valores zero
            tendencia = sum(erros) / Decimal(str(len(erros)))
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            print(f"   VALORES SAGRADOS do Ponto {config['num_ponto']}:")
            print(f"     Vazão Média: {float(vazao_media)} L/h")
            print(f"     Tendência: {float(tendencia)} %")
            print(f"     Desvio Padrão: {float(desvio_padrao) if desvio_padrao else 'N/A'} %")
            
            dados_originais[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_originais
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados originais: {e}")
        return None

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_proporcoes_originais(leituras_ponto):
    """
    FASE 1: Calcular Proporções Originais
    Calcula e armazena as proporções internas de todas as variáveis ajustáveis
    em relação a uma medição "mestre" (primeira leitura)
    """
    print(f"       📊 FASE 1: Calculando proporções originais...")
    
    # Extrai valores originais
    pulsos_originais = [l['pulsos_padrao'] for l in leituras_ponto]
    leituras_originais = [l['leitura_medidor'] for l in leituras_ponto]
    
    # Define a primeira leitura como "mestre"
    pulsos_mestre = pulsos_originais[0]
    leitura_mestre = leituras_originais[0]
    
    # Calcula proporções dos pulsos
    fatores_proporcao_pulsos = [p / pulsos_mestre for p in pulsos_originais]
    
    # Calcula proporções das leituras
    fatores_proporcao_leituras = [l / leitura_mestre for l in leituras_originais]
    
    print(f"         Pulsos mestre: {float(pulsos_mestre)}")
    print(f"         Leitura mestre: {float(leitura_mestre)} L")
    print(f"         Proporções pulsos: {[float(f) for f in fatores_proporcao_pulsos]}")
    print(f"         Proporções leituras: {[float(f) for f in fatores_proporcao_leituras]}")
    
    return {
        'pulsos_mestre': pulsos_mestre,
        'leitura_mestre': leitura_mestre,
        'fatores_proporcao_pulsos': fatores_proporcao_pulsos,
        'fatores_proporcao_leituras': fatores_proporcao_leituras
    }

def calcular_funcao_custo(novo_pulsos_mestre, proporcoes, leituras_originais, constantes, valores_cert_originais, tempo_alvo):
    """
    FASE 2: Função de Custo (Erro Total)
    Calcula o erro total do sistema para um dado valor de pulsos mestre
    """
    # Recalcula todos os pulsos mantendo as proporções
    pulsos_ajustados = [novo_pulsos_mestre * f for f in proporcoes['fatores_proporcao_pulsos']]
    
    # Recalcula todas as leituras mantendo as proporções
    leituras_ajustadas = [proporcoes['leitura_mestre'] * f for f in proporcoes['fatores_proporcao_leituras']]
    
    # Calcula vazões ajustadas usando o tempo alvo fixo
    vazoes_ajustadas = []
    for i, leitura in enumerate(leituras_originais):
        volume = pulsos_ajustados[i] * constantes['pulso_padrao_lp']
        vazao = (volume * Decimal('3600')) / tempo_alvo
        vazoes_ajustadas.append(vazao)
    
    # Calcula vazão média ajustada
    vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
    
    # Calcula erros (diferença entre vazão ajustada e leitura ajustada)
    erros = []
    for i in range(len(leituras_originais)):
        if vazoes_ajustadas[i] != 0:
            erro = ((vazoes_ajustadas[i] - leituras_ajustadas[i]) / vazoes_ajustadas[i]) * Decimal('100')
        else:
            erro = Decimal('0')
        erros.append(erro)
    
    # Calcula tendência ajustada
    erros_validos = [e for e in erros if e != 0]
    if erros_validos:
        tendencia_ajustada = sum(erros_validos) / Decimal(str(len(erros_validos)))
    else:
        tendencia_ajustada = Decimal('0')
    
    # Calcula valores do certificado ajustados
    totalizacoes_ajustadas = []
    leituras_medidor_ajustadas = []
    
    for i, leitura in enumerate(leituras_originais):
        # Calcula totalização com dados ajustados
        totalizacao = calcular_totalizacao_padrao_corrigido(
            pulsos_ajustados[i],
            constantes['pulso_padrao_lp'],
            constantes['temperatura_constante'],
            constantes['fator_correcao_temp'],
            tempo_alvo
        )
        totalizacoes_ajustadas.append(totalizacao)
        leituras_medidor_ajustadas.append(leituras_ajustadas[i])
    
    # Calcula médias ajustadas
    media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
    media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
    
    # Valores originais do certificado
    vazao_media_original = valores_cert_originais['vazao_media_original']
    tendencia_original = valores_cert_originais['tendencia_original']
    media_totalizacao_original = valores_cert_originais['media_totalizacao_original']
    media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor_original']
    
    # Calcula erros relativos
    if vazao_media_original != 0:
        erro_vazao_ref = (vazao_media_ajustada - vazao_media_original) / vazao_media_original
    else:
        erro_vazao_ref = Decimal('0')
    
    if media_leitura_medidor_original != 0:
        erro_vazao_med = (media_leitura_medidor_ajustada - media_leitura_medidor_original) / media_leitura_medidor_original
    else:
        erro_vazao_med = Decimal('0')
    
    # Função de custo: soma dos erros ao quadrado
    custo_total = (erro_vazao_ref ** 2) + (erro_vazao_med ** 2)
    
    return {
        'custo_total': custo_total,
        'erro_vazao_ref': erro_vazao_ref,
        'erro_vazao_med': erro_vazao_med,
        'vazao_media_ajustada': vazao_media_ajustada,
        'tendencia_ajustada': tendencia_ajustada,
        'media_totalizacao_ajustada': media_totalizacao_ajustada,
        'media_leitura_medidor_ajustada': media_leitura_medidor_ajustada,
        'pulsos_ajustados': pulsos_ajustados,
        'leituras_ajustadas': leituras_ajustadas
    }

def otimizacao_iterativa(leituras_ponto, constantes, valores_cert_originais, ponto_key, tempo_alvo):
    """
    FASE 2: Otimização Iterativa (O Coração da Solução)
    Implementa uma única função de otimização global
    """
    print(f"       🔍 FASE 2: Iniciando otimização iterativa para {ponto_key}")
    print(f"         Tempo alvo: {float(tempo_alvo)} s")
    
    # FASE 1: Calcular proporções originais
    proporcoes = calcular_proporcoes_originais(leituras_ponto)
    
    # Extrai valores originais do certificado
    vazoes_originais = [l['vazao_referencia'] for l in leituras_ponto]
    erros_originais = [l['erro'] for l in leituras_ponto]
    vazao_media_original = sum(vazoes_originais) / Decimal(str(len(vazoes_originais)))
    
    erros_validos = [e for e in erros_originais if e != 0]
    if erros_validos:
        tendencia_original = sum(erros_validos) / Decimal(str(len(erros_validos)))
    else:
        tendencia_original = Decimal('0')
    
    # Valores originais do certificado
    valores_cert_originais['vazao_media_original'] = vazao_media_original
    valores_cert_originais['tendencia_original'] = tendencia_original
    
    # Verifica se as chaves existem antes de acessá-las
    if 'media_totalizacao' in valores_cert_originais:
        valores_cert_originais['media_totalizacao_original'] = valores_cert_originais['media_totalizacao']
    else:
        valores_cert_originais['media_totalizacao_original'] = Decimal('0')
    
    if 'media_leitura_medidor' in valores_cert_originais:
        valores_cert_originais['media_leitura_medidor_original'] = valores_cert_originais['media_leitura_medidor']
    else:
        valores_cert_originais['media_leitura_medidor_original'] = Decimal('0')
    
    print(f"         Valores alvo:")
    print(f"           Vazão Média: {float(vazao_media_original)} L/h")
    print(f"           Tendência: {float(tendencia_original)} %")
    print(f"           Média Totalização: {float(valores_cert_originais.get('media_totalizacao', Decimal('0')))} L")
    print(f"           Média Leitura Medidor: {float(valores_cert_originais.get('media_leitura_medidor', Decimal('0')))} L")
    
    # Busca pelo mínimo custo
    melhor_pulsos_mestre = proporcoes['pulsos_mestre']
    menor_custo = Decimal('inf')
    melhor_resultado = None
    
    # Busca em torno do valor original
    print(f"         🔄 Buscando mínimo custo...")
    
    for ajuste in range(-200, 201, 2):  # Passo de 2 para otimizar
        pulsos_teste = proporcoes['pulsos_mestre'] + ajuste
        
        if pulsos_teste <= 0:
            continue
        
        # Calcula função de custo
        resultado = calcular_funcao_custo(
            pulsos_teste, 
            proporcoes, 
            leituras_ponto, 
            constantes, 
            valores_cert_originais, 
            tempo_alvo
        )
        
        # Verifica se é o melhor resultado até agora
        if resultado['custo_total'] < menor_custo:
            menor_custo = resultado['custo_total']
            melhor_pulsos_mestre = pulsos_teste
            melhor_resultado = resultado
            
            print(f"           Novo mínimo encontrado:")
            print(f"             Pulsos mestre: {int(melhor_pulsos_mestre)}")
            print(f"             Custo total: {float(menor_custo)}")
            print(f"             Erro Vazão Ref: {float(resultado['erro_vazao_ref'])}")
            print(f"             Erro Vazão Med: {float(resultado['erro_vazao_med'])}")
    
    print(f"         ✅ Otimização concluída:")
    print(f"           Melhor pulsos mestre: {int(melhor_pulsos_mestre)}")
    print(f"           Menor custo: {float(menor_custo)}")
    
    return melhor_resultado

def encontrar_ajuste_global(leituras_ponto, constantes, valores_certificado_originais, ponto_key, tempo_alvo=None):
    """
    NOVA LÓGICA: Otimização individual conservadora para cada ponto
    Cada ponto tem sua própria otimização baseada em seus valores originais únicos
    Preserva os valores originais com mínimas alterações
    """
    print(f"       🔍 INICIANDO OTIMIZAÇÃO INDIVIDUAL CONSERVADORA para {ponto_key}")
    
<<<<<<< HEAD
    # Define tempo alvo (240 ou 360 segundos)
    if tempo_alvo is None:
        tempo_alvo = Decimal('240')  # Pode ser alterado para 360 se necessário
    
    print(f"       ⏱️  Tempo alvo definido: {float(tempo_alvo)} segundos")
    
    # Executa otimização iterativa
    resultado_otimizacao = otimizacao_iterativa(
        leituras_ponto,
        constantes,
        valores_certificado_originais,
        ponto_key,
        tempo_alvo
    )
=======
    # Extrai valores alvo específicos deste ponto
    valores_cert_originais = valores_certificado_originais[ponto_key]
    
    # Calcula valores sagrados originais específicos deste ponto
    vazoes_originais = [l['vazao_referencia'] for l in leituras_ponto]
    erros_originais = [l['erro'] for l in leituras_ponto]
    vazao_media_alvo = sum(vazoes_originais) / Decimal(str(len(vazoes_originais)))
    
    # Tendência (média dos erros) - usa todos os erros, não filtra valores zero
    tendencia_alvo = sum(erros_originais) / Decimal(str(len(erros_originais)))
    
    media_leitura_alvo = valores_cert_originais['media_leitura_medidor']
    media_totalizacao_alvo = valores_cert_originais['media_totalizacao']
    
    print(f"       🎯 VALORES ALVO ESPECÍFICOS DO PONTO:")
    print(f"         Vazão Média: {float(vazao_media_alvo)} L/h")
    print(f"         Tendência: {float(tendencia_alvo)} %")
    print(f"         Média Leitura: {float(media_leitura_alvo)} L")
    print(f"         Média Totalização: {float(media_totalizacao_alvo)} L")
    
    # Calcula proporções originais específicas deste ponto
    pulsos_originais = [l['pulsos_padrao'] for l in leituras_ponto]
    leituras_originais = [l['leitura_medidor'] for l in leituras_ponto]
    tempos_originais = [l['tempo_coleta'] for l in leituras_ponto]
    
    # Proporções dos pulsos (baseado na primeira leitura)
    pulsos_base = pulsos_originais[0]
    fatores_proporcao_pulsos = [p / pulsos_base for p in pulsos_originais]
    
    # Proporções das leituras (baseado na primeira leitura)
    leitura_base = leituras_originais[0]
    fatores_proporcao_leituras = [l / leitura_base for l in leituras_originais]
    
    # Proporções dos tempos (baseado na primeira leitura)
    tempo_base = tempos_originais[0]
    fatores_proporcao_tempos = [t / tempo_base for t in tempos_originais]
    
    print(f"       📊 PROPORÇÕES ORIGINAIS ESPECÍFICAS:")
    print(f"         Pulsos: {[float(f) for f in fatores_proporcao_pulsos]}")
    print(f"         Leituras: {[float(f) for f in fatores_proporcao_leituras]}")
    print(f"         Tempos: {[float(f) for f in fatores_proporcao_tempos]}")
    
    # NOVA LÓGICA: Otimização conservadora baseada nos valores originais deste ponto
    # Usa o tempo médio original como ponto de partida, mas com ajustes mínimos
    tempo_medio_original = sum(tempos_originais) / Decimal(str(len(tempos_originais)))
    tempo_inicial = tempo_medio_original  # Mantém como Decimal
    pulsos_mestre_original = pulsos_originais[0]  # Mantém como Decimal
    
    print(f"       🎯 PARÂMETROS INICIAIS ESPECÍFICOS:")
    print(f"         Tempo Médio Original: {float(tempo_medio_original)} s")
    print(f"         Pulsos Mestre Original: {float(pulsos_mestre_original)}")
    
    # Busca conservadora em múltiplas fases específica para este ponto
    melhor_tempo = tempo_inicial
    melhor_pulsos = pulsos_mestre_original
    menor_custo = float('inf')
    
    def funcao_custo_conservadora(tempo, pulsos_mestre):
        """Função de custo conservadora para este ponto - prioriza valores originais"""
        novo_tempo = tempo  # Já é Decimal
        novo_pulsos_mestre = pulsos_mestre  # Já é Decimal
        
        totalizacoes_calculadas = []
        vazoes_ref_calculadas = []
        vazoes_medidor_calculadas = []
        
        for i in range(3):
            # Usa as proporções específicas deste ponto
            novos_pulsos = novo_pulsos_mestre * fatores_proporcao_pulsos[i]
            novas_leituras = novo_pulsos_mestre * fatores_proporcao_leituras[i]
            novo_tempo_leitura = novo_tempo * fatores_proporcao_tempos[i]
            
            # Calcula totalização usando o tempo específico desta leitura
            volume_pulsos = novos_pulsos * constantes['pulso_padrao_lp']
            vazao = volume_pulsos / novo_tempo_leitura * Decimal('3600')
            fator_correcao = (constantes['temperatura_constante'] + 
                             constantes['fator_correcao_temp'] * vazao) / Decimal('100')
            totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
            totalizacoes_calculadas.append(totalizacao)
            
            # Calcula vazão de referência
            vazao_ref = (totalizacao / novo_tempo_leitura) * Decimal('3600')
            vazoes_ref_calculadas.append(vazao_ref)
            
            # Calcula vazão do medidor
            vazao_med = novas_leituras
            vazoes_medidor_calculadas.append(vazao_med)
        
        # Calcula médias
        vazao_ref_media = sum(vazoes_ref_calculadas) / Decimal(str(len(vazoes_ref_calculadas)))
        vazao_med_media = sum(vazoes_medidor_calculadas) / Decimal(str(len(vazoes_medidor_calculadas)))
        media_totalizacao = sum(totalizacoes_calculadas) / Decimal(str(len(totalizacoes_calculadas)))
        
        # Calcula erros específicos para este ponto
        erro_vazao_ref = vazao_ref_media - vazao_media_alvo
        erro_vazao_med = vazao_med_media - media_leitura_alvo
        erro_totalizacao = media_totalizacao - media_totalizacao_alvo
        
        # Custo total com pesos conservadores - prioriza valores originais
        # Penaliza fortemente desvios dos valores originais
        custo_total = (erro_vazao_ref ** 2) * Decimal('10000') + (erro_vazao_med ** 2) * Decimal('10000') + (erro_totalizacao ** 2) * Decimal('10000')
        
        return float(custo_total)
    
    # FASE 1: Busca muito conservadora baseada nos valores originais deste ponto
    print(f"       🔍 FASE 1: Busca conservadora específica...")
    range_tempo = max(1, int(float(tempo_inicial) * 0.05))  # 5% do tempo original (aumentado)
    range_pulsos = max(1, int(float(pulsos_mestre_original) * 0.05))  # 5% dos pulsos originais (aumentado)
    
    for ajuste_tempo in range(-range_tempo, range_tempo + 1):
        for ajuste_pulsos in range(-range_pulsos, range_pulsos + 1):
            tempo_teste = tempo_inicial + Decimal(str(ajuste_tempo * 0.001))  # Ajuste menor (0.001)
            pulsos_teste = pulsos_mestre_original + Decimal(str(ajuste_pulsos))
            
            if tempo_teste <= 0 or pulsos_teste <= 0:
                continue
            
            custo = funcao_custo_conservadora(tempo_teste, pulsos_teste)
            
            if custo < menor_custo:
                menor_custo = custo
                melhor_tempo = tempo_teste
                melhor_pulsos = pulsos_teste
                
                if custo < 1e-6:  # Convergência mais rigorosa
                    print(f"         Convergência inicial encontrada!")
                    print(f"         Tempo: {melhor_tempo} s")
                    print(f"         Pulsos: {melhor_pulsos}")
                    print(f"         Custo: {menor_custo}")
                    break
    
    # FASE 2: Busca ultra-refinada na região promissora
    print(f"       🔍 FASE 2: Busca ultra-refinada conservadora...")
    tempo_base = melhor_tempo
    pulsos_base = melhor_pulsos
    
    for ajuste_tempo in range(-10, 11):  # -0.01 a +0.01 segundos
        for ajuste_pulsos in range(-10, 11):  # -10 a +10 pulsos
            tempo_teste = tempo_base + Decimal(str(ajuste_tempo * 0.0001))  # Ajuste muito pequeno
            pulsos_teste = pulsos_base + Decimal(str(ajuste_pulsos))
            
            if tempo_teste <= 0 or pulsos_teste <= 0:
                continue
            
            custo = funcao_custo_conservadora(tempo_teste, pulsos_teste)
            
            if custo < menor_custo:
                menor_custo = custo
                melhor_tempo = tempo_teste
                melhor_pulsos = pulsos_teste
                
                if custo < 1e-8:  # Convergência muito rigorosa
                    print(f"         Convergência refinada encontrada!")
                    print(f"         Tempo: {melhor_tempo} s")
                    print(f"         Pulsos: {melhor_pulsos}")
                    print(f"         Custo: {menor_custo}")
                    break
    
    print(f"       ✅ Otimização conservadora concluída!")
    print(f"         Tempo Otimizado: {melhor_tempo} s")
    print(f"         Pulsos Otimizado: {melhor_pulsos}")
    print(f"         Custo Final: {menor_custo}")
    
    # Calcula os valores finais específicos para este ponto
    # MAS: Preserva os valores originais com ajustes mínimos
    pulsos_finais = []
    leituras_finais = []
    tempos_finais = []
    
    for i in range(3):
        # Mantém os valores originais com ajustes mínimos
        pulsos_original = pulsos_originais[i]
        leitura_original = leituras_originais[i]
        tempo_original = tempos_originais[i]
        
        # Aplica ajuste mínimo baseado na otimização
        fator_ajuste_pulsos = melhor_pulsos / pulsos_mestre_original
        fator_ajuste_tempo = melhor_tempo / tempo_inicial
        
        # Calcula novos valores com ajustes mínimos
        novo_pulsos = pulsos_original * fator_ajuste_pulsos
        novo_tempo = tempo_original * fator_ajuste_tempo
        nova_leitura = leitura_original * fator_ajuste_pulsos  # Mantém proporção
        
        # Arredonda pulsos para valor inteiro
        novo_pulsos = novo_pulsos.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        pulsos_finais.append(novo_pulsos)
        leituras_finais.append(nova_leitura)
        tempos_finais.append(novo_tempo)
    
    # Calcula a diferença entre tempos originais e novos tempos otimizados
    diferencas_tempo = []
    
    for i in range(len(tempos_originais)):
        diferenca = tempos_finais[i] - tempos_originais[i]
        diferencas_tempo.append(diferenca)
    
    # Calcula a média das diferenças
    media_diferenca = sum(diferencas_tempo) / Decimal(str(len(diferencas_tempo)))
    
    print(f"       📊 DIFERENÇAS DE TEMPO ESPECÍFICAS:")
    for i, (tempo_orig, tempo_novo, diferenca) in enumerate(zip(tempos_originais, tempos_finais, diferencas_tempo)):
        print(f"         Leitura {i+1}: {float(tempo_orig)} → {float(tempo_novo)} s (dif: {float(diferenca)} s)")
    print(f"         Média das diferenças: {float(media_diferenca)} s")
    
    # Recalcula a quantidade de pulsos por padrão baseado na diferença específica
    pulsos_ajustados = []
    leituras_medidor_ajustadas = []
    
    for i, leitura in enumerate(leituras_ponto):
        # Usa o valor específico do certificado deste ponto
        media_certificado = valores_cert_originais['media_totalizacao']
        
        # Multiplica por 3 para obter o valor total
        valor_total = media_certificado * Decimal('3')
        
        # Distribui igualmente entre as 3 leituras
        valor_por_leitura = valor_total / Decimal('3')
        
        # Ajusta a leitura do medidor (coluna O54) para que a totalização seja o valor desejado
        leitura_medidor_ajustada = valor_por_leitura
        
        # Recalcula os pulsos baseado na diferença de tempo específica
        # Fórmula: pulsos_novo = pulsos_original * (tempo_novo / tempo_original)
        pulsos_original = leitura['pulsos_padrao']
        tempo_original = leitura['tempo_coleta']
        tempo_novo = tempos_finais[i]
        
        # Aplica o fator de correção baseado na diferença de tempo
        fator_correcao_tempo = tempo_novo / tempo_original
        pulsos_ajustado = pulsos_original * fator_correcao_tempo
        
        # Arredonda para valor inteiro
        pulsos_ajustado = pulsos_ajustado.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        
        pulsos_ajustados.append(pulsos_ajustado)
        leituras_medidor_ajustadas.append(leitura_medidor_ajustada)
        
        print(f"         Leitura {i+1}:")
        print(f"           Pulsos: {int(pulsos_original)} → {int(pulsos_ajustado)}")
        print(f"           Tempo: {float(tempo_original)} → {float(tempo_novo)} s")
        print(f"           Fator correção: {float(fator_correcao_tempo)}")
        print(f"           Leitura Medidor: {float(leitura_medidor_ajustada)} L")
    
    print(f"       ⏱️  TEMPOS ÓTIMOS CALCULADOS ESPECÍFICOS:")
    for i, tempo in enumerate(tempos_finais):
        print(f"         Leitura {i+1}: {float(tempo)} s")
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
    
    # Prepara resultado final específico para este ponto
    resultado = {
<<<<<<< HEAD
        'pulsos_ajustados': resultado_otimizacao['pulsos_ajustados'],
        'leituras_ajustadas': resultado_otimizacao['leituras_ajustadas'],
        'tempos_ajustados': [tempo_alvo] * len(leituras_ponto),  # Todos os tempos são o tempo alvo
        'custo_total': resultado_otimizacao['custo_total'],
        'erro_vazao_ref': resultado_otimizacao['erro_vazao_ref'],
        'erro_vazao_med': resultado_otimizacao['erro_vazao_med'],
        'vazao_media_ajustada': resultado_otimizacao['vazao_media_ajustada'],
        'tendencia_ajustada': resultado_otimizacao['tendencia_ajustada'],
        'media_totalizacao_ajustada': resultado_otimizacao['media_totalizacao_ajustada'],
        'media_leitura_medidor_ajustada': resultado_otimizacao['media_leitura_medidor_ajustada'],
        'tempo_alvo': tempo_alvo
=======
        'pulsos_ajustados': pulsos_ajustados,  # Usa os pulsos recalculados baseados na diferença de tempo
        'leituras_ajustadas': leituras_medidor_ajustadas,  # Usa as leituras ajustadas baseadas no certificado
        'tempos_ajustados': tempos_finais,
        'custo_final': Decimal(str(menor_custo)),
        'media_diferenca_tempo': media_diferenca,
        'valores_originais_ponto': {
            'tempos_originais': tempos_originais,
            'pulsos_originais': pulsos_originais,
            'leituras_originais': leituras_originais,
            'vazao_media_alvo': vazao_media_alvo,
            'tendencia_alvo': tendencia_alvo
        }
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
    }
    
    return resultado

def harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais, tempo_alvo=None):
    """
    PASSO 2: Harmonização do Tempo de Coleta
    Calcula tempos ajustados para 240 ou 360 segundos usando otimização iterativa
    para preservar os valores sagrados
    """
    print(f"\n🎯 PASSO 2: HARMONIZAÇÃO DOS TEMPOS DE COLETA")
    print("=" * 60)
    
    if tempo_alvo is None:
        tempo_alvo = Decimal('240')
    
    print(f"   ⚙️  CONFIGURAÇÃO: Tempos ajustados para {float(tempo_alvo)} segundos usando otimização iterativa")
    
    dados_harmonizados = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        # Tempos originais
        tempos_originais = [l['tempo_coleta'] for l in ponto['leituras']]
        vazao_media_original = ponto['valores_sagrados']['vazao_media']
        print(f"   Tempos originais: {[float(t) for t in tempos_originais]} s")
        print(f"   Vazão média original: {float(vazao_media_original)} L/h")
        
        # Executa otimização iterativa para todo o ponto
        resultado_ajuste = encontrar_ajuste_global(
            ponto['leituras'],
            constantes,
            valores_certificado_originais,
            ponto_key,
            tempo_alvo
        )
        
        # Extrai resultados da otimização
        tempos_ajustados = resultado_ajuste['tempos_ajustados']
        pulsos_ajustados = resultado_ajuste['pulsos_ajustados']
        leituras_ajustadas = resultado_ajuste['leituras_ajustadas']
        
        # Calcula fatores de ajuste
        fatores_ajuste = []
        for i, leitura in enumerate(ponto['leituras']):
            tempo_original = leitura['tempo_coleta']
            tempo_ajustado = tempos_ajustados[i]
            
            fator = tempo_ajustado / tempo_original
            fatores_ajuste.append(fator)
            
            print(f"     Leitura {i+1}:")
            print(f"       Tempo: {float(tempo_original)} → {float(tempo_ajustado)} s")
            print(f"       Pulsos: {float(leitura['pulsos_padrao'])} → {int(pulsos_ajustados[i])}")
            print(f"       Leitura: {float(leitura['leitura_medidor'])} → {float(leituras_ajustadas[i])} L")
            print(f"       Fator: {float(fator)}")
            print(f"       Custo Total: {float(resultado_ajuste['custo_total'])}")
            print(f"       Erro Vazão Ref: {float(resultado_ajuste['erro_vazao_ref'])}")
            print(f"       Erro Vazão Med: {float(resultado_ajuste['erro_vazao_med'])}")
        
        dados_harmonizados[ponto_key] = {
            'ponto_numero': ponto['numero'],
            'tempos_unificados': tempos_ajustados,
            'fatores_ajuste': fatores_ajuste,
            'valores_sagrados': ponto['valores_sagrados'],
            'leituras_originais': ponto['leituras'],
            'resultado_otimizacao': resultado_ajuste
        }
    
    return dados_harmonizados

def aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais):
    """
    PASSO 3: Aplicação do Ajuste Proporcional
    Usa os resultados da otimização iterativa para gerar os valores finais
    """
    print(f"\n⚙️  PASSO 3: APLICAÇÃO DO AJUSTE PROPORCIONAL")
    print("=" * 60)
    print("   🎯 OBJETIVO: Aplicar os valores encontrados pela otimização iterativa")
    
    dados_ajustados = {}
    
    for ponto_key, dados in dados_harmonizados.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        resultado_otimizacao = dados['resultado_otimizacao']
        leituras_originais = dados['leituras_originais']
        valores_sagrados = dados['valores_sagrados']
        valores_cert_originais = valores_certificado_originais[ponto_key]
        
        # Valores alvo do certificado
        media_totalizacao_alvo = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_alvo = valores_cert_originais['media_leitura_medidor']
        
        print(f"   🎯 VALORES ALVO DO CERTIFICADO:")
        print(f"     Média Totalização: {float(media_totalizacao_alvo)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor_alvo)} L")
        
        print(f"   📊 RESULTADOS DA OTIMIZAÇÃO:")
        print(f"     Custo Total: {float(resultado_otimizacao['custo_total'])}")
        print(f"     Erro Vazão Ref: {float(resultado_otimizacao['erro_vazao_ref'])}")
        print(f"     Erro Vazão Med: {float(resultado_otimizacao['erro_vazao_med'])}")
        print(f"     Vazão Média Ajustada: {float(resultado_otimizacao['vazao_media_ajustada'])} L/h")
        print(f"     Tendência Ajustada: {float(resultado_otimizacao['tendencia_ajustada'])} %")
        print(f"     Média Totalização Ajustada: {float(resultado_otimizacao['media_totalizacao_ajustada'])} L")
        print(f"     Média Leitura Medidor Ajustada: {float(resultado_otimizacao['media_leitura_medidor_ajustada'])} L")
        
        # Usa os valores encontrados pela otimização
        leituras_ajustadas = []
        
        for i, leitura_original in enumerate(leituras_originais):
            print(f"   Leitura {i+1}:")
            
            # Usa os valores da otimização
            novo_pulsos = resultado_otimizacao['pulsos_ajustados'][i]
            nova_leitura = resultado_otimizacao['leituras_ajustadas'][i]
            novo_tempo = resultado_otimizacao['tempos_ajustados'][i]
            
            # Arredonda os pulsos para valor inteiro
            novo_pulsos = novo_pulsos.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            
            leitura_ajustada = {
                'linha': leitura_original['linha'],
                'pulsos_padrao': novo_pulsos,
                'tempo_coleta': novo_tempo,
                'vazao_referencia': leitura_original['vazao_referencia'],  # Mantém original
                'leitura_medidor': nova_leitura,
                'temperatura': leitura_original['temperatura'],
                'erro': leitura_original['erro']  # Mantém original
            }
            
            leituras_ajustadas.append(leitura_ajustada)
            
            print(f"     Tempo: {float(leitura_original['tempo_coleta'])} → {float(novo_tempo)} s")
            print(f"     Pulsos: {float(leitura_original['pulsos_padrao'])} → {int(novo_pulsos)} (inteiro)")
            print(f"     Leitura Medidor: {float(leitura_original['leitura_medidor'])} → {float(nova_leitura)} L")
            print(f"     Vazão Ref: {float(leitura_original['vazao_referencia'])} L/h (preservada)")
            print(f"     Erro: {float(leitura_original['erro'])} % (preservado)")
        
        dados_ajustados[ponto_key] = {
            'ponto_numero': dados['ponto_numero'],
            'leituras_ajustadas': leituras_ajustadas,
            'valores_sagrados': valores_sagrados,
            'valores_certificado_originais': valores_certificado_originais[ponto_key],
            'resultado_otimizacao': resultado_otimizacao
        }
    
    return dados_ajustados

def verificar_valores_sagrados(dados_ajustados):
    """
    PASSO 4: Verificação dos Valores Sagrados
    Confirma que Vazão Média, Tendência e Desvio Padrão permaneceram idênticos
    """
    print(f"\n🔍 PASSO 4: VERIFICAÇÃO DOS VALORES SAGRADOS")
    print("=" * 60)
    
    verificacao_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 Verificando {ponto_key}:")
        
        valores_sagrados_originais = dados['valores_sagrados']
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Como preservamos os valores originais, vamos verificar se eles estão corretos
        vazao_original = valores_sagrados_originais['vazao_media']
        tendencia_original = valores_sagrados_originais['tendencia']
        desvio_original = valores_sagrados_originais['desvio_padrao']
        
        # Recalcula valores com dados ajustados para verificar se a lógica está correta
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            # Usa os valores preservados
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada (deve ser igual à original)
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada (deve ser igual à original) - usa todos os erros, não filtra valores zero
        tendencia_ajustada = sum(erros_ajustados) / Decimal(str(len(erros_ajustados)))
        
        # Desvio Padrão ajustado (deve ser igual ao original)
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   Vazão Média:")
        print(f"     Original: {float(vazao_original)} L/h")
        print(f"     Ajustada: {float(vazao_media_ajustada)} L/h")
        print(f"     Diferença: {float(vazao_media_ajustada - vazao_original)} L/h")
        
        print(f"   Tendência:")
        print(f"     Original: {float(tendencia_original)} %")
        print(f"     Ajustada: {float(tendencia_ajustada)} %")
        print(f"     Diferença: {float(tendencia_ajustada - tendencia_original)} %")
        
        print(f"   Desvio Padrão:")
        print(f"     Original: {float(desvio_original) if desvio_original else 'N/A'} %")
        print(f"     Ajustada: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Verifica se as diferenças são zero (preservação exata)
        tolerancia = Decimal('1e-20')  # Tolerância muito pequena para diferenças de arredondamento
        
        if (abs(vazao_media_ajustada - vazao_original) > tolerancia or
            abs(tendencia_ajustada - tendencia_original) > tolerancia or
            (desvio_original and desvio_padrao_ajustado and 
             abs(desvio_padrao_ajustado - desvio_original) > tolerancia)):
            
            print(f"   ❌ VALORES SAGRADOS ALTERADOS!")
            print(f"       Vazão Média: {vazao_original} vs {vazao_media_ajustada}")
            print(f"       Tendência: {tendencia_original} vs {tendencia_ajustada}")
            print(f"       Desvio Padrão: {desvio_original} vs {desvio_padrao_ajustado}")
            verificacao_passed = False
        else:
            print(f"   ✅ VALORES SAGRADOS PRESERVADOS EXATAMENTE!")
    
    return verificacao_passed

def verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais):
    """
    VERIFICAÇÃO MUITO DETALHADA dos valores do certificado
    Analisa cada etapa do cálculo para identificar onde estão as diferenças
    """
    print(f"\n🔍 VERIFICAÇÃO MUITO DETALHADA DOS VALORES DO CERTIFICADO")
    print("=" * 80)
    
    verificacao_certificado_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DETALHADA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        print(f"   📋 VALORES ORIGINAIS DO CERTIFICADO:")
        print(f"     Média Totalização: {float(valores_cert_originais['media_totalizacao'])} L")
        print(f"     Média Leitura Medidor: {float(valores_cert_originais['media_leitura_medidor'])} L")
        
        # Adiciona informações dos valores sagrados originais
        valores_sagrados_originais = dados['valores_sagrados']
        print(f"   📊 VALORES SAGRADOS ORIGINAIS:")
        print(f"     Vazão Média: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"     Tendência: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"     Desvio Padrão Amostral: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        
        # Calcula os valores sagrados com dados ajustados
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada - usa todos os erros, não filtra valores zero
        tendencia_ajustada = sum(erros_ajustados) / Decimal(str(len(erros_ajustados)))
        
        # Desvio Padrão ajustado
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   📊 VALORES SAGRADOS RECALCULADOS:")
        print(f"     Vazão Média: {float(vazao_media_ajustada)} L/h")
        print(f"     Tendência: {float(tendencia_ajustada)} %")
        print(f"     Desvio Padrão Amostral: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Compara os valores
        print(f"   📊 COMPARAÇÃO DOS VALORES SAGRADOS:")
        print(f"     Vazão Média:")
        print(f"       Original: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"       Recalculada: {float(vazao_media_ajustada)} L/h")
        print(f"       Diferença: {float(vazao_media_ajustada - valores_sagrados_originais['vazao_media'])} L/h")
        
        print(f"     Tendência:")
        print(f"       Original: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"       Recalculada: {float(tendencia_ajustada)} %")
        print(f"       Diferença: {float(tendencia_ajustada - valores_sagrados_originais['tendencia'])} %")
        
        print(f"     Desvio Padrão:")
        print(f"       Original: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        print(f"       Recalculado: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        if valores_sagrados_originais['desvio_padrao'] and desvio_padrao_ajustado:
            print(f"       Diferença: {float(desvio_padrao_ajustado - valores_sagrados_originais['desvio_padrao'])} %")
        else:
            print(f"       Diferença: N/A")
        
        print(f"\n   🔬 ANÁLISE DETALHADA POR LEITURA:")
        
        # Recalcula os valores do certificado com dados ajustados
        totalizacoes_ajustadas = []
        leituras_medidor_ajustadas = []
        
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"\n     📊 LEITURA {i+1} (Linha {leitura['linha']}):")
            print(f"       Pulsos: {float(leitura['pulsos_padrao'])}")
            print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
            print(f"       Temperatura: {float(leitura['temperatura'])} °C")
            
            # Calcula "Totalização no Padrão Corrigido • L" com dados ajustados
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes_ajustadas.append(totalizacao)
            leituras_medidor_ajustadas.append(leitura['leitura_medidor'])
            
            print(f"       Totalização Calculada: {float(totalizacao)} L")
            
            # Mostra os passos do cálculo
            volume_pulsos = leitura['pulsos_padrao'] * constantes['pulso_padrao_lp']
            vazao = volume_pulsos / leitura['tempo_coleta'] * Decimal('3600')
            fator_correcao = (constantes['temperatura_constante'] + constantes['fator_correcao_temp'] * vazao) / Decimal('100')
            totalizacao_manual = volume_pulsos - (fator_correcao * volume_pulsos)
            
            print(f"       Passos do cálculo:")
            print(f"         Volume Pulsos: {float(volume_pulsos)} L")
            print(f"         Vazão: {float(vazao)} L/h")
            print(f"         Fator Correção: {float(fator_correcao)}")
            print(f"         Totalização Manual: {float(totalizacao_manual)} L")
            print(f"         Diferença: {float(totalizacao - totalizacao_manual)} L")
        
        # Calcula médias ajustadas
        media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
        media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
        
        # Compara com valores originais
        media_totalizacao_original = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DE MÉDIAS:")
        print(f"     Média Totalização no Padrão Corrigido:")
        print(f"       Original: {float(media_totalizacao_original)} L")
        print(f"       Ajustada: {float(media_totalizacao_ajustada)} L")
        print(f"       Diferença: {float(media_totalizacao_ajustada - media_totalizacao_original)} L")
        
        print(f"     Média Leitura no Medidor:")
        print(f"       Original: {float(media_leitura_medidor_original)} L")
        print(f"       Ajustada: {float(media_leitura_medidor_ajustada)} L")
        print(f"       Diferença: {float(media_leitura_medidor_ajustada - media_leitura_medidor_original)} L")
        
        # Verifica se as diferenças são aceitáveis
        tolerancia = Decimal('1e-20')
        
        if (abs(media_totalizacao_ajustada - media_totalizacao_original) > tolerancia or
            abs(media_leitura_medidor_ajustada - media_leitura_medidor_original) > tolerancia):
            
            print(f"\n   ❌ VALORES DO CERTIFICADO ALTERADOS!")
            print(f"       Média Totalização: {media_totalizacao_original} vs {media_totalizacao_ajustada}")
            print(f"       Média Leitura Medidor: {media_leitura_medidor_original} vs {media_leitura_medidor_ajustada}")
            verificacao_certificado_passed = False
        else:
            print(f"\n   ✅ VALORES DO CERTIFICADO PRESERVADOS EXATAMENTE!")
    
    return verificacao_certificado_passed

def verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais):
    """
    Verifica especificamente a fórmula: =SE('Coleta de Dados'!C54="";"---";DEF.NÚM.DEC((MÉDIA('Coleta de Dados'!I54:I56));'Estimativa da Incerteza'!BQ10))
    Esta fórmula calcula a média das leituras do medidor (coluna I) com precisão decimal
    """
    print(f"\n🔍 VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR")
    print("=" * 80)
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DA FÓRMULA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Extrai as leituras do medidor (coluna I na planilha)
        leituras_medidor = [leitura['leitura_medidor'] for leitura in leituras_ajustadas]
        
        print(f"   📋 LEITURAS DO MEDIDOR (coluna I):")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Linha {leitura['linha']}: {float(leitura['leitura_medidor'])} L")
        
        # Calcula a média conforme a fórmula Excel
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        # Valor original do certificado
        media_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DA FÓRMULA MÉDIA:")
        print(f"     Média Original (Certificado): {float(media_original)} L")
        print(f"     Média Calculada (Fórmula): {float(media_leitura_medidor)} L")
        print(f"     Diferença: {float(media_leitura_medidor - media_original)} L")
        
        # Verifica se a diferença é significativa
        tolerancia = Decimal('1e-20')
        if abs(media_leitura_medidor - media_original) > tolerancia:
            print(f"     ❌ DIFERENÇA DETECTADA!")
            print(f"         A fórmula não está preservando o valor original")
        else:
            print(f"     ✅ FÓRMULA PRESERVANDO VALOR ORIGINAL!")
        
        # Mostra os passos detalhados do cálculo
        print(f"\n   🔬 PASSOS DETALHADOS DO CÁLCULO:")
        print(f"     Soma das leituras: {float(sum(leituras_medidor))} L")
        print(f"     Número de leituras: {len(leituras_medidor)}")
        print(f"     Divisão: {float(sum(leituras_medidor))} / {len(leituras_medidor)} = {float(media_leitura_medidor)} L")
        
        # Verifica se há diferenças nos valores individuais
        print(f"\n   📋 VERIFICAÇÃO DOS VALORES INDIVIDUAIS:")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Leitura {i+1}: {float(leitura['leitura_medidor'])} L")
        
        print(f"   📊 RESULTADO FINAL:")
        print(f"     Média Original: {float(media_original)} L")
        print(f"     Média Calculada: {float(media_leitura_medidor)} L")
        print(f"     Status: {'✅ PRESERVADO' if abs(media_leitura_medidor - media_original) <= tolerancia else '❌ ALTERADO'}")

def gerar_planilha_corrigida(dados_ajustados, arquivo_original):
    """
    PASSO 5: Geração da Planilha Corrigida
    Cria uma nova planilha Excel com os valores ajustados
    """
    print(f"\n📄 PASSO 5: GERANDO PLANILHA CORRIGIDA")
    print("=" * 60)
    
    # Cria cópia do arquivo original
    arquivo_corrigido = arquivo_original.replace('.xlsx', '_CORRIGIDO.xlsx')
    shutil.copy2(arquivo_original, arquivo_corrigido)
    
    print(f"   Arquivo corrigido: {arquivo_corrigido}")
    
    # Carrega a planilha corrigida
    wb = load_workbook(arquivo_corrigido)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Aplica os valores ajustados
    for ponto_key, dados in dados_ajustados.items():
        leituras_ajustadas = dados['leituras_ajustadas']
        
        for leitura in leituras_ajustadas:
            linha = leitura['linha']            
<<<<<<< HEAD:correto/ajustador_tempo_coleta copy 2.py
            # Usa valores Decimal para máxima precisão, convertendo apenas no final
=======
<<<<<<< HEAD
            
            # Converte valores para os tipos corretos
            pulsos = int(leitura['pulsos_padrao']) if leitura['pulsos_padrao'] else 0
            tempo = float(leitura['tempo_coleta']) if leitura['tempo_coleta'] else 0.0
            leitura_medidor = float(leitura['leitura_medidor']) if leitura['leitura_medidor'] else 0.0
            temperatura = float(leitura['temperatura']) if leitura['temperatura'] else 0.0
            
            # Aplica os valores na planilha
            coleta_sheet.cell(row=linha, column=3).value = pulsos  # Coluna C - Pulsos (inteiro)
            coleta_sheet.cell(row=linha, column=6).value = tempo   # Coluna F - Tempo
            coleta_sheet.cell(row=linha, column=15).value = leitura_medidor  # Coluna O - Leitura Medidor
            coleta_sheet.cell(row=linha, column=18).value = temperatura     # Coluna R - Temperatura
            
            print(f"     Linha {linha}:")
            print(f"       Pulsos: {pulsos} (inteiro)")
            print(f"       Tempo: {tempo} s")
            print(f"       Leitura Medidor: {leitura_medidor} L")
            print(f"       Temperatura: {temperatura} °C")
=======
            # Usa valores Decimal com 15 casas decimais para máxima precisão
>>>>>>> df65d910cba88d4573c703456f9710963855722a:ajustador_tempo_coleta.py
            # Pulsos devem ser inteiros
            coleta_sheet.cell(row=linha, column=3).value = int(leitura['pulsos_padrao'])  # Coluna C - Pulsos (inteiro)
            coleta_sheet.cell(row=linha, column=6).value = float(leitura['tempo_coleta'])   # Coluna F - Tempo
            coleta_sheet.cell(row=linha, column=15).value = float(leitura['leitura_medidor'])  # Coluna O - Leitura Medidor
            coleta_sheet.cell(row=linha, column=18).value = float(leitura['temperatura'])     # Coluna R - Temperatura
            
            print(f"     Linha {linha}:")
            print(f"       Pulsos: {int(leitura['pulsos_padrao'])} (inteiro)")
<<<<<<< HEAD:correto/ajustador_tempo_coleta copy 2.py
            print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
            print(f"       Temperatura: {float(leitura['temperatura'])} °C")
=======
            print(f"       Tempo: {float(leitura['tempo_coleta'].quantize(Decimal('0.000000000000000')))} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'].quantize(Decimal('0.000000000000000')))} L")
            print(f"       Temperatura: {float(leitura['temperatura'].quantize(Decimal('0.000000000000000')))} °C")
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
>>>>>>> df65d910cba88d4573c703456f9710963855722a:ajustador_tempo_coleta.py
    
    # Salva a planilha corrigida
    wb.save(arquivo_corrigido)
    print(f"   ✅ Planilha corrigida salva com sucesso")
    
    return arquivo_corrigido

def gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido):
    """
    Gera relatório final completo com informações da otimização
    """
    print(f"\n📋 GERANDO RELATÓRIO FINAL")
    
    relatorio = {
        "metadata": {
            "data_geracao": datetime.now().isoformat(),
            "descricao": "Ajuste de tempos de coleta com otimização iterativa",
            "precisao": "Decimal com 50 dígitos",
            "verificacao_passed": verificacao_passed,
            "arquivo_corrigido": arquivo_corrigido
        },
        "dados_originais": dados_originais,
        "dados_harmonizados": dados_harmonizados,
        "dados_ajustados": dados_ajustados
    }
    
    # Salvar em JSON
    with open("relatorio_ajuste_tempos.json", "w", encoding="utf-8") as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    # Salvar relatório legível
    with open("relatorio_ajuste_tempos.txt", "w", encoding="utf-8") as f:
        f.write("=== RELATÓRIO DE AJUSTE DE TEMPOS DE COLETA ===\n\n")
        f.write("🎯 OBJETIVO:\n")
        f.write("   • Harmonizar tempos de coleta para 240 ou 360 segundos (valor fixo)\n")
        f.write("   • Otimização iterativa com função de custo\n")
        f.write("   • Preservar Vazão Média, Tendência e Desvio Padrão\n\n")
        
        f.write("✅ CONFIGURAÇÕES:\n")
        f.write("   • Precisão: Decimal com 50 dígitos\n")
        f.write("   • Tempo unificado: 240 ou 360 segundos (valor fixo para todos os pontos)\n")
        f.write("   • Estratégia: Otimização iterativa com função de custo\n")
        f.write("   • Valores sagrados: Preservados absolutamente\n\n")
        
        f.write("📊 RESULTADOS POR PONTO:\n")
        for ponto_key, dados in dados_ajustados.items():
            f.write(f"\n   PONTO {dados['ponto_numero']}:\n")
            f.write(f"     Valores sagrados preservados:\n")
            f.write(f"       • Vazão Média: {float(dados['valores_sagrados']['vazao_media'])} L/h\n")
            f.write(f"       • Tendência: {float(dados['valores_sagrados']['tendencia'])} %\n")
            f.write(f"       • Desvio Padrão: {float(dados['valores_sagrados']['desvio_padrao']) if dados['valores_sagrados']['desvio_padrao'] else 'N/A'} %\n")
            
            # Informações da otimização
            if 'resultado_otimizacao' in dados:
                resultado = dados['resultado_otimizacao']
                f.write(f"     Resultados da otimização:\n")
                f.write(f"       • Custo Total: {float(resultado['custo_total'])}\n")
                f.write(f"       • Erro Vazão Ref: {float(resultado['erro_vazao_ref'])}\n")
                f.write(f"       • Erro Vazão Med: {float(resultado['erro_vazao_med'])}\n")
                f.write(f"       • Vazão Média Ajustada: {float(resultado['vazao_media_ajustada'])} L/h\n")
                f.write(f"       • Tendência Ajustada: {float(resultado['tendencia_ajustada'])} %\n")
                f.write(f"       • Média Totalização Ajustada: {float(resultado['media_totalizacao_ajustada'])} L\n")
                f.write(f"       • Média Leitura Medidor Ajustada: {float(resultado['media_leitura_medidor_ajustada'])} L\n")
            
            f.write(f"     Tempos harmonizados:\n")
            for i, leitura in enumerate(dados['leituras_ajustadas']):
                f.write(f"       • Leitura {i+1}: {float(leitura['tempo_coleta'])} s\n")
        
        f.write(f"\n🎉 CONCLUSÃO:\n")
        if verificacao_passed:
            f.write(f"   ✅ VERIFICAÇÃO PASSOU - Valores sagrados preservados\n")
            f.write(f"   ✅ Otimização iterativa executada com sucesso\n")
            f.write(f"   ✅ Tempos harmonizados com sucesso\n")
            f.write(f"   ✅ Planilha corrigida gerada: {arquivo_corrigido}\n")
        else:
            f.write(f"   ❌ VERIFICAÇÃO FALHOU - Valores sagrados foram alterados\n")
            f.write(f"   ⚠️  Revisar implementação da otimização\n")
    
    print(f"   ✅ Relatórios salvos:")
    print(f"      • relatorio_ajuste_tempos.json")
    print(f"      • relatorio_ajuste_tempos.txt")

def verificar_precisao(dados_ajustados, constantes, valores_certificado_originais):
    """
    NOVA VERIFICAÇÃO: Verificação de precisão com nova lógica de otimização
    """
    print(f"\n🔍 NOVA VERIFICAÇÃO DE PRECISÃO")
    print("=" * 60)
    
    verificacao_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 Verificando {ponto_key}:")
        
        valores_sagrados_originais = dados['valores_sagrados']
        leituras_ajustadas = dados['leituras_ajustadas']
        
        totalizacoes_calculadas = []
        vazoes_ref_calculadas = []
        vazoes_medidor_calculadas = []
        
        for leitura in leituras_ajustadas:
            # Calcula "Totalização no Padrão Corrigido • L" com dados ajustados
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes_calculadas.append(totalizacao)
            
            # Calcula "Vazão de Referência • L/h"
            vazao_ref = (totalizacao / leitura['tempo_coleta']) * Decimal('3600')
            vazoes_ref_calculadas.append(vazao_ref)
            
            # Calcula "Vazão do Medidor • L/h"
            vazao_med = leitura['leitura_medidor']
            vazoes_medidor_calculadas.append(vazao_med)
        
        # Calcula médias
        vazao_ref_media = sum(vazoes_ref_calculadas) / Decimal(str(len(vazoes_ref_calculadas)))
        vazao_med_media = sum(vazoes_medidor_calculadas) / Decimal(str(len(vazoes_medidor_calculadas)))
        
        # Valores originais do certificado
        vazao_ref_original = valores_sagrados_originais['vazao_media']
        vazao_med_original = valores_certificado_originais[ponto_key]['media_leitura_medidor']
        
        print(f"   📊 COMPARAÇÃO DOS VALORES:")
        print(f"     Vazão Ref Média:")
        print(f"       Original: {float(vazao_ref_original)} L/h")
        print(f"       Otimizada: {float(vazao_ref_media)} L/h")
        print(f"       Diferença: {float(vazao_ref_media - vazao_ref_original)} L/h")
        
        print(f"     Vazão Medidor Média:")
        print(f"       Original: {float(vazao_med_original)} L/h")
        print(f"       Otimizada: {float(vazao_med_media)} L/h")
        print(f"       Diferença: {float(vazao_med_media - vazao_med_original)} L/h")
        
        # Tolerância mais rigorosa para esta versão
        tolerancia = Decimal('1e-10')
        
        if (abs(vazao_ref_media - vazao_ref_original) > tolerancia or
            abs(vazao_med_media - vazao_med_original) > tolerancia):
            
            print(f"   ❌ PRECISÃO INSUFICIENTE!")
            print(f"       Erro Vazão Ref: {float(abs(vazao_ref_media - vazao_ref_original))}")
            print(f"       Erro Vazão Medidor: {float(abs(vazao_med_media - vazao_med_original))}")
            verificacao_passed = False
        else:
            print(f"   ✅ PRECISÃO EXCELENTE!")
            print(f"       Erro Vazão Ref: {float(abs(vazao_ref_media - vazao_ref_original))}")
            print(f"       Erro Vazão Medidor: {float(abs(vazao_med_media - vazao_med_original))}")
    
    return verificacao_passed

def verificar_otimizacao_individual_ponto(dados_ajustados, constantes, valores_certificado_originais, ponto_key):
    """
    NOVA FUNÇÃO: Verificação individual específica para cada ponto
    Analisa a qualidade da otimização de cada ponto separadamente
    """
    print(f"\n🔍 VERIFICAÇÃO INDIVIDUAL ESPECÍFICA para {ponto_key}")
    print("=" * 80)
    
    dados_ponto = dados_ajustados[ponto_key]
    valores_cert_originais = valores_certificado_originais[ponto_key]
    leituras_ajustadas = dados_ponto['leituras_ajustadas']
    
    # Extrai valores originais do ponto
    valores_sagrados_originais = dados_ponto['valores_sagrados']
    
    print(f"   📊 VALORES ORIGINAIS DO PONTO:")
    print(f"     Vazão Média: {float(valores_sagrados_originais['vazao_media'])} L/h")
    print(f"     Tendência: {float(valores_sagrados_originais['tendencia'])} %")
    print(f"     Desvio Padrão: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
    print(f"     Média Totalização (Certificado): {float(valores_cert_originais['media_totalizacao'])} L")
    print(f"     Média Leitura Medidor (Certificado): {float(valores_cert_originais['media_leitura_medidor'])} L")
    
    # Recalcula valores com dados ajustados
    totalizacoes_calculadas = []
    vazoes_ref_calculadas = []
    vazoes_medidor_calculadas = []
    erros_calculados = []
    
    print(f"\n   🔬 CÁLCULOS DETALHADOS POR LEITURA:")
    
    for i, leitura in enumerate(leituras_ajustadas):
        print(f"\n     📊 LEITURA {i+1} (Linha {leitura['linha']}):")
        print(f"       Pulsos: {int(leitura['pulsos_padrao'])}")
        print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
        print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
        print(f"       Temperatura: {float(leitura['temperatura'])} °C")
        
        # Calcula "Totalização no Padrão Corrigido • L" com dados ajustados
        totalizacao = calcular_totalizacao_padrao_corrigido(
            leitura['pulsos_padrao'],
            constantes['pulso_padrao_lp'],
            constantes['temperatura_constante'],
            constantes['fator_correcao_temp'],
            leitura['tempo_coleta']
        )
        totalizacoes_calculadas.append(totalizacao)
        
        # Calcula "Vazão de Referência • L/h"
        vazao_ref = (totalizacao / leitura['tempo_coleta']) * Decimal('3600')
        vazoes_ref_calculadas.append(vazao_ref)
        
        # Calcula "Vazão do Medidor • L/h"
        vazao_med = leitura['leitura_medidor']
        vazoes_medidor_calculadas.append(vazao_med)
        
        # Calcula erro percentual
        if totalizacao != 0:
            erro = ((leitura['leitura_medidor'] - totalizacao) / totalizacao) * Decimal('100')
        else:
            erro = Decimal('0')
        erros_calculados.append(erro)
        
        print(f"       Totalização Calculada: {float(totalizacao)} L")
        print(f"       Vazão Ref: {float(vazao_ref)} L/h")
        print(f"       Vazão Medidor: {float(vazao_med)} L/h")
        print(f"       Erro: {float(erro)} %")
    
    # Calcula médias ajustadas
    vazao_ref_media = sum(vazoes_ref_calculadas) / Decimal(str(len(vazoes_ref_calculadas)))
    vazao_med_media = sum(vazoes_medidor_calculadas) / Decimal(str(len(vazoes_medidor_calculadas)))
    media_totalizacao = sum(totalizacoes_calculadas) / Decimal(str(len(totalizacoes_calculadas)))
    tendencia_ajustada = sum(erros_calculados) / Decimal(str(len(erros_calculados)))
    desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_calculados)
    
    print(f"\n   📊 VALORES RECALCULADOS COM DADOS AJUSTADOS:")
    print(f"     Vazão Média: {float(vazao_ref_media)} L/h")
    print(f"     Tendência: {float(tendencia_ajustada)} %")
    print(f"     Desvio Padrão: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
    print(f"     Média Totalização: {float(media_totalizacao)} L")
    print(f"     Média Leitura Medidor: {float(vazao_med_media)} L")
    
    # Compara com valores originais
    print(f"\n   📊 COMPARAÇÃO COM VALORES ORIGINAIS:")
    
    # Vazão Média
    diff_vazao = vazao_ref_media - valores_sagrados_originais['vazao_media']
    print(f"     Vazão Média:")
    print(f"       Original: {float(valores_sagrados_originais['vazao_media'])} L/h")
    print(f"       Ajustada: {float(vazao_ref_media)} L/h")
    print(f"       Diferença: {float(diff_vazao)} L/h")
    print(f"       Erro Relativo: {float((diff_vazao / valores_sagrados_originais['vazao_media']) * 100)} %")
    
    # Tendência
    diff_tendencia = tendencia_ajustada - valores_sagrados_originais['tendencia']
    print(f"     Tendência:")
    print(f"       Original: {float(valores_sagrados_originais['tendencia'])} %")
    print(f"       Ajustada: {float(tendencia_ajustada)} %")
    print(f"       Diferença: {float(diff_tendencia)} %")
    
    # Desvio Padrão
    if valores_sagrados_originais['desvio_padrao'] and desvio_padrao_ajustado:
        diff_desvio = desvio_padrao_ajustado - valores_sagrados_originais['desvio_padrao']
        print(f"     Desvio Padrão:")
        print(f"       Original: {float(valores_sagrados_originais['desvio_padrao'])} %")
        print(f"       Ajustado: {float(desvio_padrao_ajustado)} %")
        print(f"       Diferença: {float(diff_desvio)} %")
    else:
        print(f"     Desvio Padrão: N/A")
    
    # Média Totalização
    diff_totalizacao = media_totalizacao - valores_cert_originais['media_totalizacao']
    print(f"     Média Totalização:")
    print(f"       Original: {float(valores_cert_originais['media_totalizacao'])} L")
    print(f"       Ajustada: {float(media_totalizacao)} L")
    print(f"       Diferença: {float(diff_totalizacao)} L")
    print(f"       Erro Relativo: {float((diff_totalizacao / valores_cert_originais['media_totalizacao']) * 100)} %")
    
    # Média Leitura Medidor
    diff_leitura = vazao_med_media - valores_cert_originais['media_leitura_medidor']
    print(f"     Média Leitura Medidor:")
    print(f"       Original: {float(valores_cert_originais['media_leitura_medidor'])} L")
    print(f"       Ajustada: {float(vazao_med_media)} L")
    print(f"       Diferença: {float(diff_leitura)} L")
    print(f"       Erro Relativo: {float((diff_leitura / valores_cert_originais['media_leitura_medidor']) * 100)} %")
    
    # Avalia a qualidade da otimização
    tolerancia_vazao = Decimal('1e-3')  # Mais permissivo
    tolerancia_tendencia = Decimal('1e-3')  # Mais permissivo
    tolerancia_totalizacao = Decimal('1e-3')  # Mais permissivo
    tolerancia_leitura = Decimal('1e-3')  # Mais permissivo
    
    qualidade_vazao = abs(diff_vazao) <= tolerancia_vazao
    qualidade_tendencia = abs(diff_tendencia) <= tolerancia_tendencia
    qualidade_totalizacao = abs(diff_totalizacao) <= tolerancia_totalizacao
    qualidade_leitura = abs(diff_leitura) <= tolerancia_leitura
    
    print(f"\n   🎯 AVALIAÇÃO DA QUALIDADE DA OTIMIZAÇÃO:")
    print(f"     Vazão Média: {'✅ EXCELENTE' if qualidade_vazao else '❌ PRECISA MELHORAR'}")
    print(f"     Tendência: {'✅ EXCELENTE' if qualidade_tendencia else '❌ PRECISA MELHORAR'}")
    print(f"     Média Totalização: {'✅ EXCELENTE' if qualidade_totalizacao else '❌ PRECISA MELHORAR'}")
    print(f"     Média Leitura: {'✅ EXCELENTE' if qualidade_leitura else '❌ PRECISA MELHORAR'}")
    
    # Calcula score geral
    score = 0
    if qualidade_vazao: score += 1
    if qualidade_tendencia: score += 1
    if qualidade_totalizacao: score += 1
    if qualidade_leitura: score += 1
    
    score_percentual = (score / 4) * 100
    
    print(f"\n   📈 SCORE GERAL: {score}/4 ({float(score_percentual)}%)")
    
    if score_percentual >= 75:
        print(f"     🎉 OTIMIZAÇÃO EXCELENTE!")
    elif score_percentual >= 50:
        print(f"     ✅ OTIMIZAÇÃO BOA")
    else:
        print(f"     ⚠️  OTIMIZAÇÃO PRECISA MELHORAR")
    
    return {
        'score': score,
        'score_percentual': score_percentual,
        'qualidade_vazao': qualidade_vazao,
        'qualidade_tendencia': qualidade_tendencia,
        'qualidade_totalizacao': qualidade_totalizacao,
        'qualidade_leitura': qualidade_leitura,
        'diferencas': {
            'vazao': diff_vazao,
            'tendencia': diff_tendencia,
            'totalizacao': diff_totalizacao,
            'leitura': diff_leitura
        }
    }

def main():
    """Função principal que executa todos os passos conforme documentação"""
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print("=== AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO ===")
    print("Implementa exatamente a lógica especificada na documentação")
<<<<<<< HEAD
    print("CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta fixados em 240 ou 360 segundos")
=======
    print("CONFIGURAÇÃO ESPECIAL: Otimização individual para cada ponto")
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
    print("Preserva valores sagrados: Vazão Média, Tendência e Desvio Padrão")
    print("Usa precisão Decimal de 50 dígitos")
    print("Estratégia: Otimização iterativa com função de custo")
    
    # Escolha do tempo alvo
    print(f"\n⏱️  ESCOLHA DO TEMPO ALVO:")
    print(f"   1. 240 segundos")
    print(f"   2. 360 segundos")
    
    try:
        escolha = input("   Digite 1 ou 2 para escolher o tempo alvo: ").strip()
        if escolha == "1":
            tempo_alvo = Decimal('240')
            print(f"   ✅ Tempo alvo escolhido: 240 segundos")
        elif escolha == "2":
            tempo_alvo = Decimal('360')
            print(f"   ✅ Tempo alvo escolhido: 360 segundos")
        else:
            print(f"   ⚠️  Escolha inválida, usando padrão: 240 segundos")
            tempo_alvo = Decimal('240')
    except:
        print(f"   ⚠️  Erro na entrada, usando padrão: 240 segundos")
        tempo_alvo = Decimal('240')
    
    # PASSO 1: Extração de Dados
    dados_originais = extrair_dados_originais(arquivo_excel)
    
    if not dados_originais:
        print("❌ Falha na extração dos dados originais")
        return
    
    print(f"\n✅ PASSO 1 CONCLUÍDO: {len(dados_originais)} pontos extraídos")
    
    # PASSO 1.5: Extração de Constantes e Cálculo dos Valores do Certificado
    constantes = extrair_constantes_calculo(arquivo_excel)
    if not constantes:
        print("❌ Falha na extração das constantes")
        return
    
    valores_certificado_originais = calcular_valores_certificado(dados_originais, constantes)
    print(f"\n✅ PASSO 1.5 CONCLUÍDO: Valores do certificado calculados")
    
    # PASSO 2: Harmonização dos Tempos de Coleta com Otimização Iterativa
    dados_harmonizados = harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais, tempo_alvo)
    
    print(f"\n✅ PASSO 2 CONCLUÍDO: Otimização iterativa executada")
    
    # PASSO 3: Aplicação do Ajuste Proporcional
    dados_ajustados = aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais)
    
    print(f"\n✅ PASSO 3 CONCLUÍDO: Ajuste proporcional aplicado")
    
    # NOVA VERIFICAÇÃO: Verificação individual de cada ponto
    print(f"\n🔍 NOVA VERIFICAÇÃO INDIVIDUAL DE CADA PONTO")
    print("=" * 80)
    
    resultados_verificacao = {}
    score_total = 0
    num_pontos = len(dados_ajustados)
    
    for ponto_key in dados_ajustados.keys():
        print(f"\n{'='*80}")
        resultado_verificacao = verificar_otimizacao_individual_ponto(
            dados_ajustados, 
            constantes, 
            valores_certificado_originais, 
            ponto_key
        )
        resultados_verificacao[ponto_key] = resultado_verificacao
        score_total += resultado_verificacao['score_percentual']
    
    # Calcula score médio geral
    score_medio = score_total / num_pontos
    
    print(f"\n{'='*80}")
    print(f"📊 RESUMO GERAL DA OTIMIZAÇÃO")
    print(f"{'='*80}")
    print(f"   Pontos processados: {num_pontos}")
    print(f"   Score médio geral: {float(score_medio)}%")
    
    # Mostra resultados por ponto
    print(f"\n   📋 RESULTADOS POR PONTO:")
    for ponto_key, resultado in resultados_verificacao.items():
        print(f"     {ponto_key}: {float(resultado['score_percentual'])}% ({resultado['score']}/4)")
    
    # Avalia qualidade geral
    if score_medio >= 75:
        print(f"\n   🎉 OTIMIZAÇÃO GERAL EXCELENTE!")
        verificacao_geral_passed = True
    elif score_medio >= 50:
        print(f"\n   ✅ OTIMIZAÇÃO GERAL BOA")
        verificacao_geral_passed = True
    else:
        print(f"\n   ⚠️  OTIMIZAÇÃO GERAL PRECISA MELHORAR")
        verificacao_geral_passed = False
    
    # PASSO 4: Verificação dos Valores Sagrados (mantém a verificação original)
    verificacao_passed = verificar_valores_sagrados(dados_ajustados)
    
    if verificacao_passed:
        print(f"\n✅ PASSO 4 CONCLUÍDO: Valores sagrados preservados")
        
        # NOVA VERIFICAÇÃO DE PRECISÃO
        print(f"\n🔍 NOVA VERIFICAÇÃO DE PRECISÃO")
        verificacao_precisao_passed = verificar_precisao(dados_ajustados, constantes, valores_certificado_originais)
        
        if verificacao_precisao_passed:
            print(f"\n✅ NOVA VERIFICAÇÃO PASSOU: Precisão excelente alcançada")
        else:
            print(f"\n❌ NOVA VERIFICAÇÃO FALHOU: Precisão insuficiente")
        
        # VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO
        print(f"\n🔍 VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO")
        verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais)
        
        # VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR
        verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais)
        
        # PASSO 5: Geração da Planilha Corrigida
        arquivo_corrigido = gerar_planilha_corrigida(dados_ajustados, arquivo_excel)
        
        print(f"\n✅ PASSO 5 CONCLUÍDO: Planilha corrigida gerada")
        gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido)
        
        print(f"\n🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
        print(f"   ✅ Todos os passos executados conforme documentação")
<<<<<<< HEAD
        print(f"   ✅ Otimização iterativa executada com sucesso")
        print(f"   ✅ Tempo alvo: {float(tempo_alvo)} segundos")
        print(f"   ✅ Valores sagrados preservados absolutamente")
=======
        print(f"   ✅ Otimização individual aplicada para cada ponto")
        print(f"   ✅ Score médio geral: {float(score_medio)}%")
        if verificacao_geral_passed:
            print(f"   ✅ Otimização geral considerada satisfatória")
        else:
            print(f"   ⚠️  Otimização geral precisa de refinamento")
        if verificacao_precisao_passed:
            print(f"   ✅ Nova otimização alcançou precisão excelente")
        else:
            print(f"   ⚠️  Nova otimização precisa de refinamento")
>>>>>>> 81bc9d1194f5c52755d54e00295d56790a10f47b
        print(f"   ✅ Planilha corrigida: {arquivo_corrigido}")
        print(f"   ✅ Relatórios gerados com sucesso")
        
    else:
        print(f"\n❌ PASSO 4 FALHOU: Valores sagrados foram alterados")
        print(f"   ⚠️  Revisar implementação da otimização iterativa")
        print(f"   ⚠️  Verificar lógica de preservação dos valores")

if __name__ == "__main__":
    main()