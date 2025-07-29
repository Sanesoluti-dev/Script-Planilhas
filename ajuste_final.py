# -*- coding: utf-8 -*-
"""
AJUSTADOR DE TEMPO DE COLETA - VERSÃO CORRIGIDA
================================================

Este script ajusta o tempo de coleta para 360 segundos e recalcula os
valores de pulso e leitura do medidor proporcionalmente para manter
a Vazão Média, Tendência e Desvio Padrão Amostral constantes.

Lógica de correção:
1.  Fixa o 'Tempo de Coleta' em 360 segundos.
2.  Calcula um 'Fator de Ajuste' = (Tempo Novo / Tempo Antigo).
3.  Aplica este fator aos valores de volume ('Qtd de Pulsos' e 'Leitura no Medidor').
4.  Mantém a temperatura original medida.
"""

import pandas as pd
from decimal import Decimal, getcontext, ROUND_HALF_UP
import shutil
from openpyxl import load_workbook

# Configura a precisão para os cálculos
getcontext().prec = 28

def converter_para_decimal(valor):
    """Converte um valor (string ou número) para Decimal com alta precisão."""
    if valor is None:
        return Decimal('0')
    if isinstance(valor, str):
        valor_limpo = valor.replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    return Decimal(str(valor))

def ajustar_planilha(arquivo_excel_original, tempo_alvo_segundos=360):
    """
    Função principal para ler, ajustar e salvar a planilha.
    """
    print(f"📄 Iniciando ajuste para o arquivo: {arquivo_excel_original}")
    print(f"🎯 Tempo de coleta alvo: {tempo_alvo_segundos} segundos")
    print("-" * 50)

    try:
        # Carrega a planilha usando openpyxl para ler os valores exatos
        wb = load_workbook(arquivo_excel_original, data_only=True)
        sheet = wb["Coleta de Dados"]
        
        # Usa pandas apenas para identificar a estrutura e as linhas
        df = pd.read_excel(arquivo_excel_original, sheet_name='Coleta de Dados', header=None)

    except FileNotFoundError:
        print(f"❌ ERRO: Arquivo não encontrado: {arquivo_excel_original}")
        return
    except KeyError:
        print("❌ ERRO: A aba 'Coleta de Dados' não foi encontrada na planilha.")
        return

    # --- Identificação dos pontos de calibração ---
    pontos_config = []
    linha_inicial = 50 # Linha de início da primeira tabela de ponto (base 0 no pandas)
    
    while True:
        # Verifica se as células de pulso nas próximas 3 linhas estão vazias ou zeradas
        # para determinar o fim dos pontos de calibração.
        # df.iloc[linha, coluna]
        try:
            pulso1 = df.iloc[linha_inicial + 3, 2]
            if pd.isna(pulso1) or pulso1 == 0:
                break
        except IndexError:
            break
            
        pontos_config.append({'linha_inicial_excel': linha_inicial + 2}) # +2 para ajustar index (1-based e header)
        linha_inicial += 9 # Pula para o próximo bloco de ponto
        
    if not pontos_config:
        print("⚠️ Nenhum ponto de calibração encontrado. Verifique a estrutura da planilha.")
        return

    print(f"✅ {len(pontos_config)} pontos de calibração identificados.")

    # --- Aplicação do ajuste proporcional ---
    tempo_novo = Decimal(str(tempo_alvo_segundos))

    for ponto in pontos_config:
        linha_base = ponto['linha_inicial_excel']
        print(f"\n🔄 Processando ponto na linha base: {linha_base}")

        # Itera sobre as 3 leituras de cada ponto
        for i in range(3):
            linha_atual = linha_base + i + 2 # Linhas 54, 55, 56 para o primeiro ponto
            
            # Extrai os valores originais com alta precisão
            pulsos_antigo = converter_para_decimal(sheet.cell(row=linha_atual, column=3).value) # Col C
            tempo_antigo = converter_para_decimal(sheet.cell(row=linha_atual, column=6).value)   # Col F
            leitura_medidor_antiga = converter_para_decimal(sheet.cell(row=linha_atual, column=15).value) # Col O
            
            # Se não houver tempo antigo, não há o que ajustar
            if tempo_antigo == 0:
                print(f"   - Linha {linha_atual}: Sem dados, pulando.")
                continue

            # 1. Calcula o fator de ajuste
            fator_ajuste = tempo_novo / tempo_antigo

            # 2. Calcula os novos valores proporcionais
            pulsos_novo = (pulsos_antigo * fator_ajuste).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            leitura_medidor_nova = leitura_medidor_antiga * fator_ajuste
            
            print(f"   - Linha {linha_atual}:")
            print(f"     Tempo: {tempo_antigo}s -> {tempo_novo}s")
            print(f"     Pulsos: {pulsos_antigo} -> {pulsos_novo}")
            print(f"     Leitura Medidor: {leitura_medidor_antiga:.4f}L -> {leitura_medidor_nova:.4f}L")

            # 3. Atualiza os valores na planilha carregada na memória (com openpyxl)
            # A temperatura original (Coluna R) não é alterada.
            sheet.cell(row=linha_atual, column=3).value = int(pulsos_novo)
            sheet.cell(row=linha_atual, column=6).value = float(tempo_novo)
            sheet.cell(row=linha_atual, column=15).value = float(leitura_medidor_nova)

    # --- Geração da nova planilha ---
    arquivo_corrigido = arquivo_excel_original.replace('.xlsx', '_CORRIGIDO.xlsx')
    shutil.copy(arquivo_excel_original, arquivo_corrigido)
    
    # Salva o workbook com as alterações no novo arquivo
    wb.save(arquivo_corrigido)
    
    print("-" * 50)
    print(f"🎉 Processo concluído com sucesso!")
    print(f"💾 Planilha corrigida salva como: {arquivo_corrigido}")


# --- EXECUÇÃO ---
if __name__ == "__main__":
    # Coloque o nome do seu arquivo Excel aqui
    # Exemplo: SAN-038-25-09-1.xlsx
    nome_do_arquivo = "SAN-038-25-09-1.xlsx" 
    
    # Certifique-se de que o arquivo Excel exista no mesmo diretório do script
    # ou forneça o caminho completo.
    try:
        # Cria um arquivo de exemplo se não existir, para o código rodar.
        # Em um caso real, você já teria o arquivo.
        pd.DataFrame([[]]).to_excel(nome_do_arquivo, index=False)
        print(f"AVISO: Criado um arquivo '{nome_do_arquivo}' de exemplo. "
              "Por favor, substitua-o pelo seu arquivo real e execute novamente.")
    except Exception:
        # Se não puder criar (permissão, etc), apenas continua
        pass
        
    ajustar_planilha(nome_do_arquivo)