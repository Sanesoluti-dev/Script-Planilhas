# -*- coding: utf-8 -*-
"""
Otimizador de Tempos de Coleta - Versão Simples e Eficiente
Decrementa tempos até encontrar valores exatos
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from decimal import Decimal, ROUND_HALF_UP, getcontext
import json
import os
import time
import numpy as np
import shutil

# Configura precisão máxima
getcontext().prec = 28



def converter_para_decimal_padrao(valor):
    """
    Converte valor para Decimal com precisão máxima
    """
    if valor is None or valor == "":
        return Decimal('0')
    
    try:
        if isinstance(valor, (int, float)):
            return Decimal(str(valor))
        elif isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            return Decimal(valor_limpo)
        else:
            return Decimal(str(valor))
    except:
        return Decimal('0')

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    try:
        valor = sheet.cell(row=linha, column=coluna).value
        return converter_para_decimal_padrao(valor)
    except Exception as e:
        print(f"       ERRO ao ler valor na linha {linha}, coluna {coluna}: {e}")
        return Decimal('0')

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal máxima
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    media = media.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    soma_quadrados = soma_quadrados.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    variancia = variancia.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    desvio_padrao = variancia.sqrt()
    desvio_padrao = desvio_padrao.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    return desvio_padrao

def extrair_dados_planilha_original(arquivo_excel):
    """
    Extrai todos os dados necessários da planilha original
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        estimativa_sheet = wb["Estimativa da Incerteza"]
        
        # Extrai constantes
        constantes = {}
        constantes['ponto_mlp'] = ler_valor_exato(coleta_sheet, 50, 9)  # I50
        constantes['pulso_equipamento_mlp'] = ler_valor_exato(coleta_sheet, 50, 30)  # AD50
        constantes['constante_correcao_temp'] = ler_valor_exato(coleta_sheet, 51, 18)  # R51
        constantes['constante_correcao_inclinacao'] = ler_valor_exato(coleta_sheet, 51, 21)  # U51
        constantes['modo_calibracao'] = coleta_sheet.cell(row=16, column=24).value  # X16
        constantes['correcao_tempo_bu23'] = ler_valor_exato(estimativa_sheet, 23, 73)  # BU23
        constantes['correcao_tempo_bw23'] = ler_valor_exato(estimativa_sheet, 23, 75)  # BW23
        constantes['correcao_temp_bu26'] = ler_valor_exato(estimativa_sheet, 26, 73)  # BU26
        constantes['correcao_temp_bw26'] = ler_valor_exato(estimativa_sheet, 26, 75)  # BW26
        
        # Extrai dados dos pontos
        pontos = []
        linha_atual = 54
        num_ponto = 1
        
        while True:
            pulsos = ler_valor_exato(coleta_sheet, linha_atual, 3)
            
            if pulsos == 0:
                pulsos2 = ler_valor_exato(coleta_sheet, linha_atual + 1, 3)
                pulsos3 = ler_valor_exato(coleta_sheet, linha_atual + 2, 3)
                
                if pulsos2 == 0 and pulsos3 == 0:
                    break
            
            ponto = {
                'numero': num_ponto,
                'linha_inicial': linha_atual,
                'leituras': [],
                'valores_originais': {}
            }
            
            # Extrai valores brutos e calculados
            for i in range(3):
                linha = linha_atual + i
                
                # Valores brutos
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)
                
                # Valores calculados (para manter iguais)
                tempo_coleta_corrigido = ler_valor_exato(coleta_sheet, linha, 27)
                temperatura_corrigida = ler_valor_exato(coleta_sheet, linha, 30)
                totalizacao_padrao_corrigido = ler_valor_exato(coleta_sheet, linha, 12)
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)
                vazao_medidor = ler_valor_exato(coleta_sheet, linha, 24)
                erro_percentual = ler_valor_exato(coleta_sheet, linha, 21)
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'tempo_coleta_corrigido': tempo_coleta_corrigido,
                    'temperatura_corrigida': temperatura_corrigida,
                    'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
                    'vazao_referencia': vazao_referencia,
                    'vazao_medidor': vazao_medidor,
                    'erro_percentual': erro_percentual
                }
                
                ponto['leituras'].append(leitura)
            
            # Valores agregados originais
            linha_agregados = linha_atual + 3
            vazao_media_original = ler_valor_exato(coleta_sheet, linha_agregados, 9)
            tendencia_original = ler_valor_exato(coleta_sheet, linha_agregados, 21)
            desvio_padrao_original = ler_valor_exato(coleta_sheet, linha_agregados, 30)
            
            ponto['valores_originais'] = {
                'vazao_media': vazao_media_original,
                'tendencia': tendencia_original,
                'desvio_padrao': desvio_padrao_original
            }
            
            pontos.append(ponto)
            num_ponto += 1
            linha_atual += 9
        
        return constantes, pontos
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados: {e}")
        return None, None

def calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_ajustados):
    """
    Calcula todas as fórmulas com tempos ajustados usando precisão máxima
    """
    resultados = []
    
    for i, leitura in enumerate(leituras):
        linha = leitura['linha']
        tempo_ajustado = tempos_ajustados[i]
        
        # FÓRMULA 1: Pulso do padrão em L/P (I51)
        pulso_padrao_lp = constantes['ponto_mlp'] / Decimal('1000')
        
        # FÓRMULA 2: Pulso do Equipamento em L/P (AD51)
        pulso_equipamento_lp = constantes['pulso_equipamento_mlp'] / Decimal('1000')
        
        # FÓRMULA 3: Tempo de Coleta Corrigido • (s) (AA54)
        tempo_coleta_bruto = tempo_ajustado
        
        tempo_coleta_corrigido = tempo_coleta_bruto - (tempo_coleta_bruto * constantes['correcao_tempo_bu23'] + constantes['correcao_tempo_bw23'])
        tempo_coleta_corrigido = tempo_coleta_corrigido.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 4: Temperatura da Água Corrigida • °C (AD54)
        temperatura_bruta = leitura['temperatura']
        temperatura_corrigida = temperatura_bruta - (temperatura_bruta * constantes['correcao_temp_bu26'] + constantes['correcao_temp_bw26'])
        temperatura_corrigida = temperatura_corrigida.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 5: Totalização no Padrão Corrigido • L (L54)
        pulsos_padrao = leitura['pulsos_padrao']
        volume_bruto = pulsos_padrao * pulso_padrao_lp
        volume_bruto = volume_bruto.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        vazao_bruta = volume_bruto / tempo_coleta_corrigido * Decimal('3600')
        vazao_bruta = vazao_bruta.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        correcao = (constantes['constante_correcao_temp'] + constantes['constante_correcao_inclinacao'] * vazao_bruta) / Decimal('100') * volume_bruto
        correcao = correcao.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        totalizacao_padrao_corrigido = volume_bruto - correcao
        totalizacao_padrao_corrigido = totalizacao_padrao_corrigido.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 6: Vazão de Referência • L/h (I54)
        vazao_referencia = totalizacao_padrao_corrigido / tempo_coleta_corrigido * Decimal('3600')
        vazao_referencia = vazao_referencia.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 7: Vazão do Medidor • L/h (X54)
        leitura_medidor = leitura['leitura_medidor']
        modo_calibracao = constantes['modo_calibracao']
        
        if modo_calibracao in ["Visual com início dinâmico", "Visual com início estática"]:
            vazao_medidor = leitura_medidor
        else:
            vazao_medidor = (leitura_medidor / tempo_coleta_corrigido) * Decimal('3600')
            vazao_medidor = vazao_medidor.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        # FÓRMULA 8: Erro % (U54)
        erro_percentual = (leitura_medidor - totalizacao_padrao_corrigido) / totalizacao_padrao_corrigido * Decimal('100')
        erro_percentual = erro_percentual.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
        
        resultado = {
            'linha': linha,
            'tempo_coleta_ajustado': tempo_ajustado,
            'tempo_coleta_corrigido': tempo_coleta_corrigido,
            'temperatura_corrigida': temperatura_corrigida,
            'totalizacao_padrao_corrigido': totalizacao_padrao_corrigido,
            'vazao_referencia': vazao_referencia,
            'vazao_medidor': vazao_medidor,
            'erro_percentual': erro_percentual
        }
        
        resultados.append(resultado)
    
    return resultados

def calcular_agregados_com_tempo_ajustado(resultados):
    """
    Calcula os valores agregados com tempos ajustados usando precisão máxima
    """
    vazoes_referencia = [r['vazao_referencia'] for r in resultados]
    erros_percentuais = [r['erro_percentual'] for r in resultados]
    
    # FÓRMULA 9: Vazão Média • L/h (I57)
    vazao_media = sum(vazoes_referencia) / Decimal(str(len(vazoes_referencia)))
    vazao_media = vazao_media.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    # FÓRMULA 10: Tendência (U57)
    tendencia = sum(erros_percentuais) / Decimal(str(len(erros_percentuais)))
    tendencia = tendencia.quantize(Decimal('0.000000000001'), rounding=ROUND_HALF_UP)
    
    # FÓRMULA 11: DESVIO PADRÃO AMOSTRAL (AD57)
    desvio_padrao = calcular_desvio_padrao_amostral(erros_percentuais)
    
    return {
        'vazao_media': vazao_media,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def otimizar_tempos_ponto_simples(leituras, constantes, valores_originais):
    """
    Otimiza os tempos de coleta usando decremento simples até encontrar valores exatos
    """
    print(f"   🔍 Iniciando otimização SIMPLES para Ponto {leituras[0]['linha']}...")
    print(f"   🎯 OBJETIVO: Vazão média exata = {float(valores_originais['vazao_media']):.3f}")
    
    # Começa com tempos originais
    tempos_atuais = [leitura['tempo_coleta'] for leitura in leituras]
    print(f"   📊 Tempos iniciais: {[float(t) for t in tempos_atuais]}")
    
    # Calcula vazão inicial
    resultados_iniciais = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
    agregados_iniciais = calcular_agregados_com_tempo_ajustado(resultados_iniciais)
    
    print(f"   📊 Vazão inicial: {float(agregados_iniciais['vazao_media']):.6f}")
    print(f"   📊 Vazão desejada: {float(valores_originais['vazao_media']):.6f}")
    
    # Verifica se já está correto
    vazao_desejada = valores_originais['vazao_media']
    vazao_atual = agregados_iniciais['vazao_media']
    
    # Arredonda para 3 casas decimais para comparação
    vazao_atual_3casas = vazao_atual.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
    vazao_desejada_3casas = vazao_desejada.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
    
    if vazao_atual_3casas == vazao_desejada_3casas:
        print(f"   ✅ Vazão já está correta! {float(vazao_atual_3casas):.3f}")
        return {
            'tempos': tempos_atuais,
            'agregados': agregados_iniciais,
            'iteracoes': 0
        }
    
    # Se a vazão atual é maior que a desejada, precisa diminuir os tempos
    if vazao_atual > vazao_desejada:
        print(f"   📉 Vazão atual ({float(vazao_atual):.6f}) > desejada ({float(vazao_desejada):.6f})")
        print(f"   🔧 Diminuindo tempos de coleta...")
        
        iteracoes = 0
        max_iteracoes = 10000  # Limite de segurança
        
        while iteracoes < max_iteracoes:
            iteracoes += 1
            
            # Decrementa todos os tempos em 0.001
            for i in range(len(tempos_atuais)):
                novo_tempo = tempos_atuais[i] - Decimal('0.001')
                
                # Verifica se está dentro da regra 239.599-240.499
                if 239.599 <= float(novo_tempo) <= 240.499:
                    tempos_atuais[i] = novo_tempo
                else:
                    print(f"   ⚠️  Tempo {i+1} atingiu limite mínimo: {float(novo_tempo):.3f}")
                    # Se um tempo atingiu o limite, para de decrementar
                    break
            else:
                # Calcula nova vazão
                resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                agregados = calcular_agregados_com_tempo_ajustado(resultados)
                vazao_atual = agregados['vazao_media']
                
                # Arredonda para 3 casas decimais
                vazao_atual_3casas = vazao_atual.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                
                if vazao_atual_3casas == vazao_desejada_3casas:
                    print(f"   ✅ Vazão encontrada após {iteracoes} iterações!")
                    print(f"   📊 Vazão final: {float(vazao_atual):.6f}")
                    return {
                        'tempos': tempos_atuais.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se a vazão ficou menor que a desejada, voltou um passo
                if vazao_atual < vazao_desejada:
                    print(f"   ⚠️  Vazão ficou menor que o desejado: {float(vazao_atual):.6f} < {float(vazao_desejada):.6f}")
                    # Volta um passo
                    for i in range(len(tempos_atuais)):
                        tempos_atuais[i] += Decimal('0.001')
                    
                    # Calcula resultado final
                    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                    agregados = calcular_agregados_com_tempo_ajustado(resultados)
                    
                    return {
                        'tempos': tempos_atuais.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes - 1
                    }
                
                continue
            
            # Se chegou aqui, um tempo atingiu o limite
            break
        
        print(f"   ❌ Não foi possível encontrar vazão exata após {iteracoes} iterações")
        return None
    
    else:
        print(f"   📈 Vazão atual ({float(vazao_atual):.6f}) < desejada ({float(vazao_desejada):.6f})")
        print(f"   🔧 Aumentando tempos de coleta...")
        
        # Verifica se os tempos já estão no limite máximo
        tempos_no_limite = [t for t in tempos_atuais if float(t) >= 240.499]
        if len(tempos_no_limite) > 0:
            print(f"   ⚠️  ALGUNS TEMPOS JÁ ESTÃO NO LIMITE MÁXIMO!")
            print(f"   📊 Tempos no limite: {[float(t) for t in tempos_no_limite]}")
            
            # Tenta uma abordagem diferente: diminui os tempos que não estão no limite
            tempos_nao_limite = [i for i, t in enumerate(tempos_atuais) if float(t) < 240.499]
            
            if len(tempos_nao_limite) > 0:
                print(f"   🔧 Tentando diminuir tempos que não estão no limite...")
                
                iteracoes = 0
                max_iteracoes = 1000
                
                while iteracoes < max_iteracoes:
                    iteracoes += 1
                    
                    # Diminui apenas os tempos que não estão no limite
                    for i in tempos_nao_limite:
                        novo_tempo = tempos_atuais[i] - Decimal('0.001')
                        
                        if 239.599 <= float(novo_tempo) <= 240.499:
                            tempos_atuais[i] = novo_tempo
                        else:
                            tempos_nao_limite.remove(i)
                            print(f"   ⚠️  Tempo {i+1} agora atingiu limite: {float(novo_tempo):.3f}")
                    
                    if not tempos_nao_limite:
                        print(f"   ❌ Todos os tempos atingiram o limite")
                        break
                    
                    # Calcula nova vazão
                    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                    agregados = calcular_agregados_com_tempo_ajustado(resultados)
                    vazao_atual = agregados['vazao_media']
                    
                    # Arredonda para 3 casas decimais
                    vazao_atual_3casas = vazao_atual.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                    
                    if vazao_atual_3casas == vazao_desejada_3casas:
                        print(f"   ✅ Vazão encontrada após {iteracoes} iterações!")
                        print(f"   📊 Vazão final: {float(vazao_atual):.6f}")
                        return {
                            'tempos': tempos_atuais.copy(),
                            'agregados': agregados,
                            'iteracoes': iteracoes
                        }
                    
                    # Se a vazão ficou menor que a desejada, voltou um passo
                    if vazao_atual < vazao_desejada:
                        print(f"   ⚠️  Vazão ficou menor que o desejado: {float(vazao_atual):.6f} < {float(vazao_desejada):.6f}")
                        # Volta um passo
                        for i in tempos_nao_limite:
                            tempos_atuais[i] += Decimal('0.001')
                        
                        # Calcula resultado final
                        resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                        agregados = calcular_agregados_com_tempo_ajustado(resultados)
                        
                        return {
                            'tempos': tempos_atuais.copy(),
                            'agregados': agregados,
                            'iteracoes': iteracoes - 1
                        }
            
            print(f"   ❌ Não foi possível otimizar com tempos no limite")
            return None
        
        iteracoes = 0
        max_iteracoes = 10000  # Limite de segurança
        
        while iteracoes < max_iteracoes:
            iteracoes += 1
            
            # Incrementa todos os tempos em 0.001
            for i in range(len(tempos_atuais)):
                novo_tempo = tempos_atuais[i] + Decimal('0.001')
                
                # Verifica se está dentro da regra 239.599-240.499
                if 239.599 <= float(novo_tempo) <= 240.499:
                    tempos_atuais[i] = novo_tempo
                else:
                    print(f"   ⚠️  Tempo {i+1} atingiu limite máximo: {float(novo_tempo):.3f}")
                    # Se um tempo atingiu o limite, para de incrementar
                    break
            else:
                # Calcula nova vazão
                resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                agregados = calcular_agregados_com_tempo_ajustado(resultados)
                vazao_atual = agregados['vazao_media']
                
                # Arredonda para 3 casas decimais
                vazao_atual_3casas = vazao_atual.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                
                if vazao_atual_3casas == vazao_desejada_3casas:
                    print(f"   ✅ Vazão encontrada após {iteracoes} iterações!")
                    print(f"   📊 Vazão final: {float(vazao_atual):.6f}")
                    return {
                        'tempos': tempos_atuais.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se a vazão ficou maior que a desejada, voltou um passo
                if vazao_atual > vazao_desejada:
                    print(f"   ⚠️  Vazão ficou maior que o desejado: {float(vazao_atual):.6f} > {float(vazao_desejada):.6f}")
                    # Volta um passo
                    for i in range(len(tempos_atuais)):
                        tempos_atuais[i] -= Decimal('0.001')
                    
                    # Calcula resultado final
                    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
                    agregados = calcular_agregados_com_tempo_ajustado(resultados)
                    
                    return {
                        'tempos': tempos_atuais.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes - 1
                    }
                
                continue
            
            # Se chegou aqui, um tempo atingiu o limite
            break
        
        print(f"   ❌ Não foi possível encontrar vazão exata após {iteracoes} iterações")
        return None

def otimizar_tempos_ponto_inteligente_v2(leituras, constantes, valores_originais):
    """
    Otimiza os tempos de coleta usando busca inteligente com incrementos menores
    """
    print(f"   🔍 Iniciando otimização INTELIGENTE V2 para Ponto {leituras[0]['linha']}...")
    print(f"   🎯 OBJETIVO: Vazão média exata = {float(valores_originais['vazao_media']):.3f}")
    
    # Começa com tempos originais
    tempos_atuais = [leitura['tempo_coleta'] for leitura in leituras]
    print(f"   📊 Tempos iniciais: {[float(t) for t in tempos_atuais]}")
    
    # Calcula vazão inicial
    resultados_iniciais = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_atuais)
    agregados_iniciais = calcular_agregados_com_tempo_ajustado(resultados_iniciais)
    
    print(f"   📊 Vazão inicial: {float(agregados_iniciais['vazao_media']):.6f}")
    print(f"   📊 Vazão desejada: {float(valores_originais['vazao_media']):.6f}")
    
    # Verifica se já está correto
    vazao_desejada = valores_originais['vazao_media']
    vazao_atual = agregados_iniciais['vazao_media']
    
    # Arredonda para 3 casas decimais para comparação
    vazao_atual_3casas = vazao_atual.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
    vazao_desejada_3casas = vazao_desejada.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
    
    if vazao_atual_3casas == vazao_desejada_3casas:
        print(f"   ✅ Vazão já está correta! {float(vazao_atual_3casas):.3f}")
        return {
            'tempos': tempos_atuais,
            'agregados': agregados_iniciais,
            'iteracoes': 0
        }
    
    # Busca o melhor valor possível
    melhor_combinacao = None
    melhor_diferenca = abs(vazao_atual - vazao_desejada)
    melhor_tempos = tempos_atuais.copy()
    melhor_agregados = agregados_iniciais
    
    # Verifica se os tempos estão no limite máximo
    tempos_no_limite = [t for t in tempos_atuais if float(t) >= 240.499]
    if len(tempos_no_limite) > 0:
        print(f"   ⚠️  ALGUNS TEMPOS ESTÃO NO LIMITE MÁXIMO!")
        print(f"   📊 Tempos no limite: {[float(t) for t in tempos_no_limite]}")
        
        # Se a vazão atual é menor que a desejada e os tempos estão no limite,
        # precisa diminuir os tempos para aumentar a vazão
        if vazao_atual < vazao_desejada:
            print(f"   🔧 Diminuindo tempos para aumentar vazão...")
            
            # Tenta diferentes decrementos
            decrementos = [Decimal('0.0001'), Decimal('0.0005'), Decimal('0.001')]
            
            for decremento in decrementos:
                print(f"   🔧 Tentando com decremento: {float(decremento)}")
                
                # Reinicia com tempos originais
                tempos_teste = [leitura['tempo_coleta'] for leitura in leituras]
                
                iteracoes = 0
                max_iteracoes = 1000
                
                while iteracoes < max_iteracoes:
                    iteracoes += 1
                    
                    # Diminui todos os tempos
                    tempos_alterados = False
                    for i in range(len(tempos_teste)):
                        novo_tempo = tempos_teste[i] - decremento
                        
                        if 239.599 <= float(novo_tempo) <= 240.499:
                            tempos_teste[i] = novo_tempo
                            tempos_alterados = True
                    
                    if not tempos_alterados:
                        print(f"   ⚠️  Todos os tempos atingiram o limite com decremento {float(decremento)}")
                        break
                    
                    # Calcula nova vazão
                    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_teste)
                    agregados = calcular_agregados_com_tempo_ajustado(resultados)
                    vazao_teste = agregados['vazao_media']
                    
                    # Verifica se é melhor
                    diferenca_atual = abs(vazao_teste - vazao_desejada)
                    if diferenca_atual < melhor_diferenca:
                        melhor_diferenca = diferenca_atual
                        melhor_tempos = tempos_teste.copy()
                        melhor_agregados = agregados
                        melhor_combinacao = {
                            'tempos': tempos_teste.copy(),
                            'agregados': agregados,
                            'iteracoes': iteracoes
                        }
                    
                    # Se chegou ao valor exato, para
                    vazao_teste_3casas = vazao_teste.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                    if vazao_teste_3casas == vazao_desejada_3casas:
                        print(f"   ✅ Vazão exata encontrada com decremento {float(decremento)}!")
                        return {
                            'tempos': tempos_teste.copy(),
                            'agregados': agregados,
                            'iteracoes': iteracoes
                        }
                    
                    # Se passou do valor desejado, para
                    if vazao_teste > vazao_desejada:
                        print(f"   ⚠️  Vazão passou do desejado: {float(vazao_teste):.6f} > {float(vazao_desejada):.6f}")
                        break
    
    # Tenta diferentes incrementos para casos normais
    incrementos = [Decimal('0.0001'), Decimal('0.0005'), Decimal('0.001')]
    
    for incremento in incrementos:
        print(f"   🔧 Tentando com incremento: {float(incremento)}")
        
        # Reinicia com tempos originais
        tempos_teste = [leitura['tempo_coleta'] for leitura in leituras]
        
        # Se a vazão atual é menor que a desejada, tenta aumentar
        if vazao_atual < vazao_desejada:
            iteracoes = 0
            max_iteracoes = 1000
            
            while iteracoes < max_iteracoes:
                iteracoes += 1
                
                # Tenta aumentar todos os tempos
                tempos_alterados = False
                for i in range(len(tempos_teste)):
                    novo_tempo = tempos_teste[i] + incremento
                    
                    if 239.599 <= float(novo_tempo) <= 240.499:
                        tempos_teste[i] = novo_tempo
                        tempos_alterados = True
                
                if not tempos_alterados:
                    print(f"   ⚠️  Todos os tempos atingiram o limite com incremento {float(incremento)}")
                    break
                
                # Calcula nova vazão
                resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_teste)
                agregados = calcular_agregados_com_tempo_ajustado(resultados)
                vazao_teste = agregados['vazao_media']
                
                # Verifica se é melhor
                diferenca_atual = abs(vazao_teste - vazao_desejada)
                if diferenca_atual < melhor_diferenca:
                    melhor_diferenca = diferenca_atual
                    melhor_tempos = tempos_teste.copy()
                    melhor_agregados = agregados
                    melhor_combinacao = {
                        'tempos': tempos_teste.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se chegou ao valor exato, para
                vazao_teste_3casas = vazao_teste.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                if vazao_teste_3casas == vazao_desejada_3casas:
                    print(f"   ✅ Vazão exata encontrada com incremento {float(incremento)}!")
                    return {
                        'tempos': tempos_teste.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se passou do valor desejado, para
                if vazao_teste > vazao_desejada:
                    print(f"   ⚠️  Vazão passou do desejado: {float(vazao_teste):.6f} > {float(vazao_desejada):.6f}")
                    break
        
        # Se a vazão atual é maior que a desejada, tenta diminuir
        else:
            iteracoes = 0
            max_iteracoes = 1000
            
            while iteracoes < max_iteracoes:
                iteracoes += 1
                
                # Tenta diminuir todos os tempos
                tempos_alterados = False
                for i in range(len(tempos_teste)):
                    novo_tempo = tempos_teste[i] - incremento
                    
                    if 239.599 <= float(novo_tempo) <= 240.499:
                        tempos_teste[i] = novo_tempo
                        tempos_alterados = True
                
                if not tempos_alterados:
                    print(f"   ⚠️  Todos os tempos atingiram o limite com incremento {float(incremento)}")
                    break
                
                # Calcula nova vazão
                resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_teste)
                agregados = calcular_agregados_com_tempo_ajustado(resultados)
                vazao_teste = agregados['vazao_media']
                
                # Verifica se é melhor
                diferenca_atual = abs(vazao_teste - vazao_desejada)
                if diferenca_atual < melhor_diferenca:
                    melhor_diferenca = diferenca_atual
                    melhor_tempos = tempos_teste.copy()
                    melhor_agregados = agregados
                    melhor_combinacao = {
                        'tempos': tempos_teste.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se chegou ao valor exato, para
                vazao_teste_3casas = vazao_teste.quantize(Decimal('0.001'), rounding=ROUND_HALF_UP)
                if vazao_teste_3casas == vazao_desejada_3casas:
                    print(f"   ✅ Vazão exata encontrada com incremento {float(incremento)}!")
                    return {
                        'tempos': tempos_teste.copy(),
                        'agregados': agregados,
                        'iteracoes': iteracoes
                    }
                
                # Se passou do valor desejado, para
                if vazao_teste < vazao_desejada:
                    print(f"   ⚠️  Vazão passou do desejado: {float(vazao_teste):.6f} < {float(vazao_desejada):.6f}")
                    break
    
    if melhor_combinacao:
        print(f"   ✅ Melhor aproximação encontrada!")
        print(f"   📊 Vazão final: {float(melhor_combinacao['agregados']['vazao_media']):.6f}")
        print(f"   📊 Diferença: {float(melhor_diferenca):.6f}")
        return melhor_combinacao
    else:
        print(f"   ❌ Não foi possível encontrar uma boa aproximação")
        return None

def gerar_tempos_iniciais():
    """
    Gera valores próximos de 240.000000 para aproximação inicial
    """
    # Valores próximos de 240.000000 para aproximação inicial
    valores_base = [
        Decimal('239.990'), Decimal('239.995'), Decimal('240.000'), Decimal('240.005'), Decimal('240.010'),
        Decimal('240.015'), Decimal('240.020'), Decimal('240.025'), Decimal('240.030'), Decimal('240.035'),
        Decimal('240.040'), Decimal('240.045'), Decimal('240.050'), Decimal('240.055'), Decimal('240.060'),
        Decimal('240.065'), Decimal('240.070'), Decimal('240.075'), Decimal('240.080'), Decimal('240.085'),
        Decimal('240.090'), Decimal('240.095'), Decimal('240.100'), Decimal('240.105'), Decimal('240.110'),
        Decimal('240.115'), Decimal('240.120'), Decimal('240.125'), Decimal('240.130'), Decimal('240.135'),
        Decimal('240.140'), Decimal('240.145'), Decimal('240.150'), Decimal('240.155'), Decimal('240.160'),
        Decimal('240.165'), Decimal('240.170'), Decimal('240.175'), Decimal('240.180'), Decimal('240.185'),
        Decimal('240.190'), Decimal('240.195'), Decimal('240.200'), Decimal('240.205'), Decimal('240.210'),
        Decimal('240.215'), Decimal('240.220'), Decimal('240.225'), Decimal('240.230'), Decimal('240.235'),
        Decimal('240.240'), Decimal('240.245'), Decimal('240.250'), Decimal('240.255'), Decimal('240.260'),
        Decimal('240.265'), Decimal('240.270'), Decimal('240.275'), Decimal('240.280'), Decimal('240.285'),
        Decimal('240.290'), Decimal('240.295'), Decimal('240.300'), Decimal('240.305'), Decimal('240.310'),
        Decimal('240.315'), Decimal('240.320'), Decimal('240.325'), Decimal('240.330'), Decimal('240.335'),
        Decimal('240.340'), Decimal('240.345'), Decimal('240.350'), Decimal('240.355'), Decimal('240.360'),
        Decimal('240.365'), Decimal('240.370'), Decimal('240.375'), Decimal('240.380'), Decimal('240.385'),
        Decimal('240.390'), Decimal('240.395'), Decimal('240.400'), Decimal('240.405'), Decimal('240.410'),
        Decimal('240.415'), Decimal('240.420'), Decimal('240.425'), Decimal('240.430'), Decimal('240.435'),
        Decimal('240.440'), Decimal('240.445'), Decimal('240.450'), Decimal('240.455'), Decimal('240.460'),
        Decimal('240.465'), Decimal('240.470'), Decimal('240.475'), Decimal('240.480'), Decimal('240.485'),
        Decimal('240.490'), Decimal('240.495'), Decimal('240.500'), Decimal('240.505'), Decimal('240.510')
    ]
    
    return valores_base

def otimizar_tempos_ponto_simples_240(leituras, constantes, valores_originais):
    """
    Define todos os tempos de coleta como 240.000 segundos
    Apenas faz ajustes proporcionais nos outros valores
    """
    print(f"   🔍 Definindo tempos como 240.000 para Ponto {leituras[0]['linha']}...")
    print(f"   🎯 OBJETIVO: Vazão média = {float(valores_originais['vazao_media']):.6f}")
    
    # Define todos os tempos como 240.000
    tempos_240 = [Decimal('240.000') for _ in range(3)]
    print(f"   📊 Tempos definidos: {[float(t) for t in tempos_240]}")
    
    # Calcula vazão com tempos 240.000
    resultados = calcular_formulas_com_tempo_ajustado(leituras, constantes, tempos_240)
    agregados = calcular_agregados_com_tempo_ajustado(resultados)
    vazao_atual = agregados['vazao_media']
    
    # Calcula diferença
    vazao_desejada = valores_originais['vazao_media']
    diferenca = vazao_atual - vazao_desejada
    
    print(f"   📊 Vazão com tempos 240.000: {float(vazao_atual):.6f}")
    print(f"   📊 Vazão desejada: {float(vazao_desejada):.6f}")
    print(f"   📊 Diferença: {float(diferenca):.6f} ({'POSITIVA' if diferenca > 0 else 'NEGATIVA'})")
    
    return {
        'tempos': tempos_240,
        'agregados': agregados,
        'iteracoes': 1,
        'diferenca': diferenca
    }

def gerar_planilha_corrigida(resultados_todos_pontos, arquivo_original, arquivo_corrigido):
    """
    Gera a planilha corrigida com os tempos otimizados aplicados
    AJUSTANDO PROPORCIONALMENTE os outros valores
    CALCULA DIFERENÇAS para orientar o refinamento
    """
    print(f"\n📄 GERANDO PLANILHA CORRIGIDA...")
    
    # Tenta criar cópia do arquivo original
    try:
        shutil.copy2(arquivo_original, arquivo_corrigido)
        print(f"   ✅ Arquivo copiado com sucesso: {arquivo_corrigido}")
    except PermissionError:
        print(f"   ⚠️  Erro de permissão ao copiar arquivo. Arquivo pode estar em uso.")
        print(f"   🔧 Tentando criar novo arquivo...")
        
        # Tenta criar um novo arquivo com nome diferente
        arquivo_corrigido = arquivo_corrigido.replace('.xlsx', '_NOVO.xlsx')
        try:
            shutil.copy2(arquivo_original, arquivo_corrigido)
            print(f"   ✅ Arquivo criado com sucesso: {arquivo_corrigido}")
        except Exception as e:
            print(f"   ❌ Erro ao criar arquivo: {e}")
            print(f"   💡 Feche o Excel e tente novamente")
            return False
    except Exception as e:
        print(f"   ❌ Erro inesperado ao copiar arquivo: {e}")
        return False
    
    # Carrega a planilha corrigida
    try:
        wb = load_workbook(arquivo_corrigido)
        coleta_sheet = wb["Coleta de Dados"]
    except Exception as e:
        print(f"   ❌ Erro ao carregar planilha: {e}")
        return False
    
    pontos_aplicados = 0
    informacoes_refinamento = []
    
    for resultado in resultados_todos_pontos:
        numero_ponto = resultado['numero']
        linha_inicial = resultado['linha_inicial']
        tempos_otimizados = resultado['tempos_otimizados']
        
        print(f"   📊 Aplicando Ponto {numero_ponto} (linha {linha_inicial})...")
        print(f"      Tempos otimizados: {tempos_otimizados}")
        
        # Usa a diferença já calculada pela função otimizar_tempos_ponto_simples_240
        vazao_original = resultado['valores_desejados']['vazao_media']
        vazao_otimizada = resultado['agregados_otimizados']['vazao_media']
        diferenca_vazao = resultado['diferenca']  # Usa a diferença já calculada
        
        print(f"      Vazão original: {float(vazao_original):.6f}")
        print(f"      Vazão otimizada: {float(vazao_otimizada):.6f}")
        print(f"      Diferença: {float(diferenca_vazao):.6f} ({'POSITIVA' if diferenca_vazao > 0 else 'NEGATIVA'})")
        
        # Verifica se a diferença é aceitável (menor que 1.0)
        if abs(diferenca_vazao) > 1.0:
            print(f"      ⚠️  DIFERENÇA MUITO ALTA! Ajustando proporcionalmente...")
        
        # Aplica os 3 tempos de coleta para o ponto AJUSTANDO PROPORCIONALMENTE
        for i, tempo_otimizado in enumerate(tempos_otimizados):
            linha = linha_inicial + i
            
            # Lê valores originais
            tempo_original = coleta_sheet.cell(row=linha, column=6).value  # Coluna F - Tempo original
            pulsos_original = coleta_sheet.cell(row=linha, column=3).value  # Coluna C - Pulsos
            leitura_medidor_original = coleta_sheet.cell(row=linha, column=15).value  # Coluna O - Leitura medidor
            temperatura_original = coleta_sheet.cell(row=linha, column=18).value  # Coluna R - Temperatura
            
            if tempo_original is not None and tempo_original != 0:
                # Calcula o fator de proporção
                fator_tempo = float(tempo_otimizado) / float(tempo_original)
                
                # LIMITA O FATOR PARA EVITAR AJUSTES EXTREMOS
                fator_maximo = 1.5  # Máximo 50% de variação
                fator_minimo = 0.5   # Mínimo 50% de variação
                
                if fator_tempo > fator_maximo:
                    print(f"      ⚠️  Fator muito alto ({fator_tempo:.3f}), limitando a {fator_maximo}")
                    fator_tempo = fator_maximo
                elif fator_tempo < fator_minimo:
                    print(f"      ⚠️  Fator muito baixo ({fator_tempo:.3f}), limitando a {fator_minimo}")
                    fator_tempo = fator_minimo
                
                # Ajusta os pulsos proporcionalmente
                if pulsos_original is not None:
                    pulsos_ajustados = int(float(pulsos_original) * fator_tempo)
                    pulsos_ajustados = max(1, pulsos_ajustados)  # Garante pelo menos 1 pulso
                    coleta_sheet.cell(row=linha, column=3).value = pulsos_ajustados
                
                # Ajusta a leitura do medidor proporcionalmente
                if leitura_medidor_original is not None:
                    leitura_medidor_ajustada = float(leitura_medidor_original) * fator_tempo
                    leitura_medidor_ajustada = max(0.1, leitura_medidor_ajustada)  # Garante valor mínimo
                    coleta_sheet.cell(row=linha, column=15).value = leitura_medidor_ajustada
                
                # Mantém a temperatura original
                if temperatura_original is not None:
                    coleta_sheet.cell(row=linha, column=18).value = float(temperatura_original)
                
                # Aplica o tempo otimizado
                coleta_sheet.cell(row=linha, column=6).value = float(tempo_otimizado)
                
                print(f"      Linha {linha}:")
                print(f"        Tempo: {float(tempo_original):.6f}s → {float(tempo_otimizado):.6f}s")
                print(f"        Pulsos: {pulsos_original} → {pulsos_ajustados}")
                print(f"        Leitura: {leitura_medidor_original:.2f} → {leitura_medidor_ajustada:.2f}")
                print(f"        Fator: {fator_tempo:.6f}")
            else:
                # Se não tem tempo original, apenas aplica o tempo otimizado
                coleta_sheet.cell(row=linha, column=6).value = float(tempo_otimizado)
                print(f"      Linha {linha}: {float(tempo_otimizado):.6f}s (sem ajuste proporcional)")
        
        # Salva informações para o refinamento
        # LÓGICA CORRIGIDA: Se a vazão atual é MENOR que a desejada, precisa INCREMENTAR os tempos
        # Se a vazão atual é MAIOR que a desejada, precisa DECREMENTAR os tempos
        direcao_correta = 'INCREMENTAR' if vazao_otimizada < vazao_original else 'DECREMENTAR'
        
        info_refinamento = {
            'numero': numero_ponto,
            'linha_inicial': linha_inicial,
            'tempos_otimizados': tempos_otimizados,
            'vazao_original': float(vazao_original),
            'vazao_otimizada': float(vazao_otimizada),
            'diferenca_vazao': float(diferenca_vazao),
            'direcao_refinamento': direcao_correta,
            'magnitude_diferenca': abs(float(diferenca_vazao))
        }
        informacoes_refinamento.append(info_refinamento)
        
        pontos_aplicados += 1
    
    # Salva a planilha corrigida
    try:
        wb.save(arquivo_corrigido)
        print(f"   ✅ Planilha salva com sucesso: {arquivo_corrigido}")
    except PermissionError:
        print(f"   ⚠️  Erro de permissão ao salvar planilha. Arquivo pode estar em uso.")
        print(f"   💡 Feche o Excel e tente novamente")
        return False
    except Exception as e:
        print(f"   ❌ Erro ao salvar planilha: {e}")
        return False
    
    # Salva informações de refinamento
    try:
        with open('informacoes_refinamento.json', 'w', encoding='utf-8') as f:
            json.dump(informacoes_refinamento, f, indent=2, ensure_ascii=False)
        print(f"   ✅ Informações de refinamento salvas: informacoes_refinamento.json")
    except Exception as e:
        print(f"   ❌ Erro ao salvar informações de refinamento: {e}")
        return False
    
    print(f"\n✅ Planilha corrigida gerada com sucesso!")
    print(f"   Pontos processados: {pontos_aplicados}")
    print(f"   Arquivo salvo: {arquivo_corrigido}")
    print(f"   Informações de refinamento salvas: informacoes_refinamento.json")
    
    return True

def main():
    """
    Função principal - PROCESSA TODOS OS PONTOS DA PLANILHA CORRIGIDA
    """
    arquivo_original = "SAN-038-25-09.xlsx"
    arquivo_corrigido = "SAN-038-25-09_CORRIGIDO.xlsx"
    
    if not os.path.exists(arquivo_original):
        print(f"❌ Arquivo original não encontrado: {arquivo_original}")
        return
    
    if not os.path.exists(arquivo_corrigido):
        print(f"❌ Arquivo corrigido não encontrado: {arquivo_corrigido}")
        return
    
    print("🚀 Iniciando otimização de tempos de coleta - VERSÃO SIMPLES...")
    print("=" * 60)
    print("🔧 PROCESSANDO PLANILHA CORRIGIDA!")
    print("=" * 60)
    
    # Extrai dados da planilha original (para obter valores desejados)
    constantes_original, pontos_original = extrair_dados_planilha_original(arquivo_original)
    if constantes_original is None or pontos_original is None:
        return
    
    print(f"✅ Extraídos {len(pontos_original)} pontos da planilha original")
    
    # Extrai dados da planilha corrigida (para obter valores atuais)
    constantes_corrigido, pontos_corrigido = extrair_dados_planilha_original(arquivo_corrigido)
    if constantes_corrigido is None or pontos_corrigido is None:
        return
    
    print(f"✅ Extraídos {len(pontos_corrigido)} pontos da planilha corrigida")
    
    tempo_inicio = time.time()
    resultados_todos_pontos = []
    
    # Processa todos os pontos
    for i, (ponto_original, ponto_corrigido) in enumerate(zip(pontos_original, pontos_corrigido)):
        print(f"\n🔍 PROCESSANDO Ponto {ponto_original['numero']} (linha {ponto_original['linha_inicial']})...")
        
        print(f"   📊 Vazão desejada (original): {float(ponto_original['valores_originais']['vazao_media']):.6f}")
        print(f"   📊 Vazão atual (corrigida): {float(ponto_corrigido['valores_originais']['vazao_media']):.6f}")
        
        # Define tempos como 240.000 e calcula diferença
        melhor_combinacao = otimizar_tempos_ponto_simples_240(
            ponto_corrigido['leituras'], 
            constantes_corrigido, 
            ponto_original['valores_originais']  # Usa valores originais como objetivo
        )
        
        if melhor_combinacao is None:
            print(f"❌ Não foi possível otimizar os tempos do Ponto {ponto_original['numero']}!")
            continue
        
        # Calcula resultados com tempos otimizados
        resultados_otimizados = calcular_formulas_com_tempo_ajustado(
            ponto_corrigido['leituras'], 
            constantes_corrigido, 
            melhor_combinacao['tempos']
        )
        
        # Verifica se os valores estão corretos
        vazao_diff = abs(float(melhor_combinacao['agregados']['vazao_media'] - ponto_original['valores_originais']['vazao_media']))
        tendencia_diff = abs(float(melhor_combinacao['agregados']['tendencia'] - ponto_original['valores_originais']['tendencia']))
        
        print(f"   📊 Vazão Média Desejada: {float(ponto_original['valores_originais']['vazao_media']):.6f}")
        print(f"   📊 Vazão Média Otimizada: {float(melhor_combinacao['agregados']['vazao_media']):.6f}")
        print(f"   📊 Diferença: {vazao_diff:.8f}")
        print(f"   📊 Tempos Otimizados: {[float(t) for t in melhor_combinacao['tempos']]}")
        print(f"   📊 Iterações necessárias: {melhor_combinacao['iteracoes']}")
        
        # Salva resultado do ponto
        resultado_ponto = {
            'numero': ponto_original['numero'],
            'linha_inicial': ponto_original['linha_inicial'],
            'tempos_otimizados': [float(t) for t in melhor_combinacao['tempos']],
            'agregados_otimizados': {k: float(v) if isinstance(v, Decimal) else v for k, v in melhor_combinacao['agregados'].items()},
            'valores_desejados': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto_original['valores_originais'].items()},
            'valores_corrigidos': {k: float(v) if isinstance(v, Decimal) else v for k, v in ponto_corrigido['valores_originais'].items()},
            'iteracoes': melhor_combinacao['iteracoes'],
            'diferenca': float(melhor_combinacao['diferenca']),  # Inclui a diferença calculada
            'diferencas': {
                'vazao': vazao_diff,
                'tendencia': tendencia_diff
            }
        }
        
        resultados_todos_pontos.append(resultado_ponto)
    
    tempo_decorrido = time.time() - tempo_inicio
    
    print(f"\n📊 RESUMO FINAL:")
    print(f"   Pontos processados: {len(resultados_todos_pontos)}/{len(pontos_original)}")
    print(f"   Tempo total: {tempo_decorrido:.2f} segundos")
    print(f"   Tempo médio por ponto: {tempo_decorrido/len(pontos_original):.2f} segundos")
    
    # Gera a planilha corrigida com os tempos otimizados
    sucesso_planilha = gerar_planilha_corrigida(resultados_todos_pontos, arquivo_original, arquivo_corrigido)
    
    if sucesso_planilha:
        # Salva resultado completo
        resultado_completo = {
            'constantes': {k: float(v) if isinstance(v, Decimal) else v for k, v in constantes_corrigido.items()},
            'pontos': resultados_todos_pontos,
            'resumo': {
                'total_pontos': len(pontos_original),
                'pontos_processados': len(resultados_todos_pontos),
                'tempo_total': tempo_decorrido,
                'tempo_medio_por_ponto': tempo_decorrido/len(pontos_original) if pontos_original else 0
            }
        }
        
        with open('resultados_otimizacao_tempos_corrigidos.json', 'w', encoding='utf-8') as f:
            json.dump(resultado_completo, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ Resultado completo salvo em: resultados_otimizacao_tempos_corrigidos.json")
        print(f"✅ Planilha corrigida gerada: {arquivo_corrigido}")
        print("🎉 Otimização de todos os pontos da planilha corrigida concluída!")
    else:
        print("❌ Erro ao gerar planilha corrigida!")

if __name__ == "__main__":
    main() 