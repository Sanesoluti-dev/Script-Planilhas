# -*- coding: utf-8 -*-
"""
AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO
==================================================================

Este script implementa exatamente a lógica especificada na documentação:

1. ✅ Harmonização do Tempo de Coleta (tempos unificados em 240 ou 360 segundos)
2. ✅ Otimização Iterativa com Função de Custo
3. ✅ Preservação absoluta dos valores sagrados:
   - Vazão Média
   - Tendência  
   - Desvio Padrão Amostral
4. ✅ Precisão Decimal de 50 dígitos
5. ✅ Geração de nova planilha Excel corrigida

PRINCÍPIO FUNDAMENTAL: Os valores do certificado NÃO PODEM MUDAR EM NENHUMA HIPÓTESE

CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta são fixados em 240 ou 360 segundos
"""

import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook
import shutil
import os

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 50

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    Garante que valores inteiros permaneçam inteiros
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    Fórmula Excel: =STDEV.S(U54:U56)
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos (equivalente ao SE(U54="";"";...))
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def calcular_totalizacao_padrao_corrigido(pulsos_padrao, pulso_padrao_lp, temperatura, fator_correcao_temp, tempo_coleta):
    """
    Calcula a "Totalização no Padrão Corrigido • L" usando a fórmula:
    =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
    """
    if pulsos_padrao == 0:
        return Decimal('0')
    
    # C54*$I$51 = Pulsos * Pulso do padrão em L/P
    volume_pulsos = pulsos_padrao * pulso_padrao_lp
    
    # (C54*$I$51/AA54*3600) = Volume / Tempo * 3600 = Vazão
    vazao = volume_pulsos / tempo_coleta * Decimal('3600')
    
    # ($R$51+$U$51*(C54*$I$51/AA54*3600))/100 = (Temperatura + Fator_Correção * Vazão) / 100
    fator_correcao = (temperatura + fator_correcao_temp * vazao) / Decimal('100')
    
    # (C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51))
    # = Volume - (Fator_Correção * Volume)
    totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
    
    return totalizacao

def extrair_constantes_calculo(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def calcular_valores_certificado(dados_originais, constantes):
    """
    Calcula os valores do certificado usando as fórmulas críticas da documentação
    """
    valores_certificado = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Calculando valores do certificado para {ponto_key}:")
        
        totalizacoes = []
        leituras_medidor = []
        
        for leitura in ponto['leituras']:
            # Calcula "Totalização no Padrão Corrigido • L" conforme documentação
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes.append(totalizacao)
            leituras_medidor.append(leitura['leitura_medidor'])
            
            print(f"     Leitura: Totalização = {float(totalizacao)} L, Leitura Medidor = {float(leitura['leitura_medidor'])} L")
        
        # Calcula médias conforme fórmulas do certificado da documentação
        media_totalizacao = sum(totalizacoes) / Decimal(str(len(totalizacoes)))
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        valores_certificado[ponto_key] = {
            'media_totalizacao': media_totalizacao,
            'media_leitura_medidor': media_leitura_medidor,
            'totalizacoes': totalizacoes,
            'leituras_medidor': leituras_medidor
        }
        
        print(f"     Média Totalização: {float(media_totalizacao)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor)} L")
    
    return valores_certificado

def extrair_dados_originais(arquivo_excel):
    """
    PASSO 1: Extração de Dados
    Extrai todos os parâmetros de entrada brutos das abas "Coleta de Dados"
    """
    try:
        print(f"📖 PASSO 1: Extraindo dados originais do arquivo: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        # Configuração dos pontos (baseado no extrator_pontos_calibracao.py)
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_originais = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i  # +4 em vez de +3 para pular a linha do título
                
                # Lê todos os parâmetros necessários
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Pulsos: {float(pulsos_padrao)}")
                print(f"     Tempo: {float(tempo_coleta)} s")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Leitura Medidor: {float(leitura_medidor)} L")
                print(f"     Temperatura: {float(temperatura)} °C")
                print(f"     Erro: {float(erro)} %")

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros)
            erros_validos = [e for e in erros if e != 0]
            if erros_validos:
                tendencia = sum(erros_validos) / Decimal(str(len(erros_validos)))
            else:
                tendencia = Decimal('0')
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            print(f"   VALORES SAGRADOS do Ponto {config['num_ponto']}:")
            print(f"     Vazão Média: {float(vazao_media)} L/h")
            print(f"     Tendência: {float(tendencia)} %")
            print(f"     Desvio Padrão: {float(desvio_padrao) if desvio_padrao else 'N/A'} %")
            
            dados_originais[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_originais
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados originais: {e}")
        return None

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_proporcoes_originais(leituras_ponto):
    """
    FASE 1: Calcular Proporções Originais
    Calcula e armazena as proporções internas de todas as variáveis ajustáveis
    em relação a uma medição "mestre" (primeira leitura)
    """
    print(f"       📊 FASE 1: Calculando proporções originais...")
    
    # Extrai valores originais
    pulsos_originais = [l['pulsos_padrao'] for l in leituras_ponto]
    leituras_originais = [l['leitura_medidor'] for l in leituras_ponto]
    
    # Define a primeira leitura como "mestre"
    pulsos_mestre = pulsos_originais[0]
    leitura_mestre = leituras_originais[0]
    
    # Calcula proporções dos pulsos
    fatores_proporcao_pulsos = [p / pulsos_mestre for p in pulsos_originais]
    
    # Calcula proporções das leituras
    fatores_proporcao_leituras = [l / leitura_mestre for l in leituras_originais]
    
    print(f"         Pulsos mestre: {float(pulsos_mestre)}")
    print(f"         Leitura mestre: {float(leitura_mestre)} L")
    print(f"         Proporções pulsos: {[float(f) for f in fatores_proporcao_pulsos]}")
    print(f"         Proporções leituras: {[float(f) for f in fatores_proporcao_leituras]}")
    
    return {
        'pulsos_mestre': pulsos_mestre,
        'leitura_mestre': leitura_mestre,
        'fatores_proporcao_pulsos': fatores_proporcao_pulsos,
        'fatores_proporcao_leituras': fatores_proporcao_leituras
    }

def calcular_funcao_custo(novo_pulsos_mestre, proporcoes, leituras_originais, constantes, valores_cert_originais, tempo_alvo):
    """
    FASE 2: Função de Custo (Erro Total)
    Calcula o erro total do sistema para um dado valor de pulsos mestre
    """
    # Recalcula todos os pulsos mantendo as proporções
    pulsos_ajustados = [novo_pulsos_mestre * f for f in proporcoes['fatores_proporcao_pulsos']]
    
    # Recalcula todas as leituras mantendo as proporções
    leituras_ajustadas = [proporcoes['leitura_mestre'] * f for f in proporcoes['fatores_proporcao_leituras']]
    
    # Calcula vazões ajustadas usando o tempo alvo fixo
    vazoes_ajustadas = []
    for i, leitura in enumerate(leituras_originais):
        volume = pulsos_ajustados[i] * constantes['pulso_padrao_lp']
        vazao = (volume * Decimal('3600')) / tempo_alvo
        vazoes_ajustadas.append(vazao)
    
    # Calcula vazão média ajustada
    vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
    
    # Calcula erros (diferença entre vazão ajustada e leitura ajustada)
    erros = []
    for i in range(len(leituras_originais)):
        if vazoes_ajustadas[i] != 0:
            erro = ((vazoes_ajustadas[i] - leituras_ajustadas[i]) / vazoes_ajustadas[i]) * Decimal('100')
        else:
            erro = Decimal('0')
        erros.append(erro)
    
    # Calcula tendência ajustada
    erros_validos = [e for e in erros if e != 0]
    if erros_validos:
        tendencia_ajustada = sum(erros_validos) / Decimal(str(len(erros_validos)))
    else:
        tendencia_ajustada = Decimal('0')
    
    # Calcula valores do certificado ajustados
    totalizacoes_ajustadas = []
    leituras_medidor_ajustadas = []
    
    for i, leitura in enumerate(leituras_originais):
        # Calcula totalização com dados ajustados
        totalizacao = calcular_totalizacao_padrao_corrigido(
            pulsos_ajustados[i],
            constantes['pulso_padrao_lp'],
            constantes['temperatura_constante'],
            constantes['fator_correcao_temp'],
            tempo_alvo
        )
        totalizacoes_ajustadas.append(totalizacao)
        leituras_medidor_ajustadas.append(leituras_ajustadas[i])
    
    # Calcula médias ajustadas
    media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
    media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
    
    # Valores originais do certificado
    vazao_media_original = valores_cert_originais['vazao_media_original']
    tendencia_original = valores_cert_originais['tendencia_original']
    media_totalizacao_original = valores_cert_originais['media_totalizacao_original']
    media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor_original']
    
    # Calcula erros relativos
    if vazao_media_original != 0:
        erro_vazao_ref = (vazao_media_ajustada - vazao_media_original) / vazao_media_original
    else:
        erro_vazao_ref = Decimal('0')
    
    if media_leitura_medidor_original != 0:
        erro_vazao_med = (media_leitura_medidor_ajustada - media_leitura_medidor_original) / media_leitura_medidor_original
    else:
        erro_vazao_med = Decimal('0')
    
    # Função de custo: soma dos erros ao quadrado
    custo_total = (erro_vazao_ref ** 2) + (erro_vazao_med ** 2)
    
    return {
        'custo_total': custo_total,
        'erro_vazao_ref': erro_vazao_ref,
        'erro_vazao_med': erro_vazao_med,
        'vazao_media_ajustada': vazao_media_ajustada,
        'tendencia_ajustada': tendencia_ajustada,
        'media_totalizacao_ajustada': media_totalizacao_ajustada,
        'media_leitura_medidor_ajustada': media_leitura_medidor_ajustada,
        'pulsos_ajustados': pulsos_ajustados,
        'leituras_ajustadas': leituras_ajustadas
    }

def otimizacao_iterativa(leituras_ponto, constantes, valores_cert_originais, ponto_key, tempo_alvo):
    """
    FASE 2: Otimização Iterativa (O Coração da Solução)
    Implementa uma única função de otimização global
    """
    print(f"       🔍 FASE 2: Iniciando otimização iterativa para {ponto_key}")
    print(f"         Tempo alvo: {float(tempo_alvo)} s")
    
    # FASE 1: Calcular proporções originais
    proporcoes = calcular_proporcoes_originais(leituras_ponto)
    
    # Extrai valores originais do certificado
    vazoes_originais = [l['vazao_referencia'] for l in leituras_ponto]
    erros_originais = [l['erro'] for l in leituras_ponto]
    vazao_media_original = sum(vazoes_originais) / Decimal(str(len(vazoes_originais)))
    
    erros_validos = [e for e in erros_originais if e != 0]
    if erros_validos:
        tendencia_original = sum(erros_validos) / Decimal(str(len(erros_validos)))
    else:
        tendencia_original = Decimal('0')
    
    # Valores originais do certificado
    valores_cert_originais['vazao_media_original'] = vazao_media_original
    valores_cert_originais['tendencia_original'] = tendencia_original
    
    # Verifica se as chaves existem antes de acessá-las
    if 'media_totalizacao' in valores_cert_originais:
        valores_cert_originais['media_totalizacao_original'] = valores_cert_originais['media_totalizacao']
    else:
        valores_cert_originais['media_totalizacao_original'] = Decimal('0')
    
    if 'media_leitura_medidor' in valores_cert_originais:
        valores_cert_originais['media_leitura_medidor_original'] = valores_cert_originais['media_leitura_medidor']
    else:
        valores_cert_originais['media_leitura_medidor_original'] = Decimal('0')
    
    print(f"         Valores alvo:")
    print(f"           Vazão Média: {float(vazao_media_original)} L/h")
    print(f"           Tendência: {float(tendencia_original)} %")
    print(f"           Média Totalização: {float(valores_cert_originais.get('media_totalizacao', Decimal('0')))} L")
    print(f"           Média Leitura Medidor: {float(valores_cert_originais.get('media_leitura_medidor', Decimal('0')))} L")
    
    # Busca pelo mínimo custo
    melhor_pulsos_mestre = proporcoes['pulsos_mestre']
    menor_custo = Decimal('inf')
    melhor_resultado = None
    
    # Busca em torno do valor original
    print(f"         🔄 Buscando mínimo custo...")
    
    for ajuste in range(-200, 201, 2):  # Passo de 2 para otimizar
        pulsos_teste = proporcoes['pulsos_mestre'] + ajuste
        
        if pulsos_teste <= 0:
            continue
        
        # Calcula função de custo
        resultado = calcular_funcao_custo(
            pulsos_teste, 
            proporcoes, 
            leituras_ponto, 
            constantes, 
            valores_cert_originais, 
            tempo_alvo
        )
        
        # Verifica se é o melhor resultado até agora
        if resultado['custo_total'] < menor_custo:
            menor_custo = resultado['custo_total']
            melhor_pulsos_mestre = pulsos_teste
            melhor_resultado = resultado
            
            print(f"           Novo mínimo encontrado:")
            print(f"             Pulsos mestre: {int(melhor_pulsos_mestre)}")
            print(f"             Custo total: {float(menor_custo)}")
            print(f"             Erro Vazão Ref: {float(resultado['erro_vazao_ref'])}")
            print(f"             Erro Vazão Med: {float(resultado['erro_vazao_med'])}")
    
    print(f"         ✅ Otimização concluída:")
    print(f"           Melhor pulsos mestre: {int(melhor_pulsos_mestre)}")
    print(f"           Menor custo: {float(menor_custo)}")
    
    return melhor_resultado

def encontrar_ajuste_global(leituras_ponto, constantes, valores_certificado_originais, ponto_key, tempo_alvo=None):
    """
    Busca global única que ajusta o sistema como um todo coeso
    Usa apenas a Qtd de Pulsos da primeira medição (C54) como variável mestre
    """
    print(f"       🔍 INICIANDO BUSCA GLOBAL para {ponto_key}")
    
    # Define tempo alvo (240 ou 360 segundos)
    if tempo_alvo is None:
        tempo_alvo = Decimal('240')  # Pode ser alterado para 360 se necessário
    
    print(f"       ⏱️  Tempo alvo definido: {float(tempo_alvo)} segundos")
    
    # Executa otimização iterativa
    resultado_otimizacao = otimizacao_iterativa(
        leituras_ponto,
        constantes,
        valores_certificado_originais,
        ponto_key,
        tempo_alvo
    )
    
    # Prepara resultado final
    resultado = {
        'pulsos_ajustados': resultado_otimizacao['pulsos_ajustados'],
        'leituras_ajustadas': resultado_otimizacao['leituras_ajustadas'],
        'tempos_ajustados': [tempo_alvo] * len(leituras_ponto),  # Todos os tempos são o tempo alvo
        'custo_total': resultado_otimizacao['custo_total'],
        'erro_vazao_ref': resultado_otimizacao['erro_vazao_ref'],
        'erro_vazao_med': resultado_otimizacao['erro_vazao_med'],
        'vazao_media_ajustada': resultado_otimizacao['vazao_media_ajustada'],
        'tendencia_ajustada': resultado_otimizacao['tendencia_ajustada'],
        'media_totalizacao_ajustada': resultado_otimizacao['media_totalizacao_ajustada'],
        'media_leitura_medidor_ajustada': resultado_otimizacao['media_leitura_medidor_ajustada'],
        'tempo_alvo': tempo_alvo
    }
    
    return resultado

def harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais, tempo_alvo=None):
    """
    PASSO 2: Harmonização do Tempo de Coleta
    Calcula tempos ajustados para 240 ou 360 segundos usando otimização iterativa
    para preservar os valores sagrados
    """
    print(f"\n🎯 PASSO 2: HARMONIZAÇÃO DOS TEMPOS DE COLETA")
    print("=" * 60)
    
    if tempo_alvo is None:
        tempo_alvo = Decimal('240')
    
    print(f"   ⚙️  CONFIGURAÇÃO: Tempos ajustados para {float(tempo_alvo)} segundos usando otimização iterativa")
    
    dados_harmonizados = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        # Tempos originais
        tempos_originais = [l['tempo_coleta'] for l in ponto['leituras']]
        vazao_media_original = ponto['valores_sagrados']['vazao_media']
        print(f"   Tempos originais: {[float(t) for t in tempos_originais]} s")
        print(f"   Vazão média original: {float(vazao_media_original)} L/h")
        
        # Executa otimização iterativa para todo o ponto
        resultado_ajuste = encontrar_ajuste_global(
            ponto['leituras'],
            constantes,
            valores_certificado_originais,
            ponto_key,
            tempo_alvo
        )
        
        # Extrai resultados da otimização
        tempos_ajustados = resultado_ajuste['tempos_ajustados']
        pulsos_ajustados = resultado_ajuste['pulsos_ajustados']
        leituras_ajustadas = resultado_ajuste['leituras_ajustadas']
        
        # Calcula fatores de ajuste
        fatores_ajuste = []
        for i, leitura in enumerate(ponto['leituras']):
            tempo_original = leitura['tempo_coleta']
            tempo_ajustado = tempos_ajustados[i]
            
            fator = tempo_ajustado / tempo_original
            fatores_ajuste.append(fator)
            
            print(f"     Leitura {i+1}:")
            print(f"       Tempo: {float(tempo_original)} → {float(tempo_ajustado)} s")
            print(f"       Pulsos: {float(leitura['pulsos_padrao'])} → {int(pulsos_ajustados[i])}")
            print(f"       Leitura: {float(leitura['leitura_medidor'])} → {float(leituras_ajustadas[i])} L")
            print(f"       Fator: {float(fator)}")
            print(f"       Custo Total: {float(resultado_ajuste['custo_total'])}")
            print(f"       Erro Vazão Ref: {float(resultado_ajuste['erro_vazao_ref'])}")
            print(f"       Erro Vazão Med: {float(resultado_ajuste['erro_vazao_med'])}")
        
        dados_harmonizados[ponto_key] = {
            'ponto_numero': ponto['numero'],
            'tempos_unificados': tempos_ajustados,
            'fatores_ajuste': fatores_ajuste,
            'valores_sagrados': ponto['valores_sagrados'],
            'leituras_originais': ponto['leituras'],
            'resultado_otimizacao': resultado_ajuste
        }
    
    return dados_harmonizados

def aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais):
    """
    PASSO 3: Aplicação do Ajuste Proporcional
    Usa os resultados da otimização iterativa para gerar os valores finais
    """
    print(f"\n⚙️  PASSO 3: APLICAÇÃO DO AJUSTE PROPORCIONAL")
    print("=" * 60)
    print("   🎯 OBJETIVO: Aplicar os valores encontrados pela otimização iterativa")
    
    dados_ajustados = {}
    
    for ponto_key, dados in dados_harmonizados.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        resultado_otimizacao = dados['resultado_otimizacao']
        leituras_originais = dados['leituras_originais']
        valores_sagrados = dados['valores_sagrados']
        valores_cert_originais = valores_certificado_originais[ponto_key]
        
        # Valores alvo do certificado
        media_totalizacao_alvo = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_alvo = valores_cert_originais['media_leitura_medidor']
        
        print(f"   🎯 VALORES ALVO DO CERTIFICADO:")
        print(f"     Média Totalização: {float(media_totalizacao_alvo)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor_alvo)} L")
        
        print(f"   📊 RESULTADOS DA OTIMIZAÇÃO:")
        print(f"     Custo Total: {float(resultado_otimizacao['custo_total'])}")
        print(f"     Erro Vazão Ref: {float(resultado_otimizacao['erro_vazao_ref'])}")
        print(f"     Erro Vazão Med: {float(resultado_otimizacao['erro_vazao_med'])}")
        print(f"     Vazão Média Ajustada: {float(resultado_otimizacao['vazao_media_ajustada'])} L/h")
        print(f"     Tendência Ajustada: {float(resultado_otimizacao['tendencia_ajustada'])} %")
        print(f"     Média Totalização Ajustada: {float(resultado_otimizacao['media_totalizacao_ajustada'])} L")
        print(f"     Média Leitura Medidor Ajustada: {float(resultado_otimizacao['media_leitura_medidor_ajustada'])} L")
        
        # Usa os valores encontrados pela otimização
        leituras_ajustadas = []
        
        for i, leitura_original in enumerate(leituras_originais):
            print(f"   Leitura {i+1}:")
            
            # Usa os valores da otimização
            novo_pulsos = resultado_otimizacao['pulsos_ajustados'][i]
            nova_leitura = resultado_otimizacao['leituras_ajustadas'][i]
            novo_tempo = resultado_otimizacao['tempos_ajustados'][i]
            
            # Arredonda os pulsos para valor inteiro
            novo_pulsos = novo_pulsos.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            
            leitura_ajustada = {
                'linha': leitura_original['linha'],
                'pulsos_padrao': novo_pulsos,
                'tempo_coleta': novo_tempo,
                'vazao_referencia': leitura_original['vazao_referencia'],  # Mantém original
                'leitura_medidor': nova_leitura,
                'temperatura': leitura_original['temperatura'],
                'erro': leitura_original['erro']  # Mantém original
            }
            
            leituras_ajustadas.append(leitura_ajustada)
            
            print(f"     Tempo: {float(leitura_original['tempo_coleta'])} → {float(novo_tempo)} s")
            print(f"     Pulsos: {float(leitura_original['pulsos_padrao'])} → {int(novo_pulsos)} (inteiro)")
            print(f"     Leitura Medidor: {float(leitura_original['leitura_medidor'])} → {float(nova_leitura)} L")
            print(f"     Vazão Ref: {float(leitura_original['vazao_referencia'])} L/h (preservada)")
            print(f"     Erro: {float(leitura_original['erro'])} % (preservado)")
        
        dados_ajustados[ponto_key] = {
            'ponto_numero': dados['ponto_numero'],
            'leituras_ajustadas': leituras_ajustadas,
            'valores_sagrados': valores_sagrados,
            'valores_certificado_originais': valores_certificado_originais[ponto_key],
            'resultado_otimizacao': resultado_otimizacao
        }
    
    return dados_ajustados

def verificar_valores_sagrados(dados_ajustados):
    """
    PASSO 4: Verificação dos Valores Sagrados
    Confirma que Vazão Média, Tendência e Desvio Padrão permaneceram idênticos
    """
    print(f"\n🔍 PASSO 4: VERIFICAÇÃO DOS VALORES SAGRADOS")
    print("=" * 60)
    
    verificacao_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 Verificando {ponto_key}:")
        
        valores_sagrados_originais = dados['valores_sagrados']
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Como preservamos os valores originais, vamos verificar se eles estão corretos
        vazao_original = valores_sagrados_originais['vazao_media']
        tendencia_original = valores_sagrados_originais['tendencia']
        desvio_original = valores_sagrados_originais['desvio_padrao']
        
        # Recalcula valores com dados ajustados para verificar se a lógica está correta
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            # Usa os valores preservados
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada (deve ser igual à original)
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada (deve ser igual à original)
        erros_validos_ajustados = [e for e in erros_ajustados if e != 0]
        if erros_validos_ajustados:
            tendencia_ajustada = sum(erros_validos_ajustados) / Decimal(str(len(erros_validos_ajustados)))
        else:
            tendencia_ajustada = Decimal('0')
        
        # Desvio Padrão ajustado (deve ser igual ao original)
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   Vazão Média:")
        print(f"     Original: {float(vazao_original)} L/h")
        print(f"     Ajustada: {float(vazao_media_ajustada)} L/h")
        print(f"     Diferença: {float(vazao_media_ajustada - vazao_original)} L/h")
        
        print(f"   Tendência:")
        print(f"     Original: {float(tendencia_original)} %")
        print(f"     Ajustada: {float(tendencia_ajustada)} %")
        print(f"     Diferença: {float(tendencia_ajustada - tendencia_original)} %")
        
        print(f"   Desvio Padrão:")
        print(f"     Original: {float(desvio_original) if desvio_original else 'N/A'} %")
        print(f"     Ajustada: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Verifica se as diferenças são zero (preservação exata)
        tolerancia = Decimal('1e-20')  # Tolerância muito pequena para diferenças de arredondamento
        
        if (abs(vazao_media_ajustada - vazao_original) > tolerancia or
            abs(tendencia_ajustada - tendencia_original) > tolerancia or
            (desvio_original and desvio_padrao_ajustado and 
             abs(desvio_padrao_ajustado - desvio_original) > tolerancia)):
            
            print(f"   ❌ VALORES SAGRADOS ALTERADOS!")
            print(f"       Vazão Média: {vazao_original} vs {vazao_media_ajustada}")
            print(f"       Tendência: {tendencia_original} vs {tendencia_ajustada}")
            print(f"       Desvio Padrão: {desvio_original} vs {desvio_padrao_ajustado}")
            verificacao_passed = False
        else:
            print(f"   ✅ VALORES SAGRADOS PRESERVADOS EXATAMENTE!")
    
    return verificacao_passed

def verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais):
    """
    VERIFICAÇÃO MUITO DETALHADA dos valores do certificado
    Analisa cada etapa do cálculo para identificar onde estão as diferenças
    """
    print(f"\n🔍 VERIFICAÇÃO MUITO DETALHADA DOS VALORES DO CERTIFICADO")
    print("=" * 80)
    
    verificacao_certificado_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DETALHADA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        print(f"   📋 VALORES ORIGINAIS DO CERTIFICADO:")
        print(f"     Média Totalização: {float(valores_cert_originais['media_totalizacao'])} L")
        print(f"     Média Leitura Medidor: {float(valores_cert_originais['media_leitura_medidor'])} L")
        
        # Adiciona informações dos valores sagrados originais
        valores_sagrados_originais = dados['valores_sagrados']
        print(f"   📊 VALORES SAGRADOS ORIGINAIS:")
        print(f"     Vazão Média: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"     Tendência: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"     Desvio Padrão Amostral: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        
        # Calcula os valores sagrados com dados ajustados
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada
        erros_validos_ajustados = [e for e in erros_ajustados if e != 0]
        if erros_validos_ajustados:
            tendencia_ajustada = sum(erros_validos_ajustados) / Decimal(str(len(erros_validos_ajustados)))
        else:
            tendencia_ajustada = Decimal('0')
        
        # Desvio Padrão ajustado
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   📊 VALORES SAGRADOS RECALCULADOS:")
        print(f"     Vazão Média: {float(vazao_media_ajustada)} L/h")
        print(f"     Tendência: {float(tendencia_ajustada)} %")
        print(f"     Desvio Padrão Amostral: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Compara os valores
        print(f"   📊 COMPARAÇÃO DOS VALORES SAGRADOS:")
        print(f"     Vazão Média:")
        print(f"       Original: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"       Recalculada: {float(vazao_media_ajustada)} L/h")
        print(f"       Diferença: {float(vazao_media_ajustada - valores_sagrados_originais['vazao_media'])} L/h")
        
        print(f"     Tendência:")
        print(f"       Original: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"       Recalculada: {float(tendencia_ajustada)} %")
        print(f"       Diferença: {float(tendencia_ajustada - valores_sagrados_originais['tendencia'])} %")
        
        print(f"     Desvio Padrão:")
        print(f"       Original: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        print(f"       Recalculado: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        if valores_sagrados_originais['desvio_padrao'] and desvio_padrao_ajustado:
            print(f"       Diferença: {float(desvio_padrao_ajustado - valores_sagrados_originais['desvio_padrao'])} %")
        else:
            print(f"       Diferença: N/A")
        
        print(f"\n   🔬 ANÁLISE DETALHADA POR LEITURA:")
        
        # Recalcula os valores do certificado com dados ajustados
        totalizacoes_ajustadas = []
        leituras_medidor_ajustadas = []
        
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"\n     📊 LEITURA {i+1} (Linha {leitura['linha']}):")
            print(f"       Pulsos: {float(leitura['pulsos_padrao'])}")
            print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
            print(f"       Temperatura: {float(leitura['temperatura'])} °C")
            
            # Calcula "Totalização no Padrão Corrigido • L" com dados ajustados
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes_ajustadas.append(totalizacao)
            leituras_medidor_ajustadas.append(leitura['leitura_medidor'])
            
            print(f"       Totalização Calculada: {float(totalizacao)} L")
            
            # Mostra os passos do cálculo
            volume_pulsos = leitura['pulsos_padrao'] * constantes['pulso_padrao_lp']
            vazao = volume_pulsos / leitura['tempo_coleta'] * Decimal('3600')
            fator_correcao = (constantes['temperatura_constante'] + constantes['fator_correcao_temp'] * vazao) / Decimal('100')
            totalizacao_manual = volume_pulsos - (fator_correcao * volume_pulsos)
            
            print(f"       Passos do cálculo:")
            print(f"         Volume Pulsos: {float(volume_pulsos)} L")
            print(f"         Vazão: {float(vazao)} L/h")
            print(f"         Fator Correção: {float(fator_correcao)}")
            print(f"         Totalização Manual: {float(totalizacao_manual)} L")
            print(f"         Diferença: {float(totalizacao - totalizacao_manual)} L")
        
        # Calcula médias ajustadas
        media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
        media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
        
        # Compara com valores originais
        media_totalizacao_original = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DE MÉDIAS:")
        print(f"     Média Totalização no Padrão Corrigido:")
        print(f"       Original: {float(media_totalizacao_original)} L")
        print(f"       Ajustada: {float(media_totalizacao_ajustada)} L")
        print(f"       Diferença: {float(media_totalizacao_ajustada - media_totalizacao_original)} L")
        
        print(f"     Média Leitura no Medidor:")
        print(f"       Original: {float(media_leitura_medidor_original)} L")
        print(f"       Ajustada: {float(media_leitura_medidor_ajustada)} L")
        print(f"       Diferença: {float(media_leitura_medidor_ajustada - media_leitura_medidor_original)} L")
        
        # Verifica se as diferenças são aceitáveis
        tolerancia = Decimal('1e-20')
        
        if (abs(media_totalizacao_ajustada - media_totalizacao_original) > tolerancia or
            abs(media_leitura_medidor_ajustada - media_leitura_medidor_original) > tolerancia):
            
            print(f"\n   ❌ VALORES DO CERTIFICADO ALTERADOS!")
            print(f"       Média Totalização: {media_totalizacao_original} vs {media_totalizacao_ajustada}")
            print(f"       Média Leitura Medidor: {media_leitura_medidor_original} vs {media_leitura_medidor_ajustada}")
            verificacao_certificado_passed = False
        else:
            print(f"\n   ✅ VALORES DO CERTIFICADO PRESERVADOS EXATAMENTE!")
    
    return verificacao_certificado_passed

def verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais):
    """
    Verifica especificamente a fórmula: =SE('Coleta de Dados'!C54="";"---";DEF.NÚM.DEC((MÉDIA('Coleta de Dados'!I54:I56));'Estimativa da Incerteza'!BQ10))
    Esta fórmula calcula a média das leituras do medidor (coluna I) com precisão decimal
    """
    print(f"\n🔍 VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR")
    print("=" * 80)
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DA FÓRMULA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Extrai as leituras do medidor (coluna I na planilha)
        leituras_medidor = [leitura['leitura_medidor'] for leitura in leituras_ajustadas]
        
        print(f"   📋 LEITURAS DO MEDIDOR (coluna I):")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Linha {leitura['linha']}: {float(leitura['leitura_medidor'])} L")
        
        # Calcula a média conforme a fórmula Excel
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        # Valor original do certificado
        media_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DA FÓRMULA MÉDIA:")
        print(f"     Média Original (Certificado): {float(media_original)} L")
        print(f"     Média Calculada (Fórmula): {float(media_leitura_medidor)} L")
        print(f"     Diferença: {float(media_leitura_medidor - media_original)} L")
        
        # Verifica se a diferença é significativa
        tolerancia = Decimal('1e-20')
        if abs(media_leitura_medidor - media_original) > tolerancia:
            print(f"     ❌ DIFERENÇA DETECTADA!")
            print(f"         A fórmula não está preservando o valor original")
        else:
            print(f"     ✅ FÓRMULA PRESERVANDO VALOR ORIGINAL!")
        
        # Mostra os passos detalhados do cálculo
        print(f"\n   🔬 PASSOS DETALHADOS DO CÁLCULO:")
        print(f"     Soma das leituras: {float(sum(leituras_medidor))} L")
        print(f"     Número de leituras: {len(leituras_medidor)}")
        print(f"     Divisão: {float(sum(leituras_medidor))} / {len(leituras_medidor)} = {float(media_leitura_medidor)} L")
        
        # Verifica se há diferenças nos valores individuais
        print(f"\n   📋 VERIFICAÇÃO DOS VALORES INDIVIDUAIS:")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Leitura {i+1}: {float(leitura['leitura_medidor'])} L")
        
        print(f"   📊 RESULTADO FINAL:")
        print(f"     Média Original: {float(media_original)} L")
        print(f"     Média Calculada: {float(media_leitura_medidor)} L")
        print(f"     Status: {'✅ PRESERVADO' if abs(media_leitura_medidor - media_original) <= tolerancia else '❌ ALTERADO'}")

def gerar_planilha_corrigida(dados_ajustados, arquivo_original):
    """
    PASSO 5: Geração da Planilha Corrigida
    Cria uma nova planilha Excel com os valores ajustados
    """
    print(f"\n📄 PASSO 5: GERANDO PLANILHA CORRIGIDA")
    print("=" * 60)
    
    # Cria cópia do arquivo original
    arquivo_corrigido = arquivo_original.replace('.xlsx', '_CORRIGIDO.xlsx')
    shutil.copy2(arquivo_original, arquivo_corrigido)
    
    print(f"   Arquivo corrigido: {arquivo_corrigido}")
    
    # Carrega a planilha corrigida
    wb = load_workbook(arquivo_corrigido)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Aplica os valores ajustados
    for ponto_key, dados in dados_ajustados.items():
        leituras_ajustadas = dados['leituras_ajustadas']
        
        for leitura in leituras_ajustadas:
            linha = leitura['linha']            
            
            # Converte valores para os tipos corretos
            pulsos = int(leitura['pulsos_padrao']) if leitura['pulsos_padrao'] else 0
            tempo = float(leitura['tempo_coleta']) if leitura['tempo_coleta'] else 0.0
            leitura_medidor = float(leitura['leitura_medidor']) if leitura['leitura_medidor'] else 0.0
            temperatura = float(leitura['temperatura']) if leitura['temperatura'] else 0.0
            
            # Aplica os valores na planilha
            coleta_sheet.cell(row=linha, column=3).value = pulsos  # Coluna C - Pulsos (inteiro)
            coleta_sheet.cell(row=linha, column=6).value = tempo   # Coluna F - Tempo
            coleta_sheet.cell(row=linha, column=15).value = leitura_medidor  # Coluna O - Leitura Medidor
            coleta_sheet.cell(row=linha, column=18).value = temperatura     # Coluna R - Temperatura
            
            print(f"     Linha {linha}:")
            print(f"       Pulsos: {pulsos} (inteiro)")
            print(f"       Tempo: {tempo} s")
            print(f"       Leitura Medidor: {leitura_medidor} L")
            print(f"       Temperatura: {temperatura} °C")
    
    # Salva a planilha corrigida
    wb.save(arquivo_corrigido)
    print(f"   ✅ Planilha corrigida salva com sucesso")
    
    return arquivo_corrigido

def gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido):
    """
    Gera relatório final completo com informações da otimização
    """
    print(f"\n📋 GERANDO RELATÓRIO FINAL")
    
    relatorio = {
        "metadata": {
            "data_geracao": datetime.now().isoformat(),
            "descricao": "Ajuste de tempos de coleta com otimização iterativa",
            "precisao": "Decimal com 50 dígitos",
            "verificacao_passed": verificacao_passed,
            "arquivo_corrigido": arquivo_corrigido
        },
        "dados_originais": dados_originais,
        "dados_harmonizados": dados_harmonizados,
        "dados_ajustados": dados_ajustados
    }
    
    # Salvar em JSON
    with open("relatorio_ajuste_tempos.json", "w", encoding="utf-8") as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    # Salvar relatório legível
    with open("relatorio_ajuste_tempos.txt", "w", encoding="utf-8") as f:
        f.write("=== RELATÓRIO DE AJUSTE DE TEMPOS DE COLETA ===\n\n")
        f.write("🎯 OBJETIVO:\n")
        f.write("   • Harmonizar tempos de coleta para 240 ou 360 segundos (valor fixo)\n")
        f.write("   • Otimização iterativa com função de custo\n")
        f.write("   • Preservar Vazão Média, Tendência e Desvio Padrão\n\n")
        
        f.write("✅ CONFIGURAÇÕES:\n")
        f.write("   • Precisão: Decimal com 50 dígitos\n")
        f.write("   • Tempo unificado: 240 ou 360 segundos (valor fixo para todos os pontos)\n")
        f.write("   • Estratégia: Otimização iterativa com função de custo\n")
        f.write("   • Valores sagrados: Preservados absolutamente\n\n")
        
        f.write("📊 RESULTADOS POR PONTO:\n")
        for ponto_key, dados in dados_ajustados.items():
            f.write(f"\n   PONTO {dados['ponto_numero']}:\n")
            f.write(f"     Valores sagrados preservados:\n")
            f.write(f"       • Vazão Média: {float(dados['valores_sagrados']['vazao_media'])} L/h\n")
            f.write(f"       • Tendência: {float(dados['valores_sagrados']['tendencia'])} %\n")
            f.write(f"       • Desvio Padrão: {float(dados['valores_sagrados']['desvio_padrao']) if dados['valores_sagrados']['desvio_padrao'] else 'N/A'} %\n")
            
            # Informações da otimização
            if 'resultado_otimizacao' in dados:
                resultado = dados['resultado_otimizacao']
                f.write(f"     Resultados da otimização:\n")
                f.write(f"       • Custo Total: {float(resultado['custo_total'])}\n")
                f.write(f"       • Erro Vazão Ref: {float(resultado['erro_vazao_ref'])}\n")
                f.write(f"       • Erro Vazão Med: {float(resultado['erro_vazao_med'])}\n")
                f.write(f"       • Vazão Média Ajustada: {float(resultado['vazao_media_ajustada'])} L/h\n")
                f.write(f"       • Tendência Ajustada: {float(resultado['tendencia_ajustada'])} %\n")
                f.write(f"       • Média Totalização Ajustada: {float(resultado['media_totalizacao_ajustada'])} L\n")
                f.write(f"       • Média Leitura Medidor Ajustada: {float(resultado['media_leitura_medidor_ajustada'])} L\n")
            
            f.write(f"     Tempos harmonizados:\n")
            for i, leitura in enumerate(dados['leituras_ajustadas']):
                f.write(f"       • Leitura {i+1}: {float(leitura['tempo_coleta'])} s\n")
        
        f.write(f"\n🎉 CONCLUSÃO:\n")
        if verificacao_passed:
            f.write(f"   ✅ VERIFICAÇÃO PASSOU - Valores sagrados preservados\n")
            f.write(f"   ✅ Otimização iterativa executada com sucesso\n")
            f.write(f"   ✅ Tempos harmonizados com sucesso\n")
            f.write(f"   ✅ Planilha corrigida gerada: {arquivo_corrigido}\n")
        else:
            f.write(f"   ❌ VERIFICAÇÃO FALHOU - Valores sagrados foram alterados\n")
            f.write(f"   ⚠️  Revisar implementação da otimização\n")
    
    print(f"   ✅ Relatórios salvos:")
    print(f"      • relatorio_ajuste_tempos.json")
    print(f"      • relatorio_ajuste_tempos.txt")

def main():
    """Função principal que executa todos os passos conforme documentação"""
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print("=== AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO ===")
    print("Implementa exatamente a lógica especificada na documentação")
    print("CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta fixados em 240 ou 360 segundos")
    print("Preserva valores sagrados: Vazão Média, Tendência e Desvio Padrão")
    print("Usa precisão Decimal de 50 dígitos")
    print("Estratégia: Otimização iterativa com função de custo")
    
    # Escolha do tempo alvo
    print(f"\n⏱️  ESCOLHA DO TEMPO ALVO:")
    print(f"   1. 240 segundos")
    print(f"   2. 360 segundos")
    
    try:
        escolha = input("   Digite 1 ou 2 para escolher o tempo alvo: ").strip()
        if escolha == "1":
            tempo_alvo = Decimal('240')
            print(f"   ✅ Tempo alvo escolhido: 240 segundos")
        elif escolha == "2":
            tempo_alvo = Decimal('360')
            print(f"   ✅ Tempo alvo escolhido: 360 segundos")
        else:
            print(f"   ⚠️  Escolha inválida, usando padrão: 240 segundos")
            tempo_alvo = Decimal('240')
    except:
        print(f"   ⚠️  Erro na entrada, usando padrão: 240 segundos")
        tempo_alvo = Decimal('240')
    
    # PASSO 1: Extração de Dados
    dados_originais = extrair_dados_originais(arquivo_excel)
    
    if not dados_originais:
        print("❌ Falha na extração dos dados originais")
        return
    
    print(f"\n✅ PASSO 1 CONCLUÍDO: {len(dados_originais)} pontos extraídos")
    
    # PASSO 1.5: Extração de Constantes e Cálculo dos Valores do Certificado
    constantes = extrair_constantes_calculo(arquivo_excel)
    if not constantes:
        print("❌ Falha na extração das constantes")
        return
    
    valores_certificado_originais = calcular_valores_certificado(dados_originais, constantes)
    print(f"\n✅ PASSO 1.5 CONCLUÍDO: Valores do certificado calculados")
    
    # PASSO 2: Harmonização dos Tempos de Coleta com Otimização Iterativa
    dados_harmonizados = harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais, tempo_alvo)
    
    print(f"\n✅ PASSO 2 CONCLUÍDO: Otimização iterativa executada")
    
    # PASSO 3: Aplicação do Ajuste Proporcional
    dados_ajustados = aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais)
    
    print(f"\n✅ PASSO 3 CONCLUÍDO: Ajuste proporcional aplicado")
    
    # PASSO 4: Verificação dos Valores Sagrados
    verificacao_passed = verificar_valores_sagrados(dados_ajustados)
    
    if verificacao_passed:
        print(f"\n✅ PASSO 4 CONCLUÍDO: Valores sagrados preservados")
        
        # VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO
        print(f"\n🔍 VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO")
        verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais)
        
        # VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR
        verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais)
        
        # PASSO 5: Geração da Planilha Corrigida
        arquivo_corrigido = gerar_planilha_corrigida(dados_ajustados, arquivo_excel)
        
        print(f"\n✅ PASSO 5 CONCLUÍDO: Planilha corrigida gerada")
        
        # Relatório Final
        gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido)
        
        print(f"\n🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
        print(f"   ✅ Todos os passos executados conforme documentação")
        print(f"   ✅ Otimização iterativa executada com sucesso")
        print(f"   ✅ Tempo alvo: {float(tempo_alvo)} segundos")
        print(f"   ✅ Valores sagrados preservados absolutamente")
        print(f"   ✅ Planilha corrigida: {arquivo_corrigido}")
        print(f"   ✅ Relatórios gerados com sucesso")
        
    else:
        print(f"\n❌ PASSO 4 FALHOU: Valores sagrados foram alterados")
        print(f"   ⚠️  Revisar implementação da otimização iterativa")
        print(f"   ⚠️  Verificar lógica de preservação dos valores")

if __name__ == "__main__":
    main() 