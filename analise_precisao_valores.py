# -*- coding: utf-8 -*-
"""
ANÁLISE DE PRECISÃO E LEITURA DE VALORES
========================================

Este script analisa como cada script está lendo valores numéricos
e verifica a precisão máxima que conseguem alcançar.
"""

import pandas as pd
from decimal import Decimal, getcontext
from openpyxl import load_workbook
import json

# Configurar precisão máxima para análise
getcontext().prec = 50

def analisar_valores_planilha():
    """Analisa os valores reais da planilha para verificar formato e precisão"""
    print("=== ANÁLISE DOS VALORES REAIS DA PLANILHA ===")
    
    wb = load_workbook("SAN-038-25-09-1.xlsx", data_only=True)
    coleta = wb["Coleta de Dados"]
    cert = wb["Emissão do Certificado"]
    
    # Valores importantes para análise
    valores_teste = [
        ("Coleta C54", coleta, 54, 3),   # Pulsos padrão
        ("Coleta F54", coleta, 54, 6),   # Tempo de coleta
        ("Coleta I54", coleta, 54, 9),   # Volume corrigido
        ("Coleta O54", coleta, 54, 15),  # Leitura medidor
        ("Coleta R54", coleta, 54, 18),  # Temperatura
        ("Cert F74", cert, 74, 6),       # Vazão padrão certificado
        ("Cert I74", cert, 74, 9),       # Volume corrigido certificado
        ("Cert O74", cert, 74, 15),      # Vazão medidor certificado
    ]
    
    print(f"{'Localização':<20} {'Valor Bruto':<25} {'Tipo':<15} {'Representação':<30} {'Casas Decimais':<15}")
    print("-" * 110)
    
    max_casas = 0
    valores_analisados = {}
    
    for nome, sheet, linha, coluna in valores_teste:
        valor_bruto = sheet.cell(row=linha, column=coluna).value
        
        # Análise detalhada do valor
        if valor_bruto is None:
            tipo_valor = "None"
            representacao = "None"
            casas_decimais = 0
        elif isinstance(valor_bruto, str):
            tipo_valor = "String"
            representacao = repr(valor_bruto)
            # Conta casas decimais na string
            if ',' in valor_bruto:
                partes = valor_bruto.split(',')
                if len(partes) > 1:
                    casas_decimais = len(partes[1])
                else:
                    casas_decimais = 0
            elif '.' in valor_bruto:
                partes = valor_bruto.split('.')
                if len(partes) > 1:
                    casas_decimais = len(partes[1])
                else:
                    casas_decimais = 0
            else:
                casas_decimais = 0
        elif isinstance(valor_bruto, (int, float)):
            tipo_valor = type(valor_bruto).__name__
            representacao = repr(valor_bruto)
            # Para números, conta casas decimais
            str_valor = str(valor_bruto)
            if '.' in str_valor:
                partes = str_valor.split('.')
                casas_decimais = len(partes[1])
            else:
                casas_decimais = 0
        else:
            tipo_valor = type(valor_bruto).__name__
            representacao = repr(valor_bruto)
            casas_decimais = 0
        
        print(f"{nome:<20} {str(valor_bruto):<25} {tipo_valor:<15} {representacao:<30} {casas_decimais:<15}")
        
        valores_analisados[nome] = {
            'valor_bruto': valor_bruto,
            'tipo': tipo_valor,
            'representacao': representacao,
            'casas_decimais': casas_decimais
        }
        
        if casas_decimais > max_casas:
            max_casas = casas_decimais
    
    print(f"\n📊 RESUMO DA ANÁLISE:")
    print(f"   • Máximo de casas decimais encontradas: {max_casas}")
    print(f"   • Valores com vírgula como separador: {sum(1 for v in valores_analisados.values() if isinstance(v['valor_bruto'], str) and ',' in str(v['valor_bruto']))}")
    print(f"   • Valores com ponto como separador: {sum(1 for v in valores_analisados.values() if isinstance(v['valor_bruto'], str) and '.' in str(v['valor_bruto']))}")
    
    return valores_analisados, max_casas

def testar_funcoes_conversao():
    """Testa as diferentes funções de conversão dos scripts"""
    print("\n=== TESTE DAS FUNÇÕES DE CONVERSÃO ===")
    
    # Valores de teste com diferentes formatos
    valores_teste = [
        "33.957,90",    # Formato brasileiro
        "33957.90",     # Formato americano
        "33,957.90",    # Formato com vírgula de milhares
        "33957,90",     # Formato brasileiro sem pontos
        "0.19488768926220715",  # Número com muitas casas
        "0,19488768926220715",  # Número brasileiro com muitas casas
        33858.49554066015,      # Float
        7947,                   # Inteiro
        None,                   # Valor nulo
    ]
    
    print(f"{'Valor Original':<25} {'Tipo':<10} {'Simulação':<25} {'Extrator':<25} {'Script Corrigido':<25}")
    print("-" * 110)
    
    # Função do simulacao.py
    def converter_para_decimal_simulacao(valor):
        if valor is None:
            return Decimal('0')
        if isinstance(valor, str):
            valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
            return Decimal(valor_limpo)
        return Decimal(str(valor))
    
    # Função do extrator_pontos_calibracao.py
    def get_numeric_value_extrator(valor):
        if valor is None:
            return Decimal('0')
        return Decimal(str(valor))
    
    # Função do script_corrigido.py
    def get_numeric_value_script(valor):
        if valor is None:
            return None
        if isinstance(valor, (int, float)):
            return float(valor)
        if isinstance(valor, str):
            valor_limpo = valor.strip().replace(',', '.')
            if valor_limpo.replace('.', '').replace('-', '').isdigit():
                return float(valor_limpo)
        return None
    
    for valor in valores_teste:
        try:
            sim = converter_para_decimal_simulacao(valor)
            ext = get_numeric_value_extrator(valor)
            scr = get_numeric_value_script(valor)
            
            print(f"{str(valor):<25} {type(valor).__name__:<10} {str(sim):<25} {str(ext):<25} {str(scr):<25}")
        except Exception as e:
            print(f"{str(valor):<25} {type(valor).__name__:<10} {'ERRO':<25} {'ERRO':<25} {'ERRO':<25}")

def testar_precisao_decimal():
    """Testa a precisão máxima do Decimal"""
    print("\n=== TESTE DE PRECISÃO DECIMAL ===")
    
    # Testa diferentes precisões
    precisoes = [15, 28, 50, 100]
    
    for prec in precisoes:
        getcontext().prec = prec
        valor_teste = Decimal('0.19488768926220715')
        print(f"Precisão {prec:3d}: {valor_teste}")
    
    # Testa com valores da planilha
    print(f"\n--- TESTE COM VALORES REAIS DA PLANILHA ---")
    wb = load_workbook("SAN-038-25-09-1.xlsx", data_only=True)
    cert = wb["Emissão do Certificado"]
    
    # Valor do certificado
    valor_cert = cert.cell(row=74, column=15).value  # O74
    print(f"Valor bruto do certificado: {valor_cert} (tipo: {type(valor_cert)})")
    
    # Testa conversão com diferentes precisões
    for prec in [15, 28, 50]:
        getcontext().prec = prec
        try:
            if isinstance(valor_cert, str):
                # Remove pontos de milhares e substitui vírgula por ponto
                valor_limpo = valor_cert.replace(' ', '').replace('.', '').replace(',', '.')
                decimal_valor = Decimal(valor_limpo)
            else:
                decimal_valor = Decimal(str(valor_cert))
            print(f"Precisão {prec:2d}: {decimal_valor}")
        except Exception as e:
            print(f"Precisão {prec:2d}: ERRO - {e}")

def analisar_scripts_existentes():
    """Analisa como os scripts existentes tratam os valores"""
    print("\n=== ANÁLISE DOS SCRIPTS EXISTENTES ===")
    
    scripts_info = {
        "simulacao.py": {
            "precisao": 28,
            "funcao_conversao": "converter_para_decimal",
            "tratamento_virgula": "Remove pontos de milhares e substitui vírgula por ponto",
            "biblioteca": "openpyxl"
        },
        "extrator_pontos_calibracao.py": {
            "precisao": 15,
            "funcao_conversao": "get_numeric_value",
            "tratamento_virgula": "Usa Decimal(str(valor)) diretamente",
            "biblioteca": "pandas"
        },
        "script_corrigido.py": {
            "precisao": 28,
            "funcao_conversao": "get_numeric_value",
            "tratamento_virgula": "Substitui vírgula por ponto, retorna float",
            "biblioteca": "pandas"
        },
        "extrator_dados_certificado.py": {
            "precisao": "Não especificada",
            "funcao_conversao": "Não tem conversão específica",
            "tratamento_virgula": "Usa pandas diretamente",
            "biblioteca": "pandas"
        }
    }
    
    print(f"{'Script':<30} {'Precisão':<15} {'Função':<25} {'Tratamento Vírgula':<40}")
    print("-" * 110)
    
    for script, info in scripts_info.items():
        print(f"{script:<30} {str(info['precisao']):<15} {info['funcao_conversao']:<25} {info['tratamento_virgula']:<40}")

def gerar_recomendacoes():
    """Gera recomendações baseadas na análise"""
    print("\n=== RECOMENDAÇÕES ===")
    
    print("🔍 PROBLEMAS IDENTIFICADOS:")
    print("   1. Inconsistência no tratamento de vírgulas entre scripts")
    print("   2. Diferentes precisões decimais (15 vs 28)")
    print("   3. Alguns scripts não tratam formato brasileiro")
    print("   4. Falta de padronização na conversão de valores")
    
    print("\n✅ RECOMENDAÇÕES:")
    print("   1. Padronizar função de conversão em todos os scripts")
    print("   2. Usar precisão Decimal de 28 em todos os scripts")
    print("   3. Implementar tratamento consistente para formato brasileiro")
    print("   4. Criar função única de conversão reutilizável")
    print("   5. Validar precisão máxima necessária para os cálculos")
    
    print("\n📋 FUNÇÃO RECOMENDADA:")
    print("""
def converter_para_decimal_padrao(valor):
    \"\"\"Função padronizada para converter valores para Decimal\"\"\"
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    return Decimal(str(valor))
    """)

def main():
    """Função principal"""
    print("ANÁLISE DE PRECISÃO E LEITURA DE VALORES")
    print("=" * 60)
    
    # 1. Analisar valores reais da planilha
    valores_analisados, max_casas = analisar_valores_planilha()
    
    # 2. Testar funções de conversão
    testar_funcoes_conversao()
    
    # 3. Testar precisão decimal
    testar_precisao_decimal()
    
    # 4. Analisar scripts existentes
    analisar_scripts_existentes()
    
    # 5. Gerar recomendações
    gerar_recomendacoes()
    
    # 6. Salvar relatório
    relatorio = {
        "max_casas_decimais": max_casas,
        "valores_analisados": valores_analisados,
        "recomendacoes": {
            "precisao_recomendada": 28,
            "funcao_padrao": "converter_para_decimal_padrao",
            "tratamento_virgula": "Remove pontos de milhares e substitui vírgula por ponto"
        }
    }
    
    with open("relatorio_precisao.json", "w", encoding="utf-8") as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n📄 Relatório salvo em 'relatorio_precisao.json'")

if __name__ == "__main__":
    main() 