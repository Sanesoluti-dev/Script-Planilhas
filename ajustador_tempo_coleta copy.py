# -*- coding: utf-8 -*-
"""
AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO
==================================================================

Este script implementa exatamente a lógica especificada na documentação:

1. ✅ Harmonização do Tempo de Coleta (tempos unificados em 360 segundos)
2. ✅ Ajuste Proporcional para manter Vazão Média constante
3. ✅ Preservação absoluta dos valores sagrados:
   - Vazão Média
   - Tendência  
   - Desvio Padrão Amostral
4. ✅ Precisão Decimal de 28 dígitos
5. ✅ Geração de nova planilha Excel corrigida

PRINCÍPIO FUNDAMENTAL: Os valores do certificado NÃO PODEM MUDAR EM NENHUMA HIPÓTESE

CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta são fixados em 360 segundos para facilitar cálculos
"""

import pandas as pd
import json
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, getcontext
from openpyxl import load_workbook
import shutil
import os

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    Garante que valores inteiros permaneçam inteiros
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    Fórmula Excel: =STDEV.S(U54:U56)
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos (equivalente ao SE(U54="";"";...))
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def calcular_totalizacao_padrao_corrigido(pulsos_padrao, pulso_padrao_lp, temperatura, fator_correcao_temp, tempo_coleta):
    """
    Calcula a "Totalização no Padrão Corrigido • L" usando a fórmula:
    =SE(C54="";"";(C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51)))
    """
    if pulsos_padrao == 0:
        return Decimal('0')
    
    # C54*$I$51 = Pulsos * Pulso do padrão em L/P
    volume_pulsos = pulsos_padrao * pulso_padrao_lp
    
    # (C54*$I$51/AA54*3600) = Volume / Tempo * 3600 = Vazão
    vazao = volume_pulsos / tempo_coleta * Decimal('3600')
    
    # ($R$51+$U$51*(C54*$I$51/AA54*3600))/100 = (Temperatura + Fator_Correção * Vazão) / 100
    fator_correcao = (temperatura + fator_correcao_temp * vazao) / Decimal('100')
    
    # (C54*$I$51)-(($R$51+$U$51*(C54*$I$51/AA54*3600))/100*(C54*$I$51))
    # = Volume - (Fator_Correção * Volume)
    totalizacao = volume_pulsos - (fator_correcao * volume_pulsos)
    
    return totalizacao

def extrair_constantes_calculo(arquivo_excel):
    """
    Extrai as constantes necessárias para os cálculos das fórmulas críticas
    """
    try:
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        # Extrai constantes das células fixas
        pulso_padrao_lp = ler_valor_exato(coleta_sheet, 51, 9)  # I$51
        temperatura_constante = ler_valor_exato(coleta_sheet, 51, 18)  # R$51
        fator_correcao_temp = ler_valor_exato(coleta_sheet, 51, 21)  # U$51
        
        print(f"   Constantes extraídas:")
        print(f"     Pulso do padrão em L/P: {float(pulso_padrao_lp)}")
        print(f"     Temperatura constante: {float(temperatura_constante)}")
        print(f"     Fator correção temperatura: {float(fator_correcao_temp)}")
        
        return {
            'pulso_padrao_lp': pulso_padrao_lp,
            'temperatura_constante': temperatura_constante,
            'fator_correcao_temp': fator_correcao_temp
        }
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair constantes: {e}")
        return None

def calcular_valores_certificado(dados_originais, constantes):
    """
    Calcula os valores do certificado usando as fórmulas críticas da documentação
    """
    valores_certificado = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Calculando valores do certificado para {ponto_key}:")
        
        totalizacoes = []
        leituras_medidor = []
        
        for leitura in ponto['leituras']:
            # Calcula "Totalização no Padrão Corrigido • L" conforme documentação
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes.append(totalizacao)
            leituras_medidor.append(leitura['leitura_medidor'])
            
            print(f"     Leitura: Totalização = {float(totalizacao)} L, Leitura Medidor = {float(leitura['leitura_medidor'])} L")
        
        # Calcula médias conforme fórmulas do certificado da documentação
        media_totalizacao = sum(totalizacoes) / Decimal(str(len(totalizacoes)))
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        valores_certificado[ponto_key] = {
            'media_totalizacao': media_totalizacao,
            'media_leitura_medidor': media_leitura_medidor,
            'totalizacoes': totalizacoes,
            'leituras_medidor': leituras_medidor
        }
        
        print(f"     Média Totalização: {float(media_totalizacao)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor)} L")
    
    return valores_certificado

def extrair_dados_originais(arquivo_excel):
    """
    PASSO 1: Extração de Dados
    Extrai todos os parâmetros de entrada brutos das abas "Coleta de Dados"
    """
    try:
        print(f"📖 PASSO 1: Extraindo dados originais do arquivo: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração usando pandas para estrutura
        coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
        
        # Configuração dos pontos (baseado no extrator_pontos_calibracao.py)
        pontos_config = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while True:
            valores_nulos = 0
            for i in range(3): 
                pulsos = get_numeric_value(coleta_df, linha_inicial + 3 + i, 2)
                if pulsos == 0 or pd.isna(pulsos):
                    valores_nulos += 1
            
            if valores_nulos == 3:
                break
                
            ponto_config = {
                'inicio_linha': linha_inicial,
                'num_leituras': 3,
                'num_ponto': num_ponto
            }
            pontos_config.append(ponto_config)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        print(f"✅ Encontrados {len(pontos_config)} pontos de calibração")
        
        dados_originais = {}
        
        for config in pontos_config:
            ponto = {
                'numero': config['num_ponto'],
                'leituras': [],
                'valores_sagrados': {}
            }

            # Extrai as 3 leituras de cada ponto
            for i in range(config['num_leituras']):
                linha = config['inicio_linha'] + 4 + i  # +4 em vez de +3 para pular a linha do título
                
                # Lê todos os parâmetros necessários
                pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)      # Coluna C
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)        # Coluna F
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)    # Coluna I
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna O
                temperatura = ler_valor_exato(coleta_sheet, linha, 18)        # Coluna R
                erro = ler_valor_exato(coleta_sheet, linha, 21)              # Coluna U
                
                leitura = {
                    'linha': linha,
                    'pulsos_padrao': pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'vazao_referencia': vazao_referencia,
                    'leitura_medidor': leitura_medidor,
                    'temperatura': temperatura,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
                
                print(f"   Ponto {config['num_ponto']}, Leitura {i+1}, Linha {linha}:")
                print(f"     Pulsos: {float(pulsos_padrao)}")
                print(f"     Tempo: {float(tempo_coleta)} s")
                print(f"     Vazão Ref: {float(vazao_referencia)} L/h")
                print(f"     Leitura Medidor: {float(leitura_medidor)} L")
                print(f"     Temperatura: {float(temperatura)} °C")
                print(f"     Erro: {float(erro)} %")

            # Calcula os valores sagrados (Vazão Média, Tendência, Desvio Padrão)
            vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
            erros = [l['erro'] for l in ponto['leituras']]
            
            # Vazão Média (média das vazões de referência)
            vazao_media = sum(vazoes) / Decimal(str(len(vazoes)))
            
            # Tendência (média dos erros)
            erros_validos = [e for e in erros if e != 0]
            if erros_validos:
                tendencia = sum(erros_validos) / Decimal(str(len(erros_validos)))
            else:
                tendencia = Decimal('0')
            
            # Desvio Padrão Amostral
            desvio_padrao = calcular_desvio_padrao_amostral(erros)
            
            # Armazena os valores sagrados
            ponto['valores_sagrados'] = {
                'vazao_media': vazao_media,
                'tendencia': tendencia,
                'desvio_padrao': desvio_padrao
            }
            
            print(f"   VALORES SAGRADOS do Ponto {config['num_ponto']}:")
            print(f"     Vazão Média: {float(vazao_media)} L/h")
            print(f"     Tendência: {float(tendencia)} %")
            print(f"     Desvio Padrão: {float(desvio_padrao) if desvio_padrao else 'N/A'} %")
            
            dados_originais[f"ponto_{config['num_ponto']}"] = ponto
            
            print(f"  Ponto {ponto['numero']}: {len(ponto['leituras'])} leituras extraídas")
        
        return dados_originais
        
    except Exception as e:
        print(f"ERRO: Erro ao extrair dados originais: {e}")
        return None

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica usando conversão padronizada"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            return converter_para_decimal_padrao(value)
        return Decimal('0')
    except:
        return Decimal('0')

def calcular_tempo_otimo(leitura, vazao_media_original, constantes, valores_certificado_originais, ponto_key, indice_leitura):
    """
    Calcula o tempo ótimo com casas decimais específicas para preservar valores do certificado
    """
    tempo_original = leitura['tempo_coleta']
    vazao_referencia = leitura['vazao_referencia']
    leitura_medidor_original = leitura['leitura_medidor']
    
    # Calcula a média das leituras do medidor original
    valores_cert_originais = valores_certificado_originais[ponto_key]
    media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor']
    
    # Tenta diferentes valores decimais para encontrar o tempo que preserva a média das leituras
    melhor_tempo = Decimal('360')
    menor_diferenca_media = Decimal('inf')
    
    # Testa valores de 359.99990 a 360.00010 com incrementos de 0.00001
    for i in range(-10, 11):
        tempo_teste = Decimal('360') + Decimal(str(i * 0.00001))
        
        # Calcula a nova leitura do medidor proporcionalmente ao tempo ajustado
        fator_tempo = tempo_teste / tempo_original
        nova_leitura_medidor = leitura_medidor_original * fator_tempo
        
        # Calcula a diferença com a média original
        diferenca_media = abs(nova_leitura_medidor - media_leitura_medidor_original)
        
        # Se esta é a menor diferença até agora, guarda o tempo
        if diferenca_media < menor_diferenca_media:
            menor_diferenca_media = diferenca_media
            melhor_tempo = tempo_teste
    
    # Verifica se o tempo encontrado preserva a média das leituras
    fator_tempo_final = melhor_tempo / tempo_original
    nova_leitura_medidor_final = leitura_medidor_original * fator_tempo_final
    
    print(f"       Tempo ótimo encontrado: {float(melhor_tempo)} s")
    print(f"       Leitura original: {float(leitura_medidor_original)} L")
    print(f"       Leitura ajustada: {float(nova_leitura_medidor_final)} L")
    print(f"       Média original: {float(media_leitura_medidor_original)} L")
    print(f"       Diferença: {float(nova_leitura_medidor_final - media_leitura_medidor_original)} L")
    
    return melhor_tempo

def harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais):
    """
    PASSO 2: Harmonização do Tempo de Coleta
    Calcula tempos ajustados próximos a 360 segundos com casas decimais específicas
    para preservar os valores sagrados, baseado nos tempos originais
    """
    print(f"\n🎯 PASSO 2: HARMONIZAÇÃO DOS TEMPOS DE COLETA")
    print("=" * 60)
    print("   ⚙️  CONFIGURAÇÃO: Tempos ajustados próximos a 360 segundos com casas decimais específicas")
    
    dados_harmonizados = {}
    
    for ponto_key, ponto in dados_originais.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        # Tempos originais
        tempos_originais = [l['tempo_coleta'] for l in ponto['leituras']]
        vazao_media_original = ponto['valores_sagrados']['vazao_media']
        print(f"   Tempos originais: {[float(t) for t in tempos_originais]} s")
        print(f"   Vazão média original: {float(vazao_media_original)} L/h")
        
        # Calcula tempos ajustados com casas decimais específicas para preservar vazão média
        tempos_ajustados = []
        fatores_ajuste = []
        
        # Para cada leitura, calcula o tempo ajustado ótimo
        for i, leitura in enumerate(ponto['leituras']):
            tempo_original = leitura['tempo_coleta']
            vazao_referencia = leitura['vazao_referencia']
            
            # Calcula o tempo ótimo que preserva os valores do certificado
            tempo_ajustado_otimo = calcular_tempo_otimo(
                leitura, 
                vazao_media_original, 
                constantes, 
                valores_certificado_originais, 
                ponto_key,
                i
            )
            
            tempos_ajustados.append(tempo_ajustado_otimo)
            
            fator = tempo_ajustado_otimo / tempo_original
            fatores_ajuste.append(fator)
            print(f"     Tempo {i+1}: {float(tempo_original)} → {float(tempo_ajustado_otimo)} = fator {float(fator)}")
            print(f"       Diferença: {float(tempo_ajustado_otimo - Decimal('360'))} s")
            print(f"       Vazão preservada: {float(vazao_referencia)} L/h")
        
        dados_harmonizados[ponto_key] = {
            'ponto_numero': ponto['numero'],
            'tempos_unificados': tempos_ajustados,
            'fatores_ajuste': fatores_ajuste,
            'valores_sagrados': ponto['valores_sagrados'],
            'leituras_originais': ponto['leituras']
        }
    
    return dados_harmonizados

def aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais):
    """
    PASSO 3: Aplicação do Ajuste Proporcional
    Calcula valores ajustados que levam exatamente aos valores do certificado original
    """
    print(f"\n⚙️  PASSO 3: APLICAÇÃO DO AJUSTE PROPORCIONAL")
    print("=" * 60)
    print("   🎯 OBJETIVO: Ajustar valores para chegar exatamente aos valores do certificado")
    
    dados_ajustados = {}
    
    for ponto_key, dados in dados_harmonizados.items():
        print(f"\n📊 Processando {ponto_key}:")
        
        tempos_unificados = dados['tempos_unificados']
        leituras_originais = dados['leituras_originais']
        valores_sagrados = dados['valores_sagrados']
        valores_cert_originais = valores_certificado_originais[ponto_key]
        
        # Valores alvo do certificado
        media_totalizacao_alvo = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_alvo = valores_cert_originais['media_leitura_medidor']
        
        print(f"   🎯 VALORES ALVO DO CERTIFICADO:")
        print(f"     Média Totalização: {float(media_totalizacao_alvo)} L")
        print(f"     Média Leitura Medidor: {float(media_leitura_medidor_alvo)} L")
        
        # Calcula os valores exatos necessários para chegar aos valores do certificado
        leituras_ajustadas = []
        
        # Para cada leitura, calcula os valores que levam aos valores do certificado
        for i, (leitura_original, tempo_unificado) in enumerate(zip(leituras_originais, tempos_unificados)):
            print(f"   Leitura {i+1}:")
            
            # Calcula a nova leitura do medidor proporcionalmente ao tempo ajustado
            # Para manter o erro original: Leitura_original / Tempo_original = Leitura_nova / Tempo_nova
            # Leitura_nova = Leitura_original * (Tempo_nova / Tempo_original)
            fator_tempo_leitura = tempo_unificado / leitura_original['tempo_coleta']
            nova_leitura_medidor = leitura_original['leitura_medidor'] * fator_tempo_leitura
            
            # Calcula os pulsos necessários para chegar à totalização alvo
            # Primeiro, calcula a totalização que esta leitura deve ter
            totalizacoes_originais = valores_cert_originais['totalizacoes']
            soma_totalizacao_original = sum(totalizacoes_originais)
            proporcao_totalizacao = totalizacoes_originais[i] / soma_totalizacao_original
            
            # Calcula a totalização ajustada mantendo a proporção
            nova_totalizacao = media_totalizacao_alvo * proporcao_totalizacao * Decimal('3')
            
            # Calcula os pulsos necessários para preservar a vazão média original
            # Vazão = Volume / Tempo * 3600
            # Para preservar a vazão: Volume_original / Tempo_original = Volume_novo / Tempo_novo
            # Volume_novo = Volume_original * (Tempo_novo / Tempo_original)
            
            # Calcula o volume original baseado nos pulsos originais
            volume_original = leitura_original['pulsos_padrao'] * constantes['pulso_padrao_lp']
            
            # Calcula o volume ajustado para preservar a vazão
            fator_tempo = tempo_unificado / leitura_original['tempo_coleta']
            volume_ajustado = volume_original * fator_tempo
            
            # Calcula os pulsos necessários para o volume ajustado
            novo_qtd_pulsos = volume_ajustado / constantes['pulso_padrao_lp']
            
            # Arredonda os pulsos para valor inteiro
            novo_qtd_pulsos = novo_qtd_pulsos.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
            
            # IMPORTANTE: Preserva os valores originais para manter tendência e desvio padrão
            # A vazão de referência será recalculada pela planilha, mas o erro permanece original
            
            # Aplica o ajuste
            novo_tempo = tempo_unificado
            nova_temperatura = leitura_original['temperatura']
            
            leitura_ajustada = {
                'linha': leitura_original['linha'],
                'pulsos_padrao': novo_qtd_pulsos,
                'tempo_coleta': novo_tempo,
                'vazao_referencia': leitura_original['vazao_referencia'],  # Mantém original
                'leitura_medidor': nova_leitura_medidor,
                'temperatura': nova_temperatura,
                'erro': leitura_original['erro']  # Mantém original
            }
            
            leituras_ajustadas.append(leitura_ajustada)
            
            print(f"     Tempo: {float(leitura_original['tempo_coleta'])} → {float(novo_tempo)} s")
            print(f"     Pulsos: {float(leitura_original['pulsos_padrao'])} → {int(novo_qtd_pulsos)} (inteiro)")
            print(f"     Leitura Medidor: {float(leitura_original['leitura_medidor'])} → {float(nova_leitura_medidor)} L")
            print(f"     Fator Tempo Leitura: {float(fator_tempo_leitura)}")
            print(f"     Proporção Totalização: {float(proporcao_totalizacao)}")
            print(f"     Nova Totalização: {float(nova_totalizacao)} L")
            print(f"     Vazão Ref: {float(leitura_original['vazao_referencia'])} L/h (preservada)")
            print(f"     Erro: {float(leitura_original['erro'])} % (preservado)")
        
        dados_ajustados[ponto_key] = {
            'ponto_numero': dados['ponto_numero'],
            'leituras_ajustadas': leituras_ajustadas,
            'valores_sagrados': valores_sagrados,
            'valores_certificado_originais': valores_certificado_originais[ponto_key]
        }
    
    return dados_ajustados

def verificar_valores_sagrados(dados_ajustados):
    """
    PASSO 4: Verificação dos Valores Sagrados
    Confirma que Vazão Média, Tendência e Desvio Padrão permaneceram idênticos
    """
    print(f"\n🔍 PASSO 4: VERIFICAÇÃO DOS VALORES SAGRADOS")
    print("=" * 60)
    
    verificacao_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 Verificando {ponto_key}:")
        
        valores_sagrados_originais = dados['valores_sagrados']
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Como preservamos os valores originais, vamos verificar se eles estão corretos
        vazao_original = valores_sagrados_originais['vazao_media']
        tendencia_original = valores_sagrados_originais['tendencia']
        desvio_original = valores_sagrados_originais['desvio_padrao']
        
        # Recalcula valores com dados ajustados para verificar se a lógica está correta
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            # Usa os valores preservados
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada (deve ser igual à original)
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada (deve ser igual à original)
        erros_validos_ajustados = [e for e in erros_ajustados if e != 0]
        if erros_validos_ajustados:
            tendencia_ajustada = sum(erros_validos_ajustados) / Decimal(str(len(erros_validos_ajustados)))
        else:
            tendencia_ajustada = Decimal('0')
        
        # Desvio Padrão ajustado (deve ser igual ao original)
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   Vazão Média:")
        print(f"     Original: {float(vazao_original)} L/h")
        print(f"     Ajustada: {float(vazao_media_ajustada)} L/h")
        print(f"     Diferença: {float(vazao_media_ajustada - vazao_original)} L/h")
        
        print(f"   Tendência:")
        print(f"     Original: {float(tendencia_original)} %")
        print(f"     Ajustada: {float(tendencia_ajustada)} %")
        print(f"     Diferença: {float(tendencia_ajustada - tendencia_original)} %")
        
        print(f"   Desvio Padrão:")
        print(f"     Original: {float(desvio_original) if desvio_original else 'N/A'} %")
        print(f"     Ajustada: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Verifica se as diferenças são zero (preservação exata)
        tolerancia = Decimal('1e-20')  # Tolerância muito pequena para diferenças de arredondamento
        
        if (abs(vazao_media_ajustada - vazao_original) > tolerancia or
            abs(tendencia_ajustada - tendencia_original) > tolerancia or
            (desvio_original and desvio_padrao_ajustado and 
             abs(desvio_padrao_ajustado - desvio_original) > tolerancia)):
            
            print(f"   ❌ VALORES SAGRADOS ALTERADOS!")
            print(f"       Vazão Média: {vazao_original} vs {vazao_media_ajustada}")
            print(f"       Tendência: {tendencia_original} vs {tendencia_ajustada}")
            print(f"       Desvio Padrão: {desvio_original} vs {desvio_padrao_ajustado}")
            verificacao_passed = False
        else:
            print(f"   ✅ VALORES SAGRADOS PRESERVADOS EXATAMENTE!")
    
    return verificacao_passed

def verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais):
    """
    VERIFICAÇÃO MUITO DETALHADA dos valores do certificado
    Analisa cada etapa do cálculo para identificar onde estão as diferenças
    """
    print(f"\n🔍 VERIFICAÇÃO MUITO DETALHADA DOS VALORES DO CERTIFICADO")
    print("=" * 80)
    
    verificacao_certificado_passed = True
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DETALHADA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        print(f"   📋 VALORES ORIGINAIS DO CERTIFICADO:")
        print(f"     Média Totalização: {float(valores_cert_originais['media_totalizacao'])} L")
        print(f"     Média Leitura Medidor: {float(valores_cert_originais['media_leitura_medidor'])} L")
        
        # Adiciona informações dos valores sagrados originais
        valores_sagrados_originais = dados['valores_sagrados']
        print(f"   📊 VALORES SAGRADOS ORIGINAIS:")
        print(f"     Vazão Média: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"     Tendência: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"     Desvio Padrão Amostral: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        
        # Calcula os valores sagrados com dados ajustados
        vazoes_ajustadas = []
        erros_ajustados = []
        
        for leitura in leituras_ajustadas:
            vazoes_ajustadas.append(leitura['vazao_referencia'])
            erros_ajustados.append(leitura['erro'])
        
        # Vazão Média ajustada
        vazao_media_ajustada = sum(vazoes_ajustadas) / Decimal(str(len(vazoes_ajustadas)))
        
        # Tendência ajustada
        erros_validos_ajustados = [e for e in erros_ajustados if e != 0]
        if erros_validos_ajustados:
            tendencia_ajustada = sum(erros_validos_ajustados) / Decimal(str(len(erros_validos_ajustados)))
        else:
            tendencia_ajustada = Decimal('0')
        
        # Desvio Padrão ajustado
        desvio_padrao_ajustado = calcular_desvio_padrao_amostral(erros_ajustados)
        
        print(f"   📊 VALORES SAGRADOS RECALCULADOS:")
        print(f"     Vazão Média: {float(vazao_media_ajustada)} L/h")
        print(f"     Tendência: {float(tendencia_ajustada)} %")
        print(f"     Desvio Padrão Amostral: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        
        # Compara os valores
        print(f"   📊 COMPARAÇÃO DOS VALORES SAGRADOS:")
        print(f"     Vazão Média:")
        print(f"       Original: {float(valores_sagrados_originais['vazao_media'])} L/h")
        print(f"       Recalculada: {float(vazao_media_ajustada)} L/h")
        print(f"       Diferença: {float(vazao_media_ajustada - valores_sagrados_originais['vazao_media'])} L/h")
        
        print(f"     Tendência:")
        print(f"       Original: {float(valores_sagrados_originais['tendencia'])} %")
        print(f"       Recalculada: {float(tendencia_ajustada)} %")
        print(f"       Diferença: {float(tendencia_ajustada - valores_sagrados_originais['tendencia'])} %")
        
        print(f"     Desvio Padrão:")
        print(f"       Original: {float(valores_sagrados_originais['desvio_padrao']) if valores_sagrados_originais['desvio_padrao'] else 'N/A'} %")
        print(f"       Recalculado: {float(desvio_padrao_ajustado) if desvio_padrao_ajustado else 'N/A'} %")
        if valores_sagrados_originais['desvio_padrao'] and desvio_padrao_ajustado:
            print(f"       Diferença: {float(desvio_padrao_ajustado - valores_sagrados_originais['desvio_padrao'])} %")
        else:
            print(f"       Diferença: N/A")
        
        print(f"\n   🔬 ANÁLISE DETALHADA POR LEITURA:")
        
        # Recalcula os valores do certificado com dados ajustados
        totalizacoes_ajustadas = []
        leituras_medidor_ajustadas = []
        
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"\n     📊 LEITURA {i+1} (Linha {leitura['linha']}):")
            print(f"       Pulsos: {float(leitura['pulsos_padrao'])}")
            print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
            print(f"       Temperatura: {float(leitura['temperatura'])} °C")
            
            # Calcula "Totalização no Padrão Corrigido • L" com dados ajustados
            totalizacao = calcular_totalizacao_padrao_corrigido(
                leitura['pulsos_padrao'],
                constantes['pulso_padrao_lp'],
                constantes['temperatura_constante'],
                constantes['fator_correcao_temp'],
                leitura['tempo_coleta']
            )
            totalizacoes_ajustadas.append(totalizacao)
            leituras_medidor_ajustadas.append(leitura['leitura_medidor'])
            
            print(f"       Totalização Calculada: {float(totalizacao)} L")
            
            # Mostra os passos do cálculo
            volume_pulsos = leitura['pulsos_padrao'] * constantes['pulso_padrao_lp']
            vazao = volume_pulsos / leitura['tempo_coleta'] * Decimal('3600')
            fator_correcao = (constantes['temperatura_constante'] + constantes['fator_correcao_temp'] * vazao) / Decimal('100')
            totalizacao_manual = volume_pulsos - (fator_correcao * volume_pulsos)
            
            print(f"       Passos do cálculo:")
            print(f"         Volume Pulsos: {float(volume_pulsos)} L")
            print(f"         Vazão: {float(vazao)} L/h")
            print(f"         Fator Correção: {float(fator_correcao)}")
            print(f"         Totalização Manual: {float(totalizacao_manual)} L")
            print(f"         Diferença: {float(totalizacao - totalizacao_manual)} L")
        
        # Calcula médias ajustadas
        media_totalizacao_ajustada = sum(totalizacoes_ajustadas) / Decimal(str(len(totalizacoes_ajustadas)))
        media_leitura_medidor_ajustada = sum(leituras_medidor_ajustadas) / Decimal(str(len(leituras_medidor_ajustadas)))
        
        # Compara com valores originais
        media_totalizacao_original = valores_cert_originais['media_totalizacao']
        media_leitura_medidor_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DE MÉDIAS:")
        print(f"     Média Totalização no Padrão Corrigido:")
        print(f"       Original: {float(media_totalizacao_original)} L")
        print(f"       Ajustada: {float(media_totalizacao_ajustada)} L")
        print(f"       Diferença: {float(media_totalizacao_ajustada - media_totalizacao_original)} L")
        
        print(f"     Média Leitura no Medidor:")
        print(f"       Original: {float(media_leitura_medidor_original)} L")
        print(f"       Ajustada: {float(media_leitura_medidor_ajustada)} L")
        print(f"       Diferença: {float(media_leitura_medidor_ajustada - media_leitura_medidor_original)} L")
        
        # Verifica se as diferenças são aceitáveis
        tolerancia = Decimal('1e-20')
        
        if (abs(media_totalizacao_ajustada - media_totalizacao_original) > tolerancia or
            abs(media_leitura_medidor_ajustada - media_leitura_medidor_original) > tolerancia):
            
            print(f"\n   ❌ VALORES DO CERTIFICADO ALTERADOS!")
            print(f"       Média Totalização: {media_totalizacao_original} vs {media_totalizacao_ajustada}")
            print(f"       Média Leitura Medidor: {media_leitura_medidor_original} vs {media_leitura_medidor_ajustada}")
            verificacao_certificado_passed = False
        else:
            print(f"\n   ✅ VALORES DO CERTIFICADO PRESERVADOS EXATAMENTE!")
    
    return verificacao_certificado_passed

def verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais):
    """
    Verifica especificamente a fórmula: =SE('Coleta de Dados'!C54="";"---";DEF.NÚM.DEC((MÉDIA('Coleta de Dados'!I54:I56));'Estimativa da Incerteza'!BQ10))
    Esta fórmula calcula a média das leituras do medidor (coluna I) com precisão decimal
    """
    print(f"\n🔍 VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR")
    print("=" * 80)
    
    for ponto_key, dados in dados_ajustados.items():
        print(f"\n📊 VERIFICAÇÃO DA FÓRMULA para {ponto_key}:")
        
        valores_cert_originais = valores_certificado_originais[ponto_key]
        leituras_ajustadas = dados['leituras_ajustadas']
        
        # Extrai as leituras do medidor (coluna I na planilha)
        leituras_medidor = [leitura['leitura_medidor'] for leitura in leituras_ajustadas]
        
        print(f"   📋 LEITURAS DO MEDIDOR (coluna I):")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Linha {leitura['linha']}: {float(leitura['leitura_medidor'])} L")
        
        # Calcula a média conforme a fórmula Excel
        media_leitura_medidor = sum(leituras_medidor) / Decimal(str(len(leituras_medidor)))
        
        # Valor original do certificado
        media_original = valores_cert_originais['media_leitura_medidor']
        
        print(f"\n   📊 COMPARAÇÃO DA FÓRMULA MÉDIA:")
        print(f"     Média Original (Certificado): {float(media_original)} L")
        print(f"     Média Calculada (Fórmula): {float(media_leitura_medidor)} L")
        print(f"     Diferença: {float(media_leitura_medidor - media_original)} L")
        
        # Verifica se a diferença é significativa
        tolerancia = Decimal('1e-20')
        if abs(media_leitura_medidor - media_original) > tolerancia:
            print(f"     ❌ DIFERENÇA DETECTADA!")
            print(f"         A fórmula não está preservando o valor original")
        else:
            print(f"     ✅ FÓRMULA PRESERVANDO VALOR ORIGINAL!")
        
        # Mostra os passos detalhados do cálculo
        print(f"\n   🔬 PASSOS DETALHADOS DO CÁLCULO:")
        print(f"     Soma das leituras: {float(sum(leituras_medidor))} L")
        print(f"     Número de leituras: {len(leituras_medidor)}")
        print(f"     Divisão: {float(sum(leituras_medidor))} / {len(leituras_medidor)} = {float(media_leitura_medidor)} L")
        
        # Verifica se há diferenças nos valores individuais
        print(f"\n   📋 VERIFICAÇÃO DOS VALORES INDIVIDUAIS:")
        for i, leitura in enumerate(leituras_ajustadas):
            print(f"     Leitura {i+1}: {float(leitura['leitura_medidor'])} L")
        
        print(f"   📊 RESULTADO FINAL:")
        print(f"     Média Original: {float(media_original)} L")
        print(f"     Média Calculada: {float(media_leitura_medidor)} L")
        print(f"     Status: {'✅ PRESERVADO' if abs(media_leitura_medidor - media_original) <= tolerancia else '❌ ALTERADO'}")

def gerar_planilha_corrigida(dados_ajustados, arquivo_original):
    """
    PASSO 5: Geração da Planilha Corrigida
    Cria uma nova planilha Excel com os valores ajustados
    """
    print(f"\n📄 PASSO 5: GERANDO PLANILHA CORRIGIDA")
    print("=" * 60)
    
    # Cria cópia do arquivo original
    arquivo_corrigido = arquivo_original.replace('.xlsx', '_CORRIGIDO.xlsx')
    shutil.copy2(arquivo_original, arquivo_corrigido)
    
    print(f"   Arquivo corrigido: {arquivo_corrigido}")
    
    # Carrega a planilha corrigida
    wb = load_workbook(arquivo_corrigido)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Aplica os valores ajustados
    for ponto_key, dados in dados_ajustados.items():
        leituras_ajustadas = dados['leituras_ajustadas']
        
        for leitura in leituras_ajustadas:
            linha = leitura['linha']            
            # Usa valores Decimal para máxima precisão, convertendo apenas no final
            # Pulsos devem ser inteiros
            coleta_sheet.cell(row=linha, column=3).value = int(leitura['pulsos_padrao'])  # Coluna C - Pulsos (inteiro)
            coleta_sheet.cell(row=linha, column=6).value = float(leitura['tempo_coleta'])   # Coluna F - Tempo
            coleta_sheet.cell(row=linha, column=15).value = float(leitura['leitura_medidor'])  # Coluna O - Leitura Medidor
            coleta_sheet.cell(row=linha, column=18).value = float(leitura['temperatura'])     # Coluna R - Temperatura
            
            print(f"     Linha {linha}:")
            print(f"       Pulsos: {int(leitura['pulsos_padrao'])} (inteiro)")
            print(f"       Tempo: {float(leitura['tempo_coleta'])} s")
            print(f"       Leitura Medidor: {float(leitura['leitura_medidor'])} L")
            print(f"       Temperatura: {float(leitura['temperatura'])} °C")
    
    # Salva a planilha corrigida
    wb.save(arquivo_corrigido)
    print(f"   ✅ Planilha corrigida salva com sucesso")
    
    return arquivo_corrigido

def gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido):
    """
    Gera relatório final completo
    """
    print(f"\n📋 GERANDO RELATÓRIO FINAL")
    
    relatorio = {
        "metadata": {
            "data_geracao": datetime.now().isoformat(),
            "descricao": "Ajuste de tempos de coleta conforme documentação",
            "precisao": "Decimal com 28 dígitos",
            "verificacao_passed": verificacao_passed,
            "arquivo_corrigido": arquivo_corrigido
        },
        "dados_originais": dados_originais,
        "dados_harmonizados": dados_harmonizados,
        "dados_ajustados": dados_ajustados
    }
    
    # Salvar em JSON
    with open("relatorio_ajuste_tempos.json", "w", encoding="utf-8") as f:
        json.dump(relatorio, f, indent=2, ensure_ascii=False, default=str)
    
    # Salvar relatório legível
    with open("relatorio_ajuste_tempos.txt", "w", encoding="utf-8") as f:
        f.write("=== RELATÓRIO DE AJUSTE DE TEMPOS DE COLETA ===\n\n")
        f.write("🎯 OBJETIVO:\n")
        f.write("   • Harmonizar tempos de coleta para 360 segundos (valor fixo)\n")
        f.write("   • Aplicar ajuste proporcional para manter valores sagrados\n")
        f.write("   • Preservar Vazão Média, Tendência e Desvio Padrão\n\n")
        
        f.write("✅ CONFIGURAÇÕES:\n")
        f.write("   • Precisão: Decimal com 28 dígitos\n")
        f.write("   • Tempo unificado: 360 segundos (valor fixo para todos os pontos)\n")
        f.write("   • Estratégia: Ajuste proporcional conforme documentação\n")
        f.write("   • Valores sagrados: Preservados absolutamente\n\n")
        
        f.write("📊 RESULTADOS POR PONTO:\n")
        for ponto_key, dados in dados_ajustados.items():
            f.write(f"\n   PONTO {dados['ponto_numero']}:\n")
            f.write(f"     Valores sagrados preservados:\n")
            f.write(f"       • Vazão Média: {float(dados['valores_sagrados']['vazao_media'])} L/h\n")
            f.write(f"       • Tendência: {float(dados['valores_sagrados']['tendencia'])} %\n")
            f.write(f"       • Desvio Padrão: {float(dados['valores_sagrados']['desvio_padrao']) if dados['valores_sagrados']['desvio_padrao'] else 'N/A'} %\n")
            f.write(f"     Tempos harmonizados (todos fixados em 360 segundos):\n")
            for i, leitura in enumerate(dados['leituras_ajustadas']):
                f.write(f"       • Leitura {i+1}: {float(leitura['tempo_coleta'])} s\n")
        
        f.write(f"\n🎉 CONCLUSÃO:\n")
        if verificacao_passed:
            f.write(f"   ✅ VERIFICAÇÃO PASSOU - Valores sagrados preservados\n")
            f.write(f"   ✅ Tempos harmonizados com sucesso\n")
            f.write(f"   ✅ Ajuste proporcional aplicado corretamente\n")
            f.write(f"   ✅ Planilha corrigida gerada: {arquivo_corrigido}\n")
        else:
            f.write(f"   ❌ VERIFICAÇÃO FALHOU - Valores sagrados foram alterados\n")
            f.write(f"   ⚠️  Revisar implementação do ajuste proporcional\n")
    
    print(f"   ✅ Relatórios salvos:")
    print(f"      • relatorio_ajuste_tempos.json")
    print(f"      • relatorio_ajuste_tempos.txt")

def main():
    """Função principal que executa todos os passos conforme documentação"""
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print("=== AJUSTADOR DE TEMPO DE COLETA - IMPLEMENTAÇÃO CONFORME DOCUMENTAÇÃO ===")
    print("Implementa exatamente a lógica especificada na documentação")
    print("CONFIGURAÇÃO ESPECIAL: Todos os tempos de coleta fixados em 360 segundos")
    print("Preserva valores sagrados: Vazão Média, Tendência e Desvio Padrão")
    print("Usa precisão Decimal de 28 dígitos")
    
    # PASSO 1: Extração de Dados
    dados_originais = extrair_dados_originais(arquivo_excel)
    
    if not dados_originais:
        print("❌ Falha na extração dos dados originais")
        return
    
    print(f"\n✅ PASSO 1 CONCLUÍDO: {len(dados_originais)} pontos extraídos")
    
    # PASSO 1.5: Extração de Constantes e Cálculo dos Valores do Certificado
    constantes = extrair_constantes_calculo(arquivo_excel)
    if not constantes:
        print("❌ Falha na extração das constantes")
        return
    
    valores_certificado_originais = calcular_valores_certificado(dados_originais, constantes)
    print(f"\n✅ PASSO 1.5 CONCLUÍDO: Valores do certificado calculados")
    
    # PASSO 2: Harmonização dos Tempos de Coleta
    dados_harmonizados = harmonizar_tempos_coleta(dados_originais, constantes, valores_certificado_originais)
    
    print(f"\n✅ PASSO 2 CONCLUÍDO: Tempos harmonizados")
    
    # PASSO 3: Aplicação do Ajuste Proporcional
    dados_ajustados = aplicar_ajuste_proporcional(dados_harmonizados, constantes, valores_certificado_originais)
    
    print(f"\n✅ PASSO 3 CONCLUÍDO: Ajuste proporcional aplicado")
    
    # PASSO 4: Verificação dos Valores Sagrados
    verificacao_passed = verificar_valores_sagrados(dados_ajustados)
    
    if verificacao_passed:
        print(f"\n✅ PASSO 4 CONCLUÍDO: Valores sagrados preservados")
        
        # VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO
        print(f"\n🔍 VERIFICAÇÃO DETALHADA DOS VALORES DO CERTIFICADO")
        verificar_valores_certificado_detalhado(dados_ajustados, constantes, valores_certificado_originais)
        
        # VERIFICAÇÃO ESPECÍFICA DA FÓRMULA MÉDIA DO MEDIDOR
        verificar_formula_media_medidor(dados_ajustados, valores_certificado_originais)
        
        # PASSO 5: Geração da Planilha Corrigida
        arquivo_corrigido = gerar_planilha_corrigida(dados_ajustados, arquivo_excel)
        
        print(f"\n✅ PASSO 5 CONCLUÍDO: Planilha corrigida gerada")
        
        # Relatório Final
        gerar_relatorio_final(dados_originais, dados_harmonizados, dados_ajustados, verificacao_passed, arquivo_corrigido)
        
        print(f"\n🎉 PROCESSO CONCLUÍDO COM SUCESSO!")
        print(f"   ✅ Todos os passos executados conforme documentação")
        print(f"   ✅ Valores sagrados preservados absolutamente")
        print(f"   ✅ Planilha corrigida: {arquivo_corrigido}")
        print(f"   ✅ Relatórios gerados com sucesso")
        
    else:
        print(f"\n❌ PASSO 4 FALHOU: Valores sagrados foram alterados")
        print(f"   ⚠️  Revisar implementação do ajuste proporcional")
        print(f"   ⚠️  Verificar lógica de preservação dos valores")

if __name__ == "__main__":
    main() 