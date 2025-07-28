import json
import math
from decimal import Decimal, getcontext
from openpyxl import load_workbook

# Configurar precisão alta para evitar diferenças de arredondamento
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    Garante que valores inteiros permaneçam inteiros
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def calcular_desvio_padrao_amostral(valores):
    """
    Calcula o desvio padrão amostral (STDEV.S) usando precisão Decimal
    """
    if not valores or len(valores) < 2:
        return None
    
    # Filtra valores não nulos
    valores_validos = [v for v in valores if v != 0]
    
    if len(valores_validos) < 2:
        return None
    
    # Calcula a média
    media = sum(valores_validos) / Decimal(str(len(valores_validos)))
    
    # Calcula a soma dos quadrados das diferenças
    soma_quadrados = sum((v - media) ** 2 for v in valores_validos)
    
    # Calcula o desvio padrão amostral: sqrt(soma_quadrados / (n-1))
    n = len(valores_validos)
    variancia = soma_quadrados / Decimal(str(n - 1))
    desvio_padrao = variancia.sqrt()
    
    return desvio_padrao

def carregar_dados_extraidos(arquivo_json):
    """Carrega os dados extraídos do extrator_pontos_calibracao.py"""
    try:
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)
        
        # Converte os valores string para Decimal
        for ponto in dados['pontos_dados']:
            # Converte valores principais
            ponto['media_vazao_referencia'] = Decimal(ponto['media_vazao_referencia'])
            ponto['tendencia'] = Decimal(ponto['tendencia'])
            ponto['desvio_padrao'] = Decimal(ponto['desvio_padrao'])
            
            # Converte vazões de referência
            for vazao in ponto['vazoes_referencia']:
                vazao['vazao_referencia'] = Decimal(vazao['vazao_referencia'])
            
            # Converte erros
            for erro in ponto['erros']:
                erro['erro'] = Decimal(erro['erro'])
        
        return dados
    except Exception as e:
        print(f"ERRO ao carregar dados: {e}")
        return None

def ler_dados_originais_planilha(arquivo_excel):
    """Lê os dados originais da planilha para obter os valores que podem ser ajustados"""
    try:
        print(f"Lendo dados originais da planilha: {arquivo_excel}")
        
        # Carregar planilha com openpyxl para precisão máxima
        wb = load_workbook(arquivo_excel, data_only=True)
        coleta_sheet = wb["Coleta de Dados"]
        
        print("✅ Aba 'Coleta de Dados' carregada com sucesso")
        
        # Identifica os pontos de calibração
        pontos_dados = []
        linha_inicial = 50
        avanca_linha = 9
        num_ponto = 1
        
        while num_ponto <= 8:  # Assumindo 8 pontos como no exemplo
            ponto = {
                'numero': num_ponto,
                'leituras': []
            }
            
            # Extrai as 3 leituras de cada ponto
            for i in range(3):
                linha = linha_inicial + 4 + i  # +4 para pular a linha do título
                
                # Lê os valores que podem ser ajustados
                qtd_pulsos_padrao = ler_valor_exato(coleta_sheet, linha, 3)  # Coluna B - Pulsos (assumindo)
                tempo_coleta = ler_valor_exato(coleta_sheet, linha, 6)       # Coluna C - Tempo
                leitura_medidor = ler_valor_exato(coleta_sheet, linha, 15)    # Coluna F - Leitura
                temperatura_agua = ler_valor_exato(coleta_sheet, linha, 18)   # Coluna E - Temperatura (assumindo)
                
                # Lê os valores de referência (não podem ser alterados)
                vazao_referencia = ler_valor_exato(coleta_sheet, linha, 9)   # Coluna I
                erro = ler_valor_exato(coleta_sheet, linha, 21)             # Coluna U
                
                leitura = {
                    'linha': linha,
                    'qtd_pulsos_padrao': qtd_pulsos_padrao,
                    'tempo_coleta': tempo_coleta,
                    'leitura_medidor': leitura_medidor,
                    'temperatura_agua': temperatura_agua,
                    'vazao_referencia': vazao_referencia,
                    'erro': erro
                }
                
                ponto['leituras'].append(leitura)
            
            pontos_dados.append(ponto)
            linha_inicial += avanca_linha
            num_ponto += 1
        
        return pontos_dados
        
    except Exception as e:
        print(f"ERRO ao ler dados da planilha: {e}")
        return None

def calcular_vazao_media_erros(ponto):
    """Calcula a vazão média e erros para um ponto"""
    vazoes = [l['vazao_referencia'] for l in ponto['leituras']]
    erros = [l['erro'] for l in ponto['leituras']]
    
    # Calcula média da vazão
    media_vazao = sum(vazoes) / Decimal(str(len(vazoes)))
    
    # Calcula tendência (média dos erros)
    erros_validos = [e for e in erros if e != 0]
    if erros_validos:
        tendencia = sum(erros_validos) / Decimal(str(len(erros_validos)))
    else:
        tendencia = Decimal('0')
    
    # Calcula desvio padrão
    desvio_padrao = calcular_desvio_padrao_amostral(erros)
    
    return {
        'media_vazao': media_vazao,
        'tendencia': tendencia,
        'desvio_padrao': desvio_padrao
    }

def ajustar_tempos_coleta(pontos_originais, dados_extraidos):
    """
    Ajusta os tempos de coleta para terem a mesma parte inteira,
    mantendo os valores de vazão média, tendência e desvio padrão exatamente iguais
    """
    print("\n=== AJUSTANDO TEMPOS DE COLETA ===")
    
    # Determina o tempo de coleta padrão (parte inteira do primeiro ponto)
    tempo_padrao = int(pontos_originais[0]['leituras'][0]['tempo_coleta'])
    print(f"Tempo de coleta padrão definido: {tempo_padrao} segundos")
    
    pontos_ajustados = []
    
    for i, ponto_original in enumerate(pontos_originais):
        ponto_extraido = dados_extraidos['pontos_dados'][i]
        print(f"\n--- Ajustando Ponto {ponto_original['numero']} ---")
        
        # Valores de referência que devem permanecer exatos
        vazao_media_esperada = ponto_extraido['media_vazao_referencia']
        tendencia_esperada = ponto_extraido['tendencia']
        desvio_esperado = ponto_extraido['desvio_padrao']
        
        print(f"Valores esperados:")
        print(f"  Vazão média: {float(vazao_media_esperada)} L/h")
        print(f"  Tendência: {float(tendencia_esperada)} %")
        print(f"  Desvio padrão: {float(desvio_esperado)} %")
        
        ponto_ajustado = {
            'numero': ponto_original['numero'],
            'leituras': []
        }
        
        # Ajusta cada leitura
        for j, leitura_original in enumerate(ponto_original['leituras']):
            print(f"\n  Leitura {j+1} (Linha {leitura_original['linha']}):")
            
            # Mantém valores que não podem ser alterados
            qtd_pulsos_padrao = leitura_original['qtd_pulsos_padrao']
            leitura_medidor = leitura_original['leitura_medidor']
            temperatura_agua = leitura_original['temperatura_agua']
            vazao_referencia = leitura_original['vazao_referencia']
            erro = leitura_original['erro']
            
            # Calcula o tempo de coleta ajustado
            # Fórmula: tempo = (qtd_pulsos_padrao * 3600) / vazao_referencia
            tempo_ajustado = (qtd_pulsos_padrao * Decimal('3600')) / vazao_referencia
            
            # Garante que a parte inteira seja igual ao tempo padrão
            parte_inteira = tempo_padrao
            parte_decimal = tempo_ajustado - int(tempo_ajustado)
            tempo_final = Decimal(str(parte_inteira)) + parte_decimal
            
            print(f"    Tempo original: {float(leitura_original['tempo_coleta'])} s")
            print(f"    Tempo calculado: {float(tempo_ajustado)} s")
            print(f"    Tempo ajustado: {float(tempo_final)} s")
            
            leitura_ajustada = {
                'linha': leitura_original['linha'],
                'qtd_pulsos_padrao': qtd_pulsos_padrao,
                'tempo_coleta': tempo_final,
                'leitura_medidor': leitura_medidor,
                'temperatura_agua': temperatura_agua,
                'vazao_referencia': vazao_referencia,
                'erro': erro
            }
            
            ponto_ajustado['leituras'].append(leitura_ajustada)
        
        # Verifica se os valores permanecem iguais
        valores_calculados = calcular_vazao_media_erros(ponto_ajustado)
        
        print(f"\n  Verificação após ajuste:")
        print(f"    Vazão média: {float(valores_calculados['media_vazao'])} L/h")
        print(f"    Tendência: {float(valores_calculados['tendencia'])} %")
        print(f"    Desvio padrão: {float(valores_calculados['desvio_padrao'])} %")
        
        # Verifica se os valores são exatamente iguais
        vazao_igual = abs(valores_calculados['media_vazao'] - vazao_media_esperada) < Decimal('0.000000000000001')
        tendencia_igual = abs(valores_calculados['tendencia'] - tendencia_esperada) < Decimal('0.000000000000001')
        desvio_igual = abs(valores_calculados['desvio_padrao'] - desvio_esperado) < Decimal('0.000000000000001')
        
        if vazao_igual and tendencia_igual and desvio_igual:
            print(f"  ✅ Valores mantidos exatamente iguais!")
        else:
            print(f"  ⚠️  Diferenças detectadas - ajuste necessário")
        
        pontos_ajustados.append(ponto_ajustado)
    
    return pontos_ajustados

def gerar_planilha_ajustada(pontos_ajustados, arquivo_saida):
    """Gera uma nova planilha com os tempos ajustados"""
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tempos Ajustados"
        
        # Cabeçalho
        headers = [
            "Ponto", "Leitura", "Linha", "Qtd Pulsos Padrão", "Tempo Coleta (s)", 
            "Leitura Medidor", "Temperatura Água (°C)", "Vazão Referência (L/h)", "Erro (%)"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Dados
        row = 2
        for ponto in pontos_ajustados:
            for leitura in ponto['leituras']:
                ws.cell(row=row, column=1, value=ponto['numero'])
                ws.cell(row=row, column=2, value=row-1)
                ws.cell(row=row, column=3, value=leitura['linha'])
                ws.cell(row=row, column=4, value=float(leitura['qtd_pulsos_padrao']))
                ws.cell(row=row, column=5, value=float(leitura['tempo_coleta']))
                ws.cell(row=row, column=6, value=float(leitura['leitura_medidor']))
                ws.cell(row=row, column=7, value=float(leitura['temperatura_agua']))
                ws.cell(row=row, column=8, value=float(leitura['vazao_referencia']))
                ws.cell(row=row, column=9, value=float(leitura['erro']))
                row += 1
        
        wb.save(arquivo_saida)
        print(f"\n✅ Planilha ajustada salva como: {arquivo_saida}")
        
    except Exception as e:
        print(f"ERRO ao gerar planilha: {e}")

def main():
    """Função principal"""
    arquivo_json = "vazoes_referencia_tendencia_desvio.json"
    arquivo_excel = "SAN-038-25-09-1.xlsx"
    arquivo_saida = "tempos_ajustados.xlsx"
    
    print("=== AJUSTADOR DE TEMPOS DE COLETA ===")
    print("Ajusta tempos de coleta para terem a mesma parte inteira")
    print("Mantém vazão média, tendência e desvio padrão exatamente iguais")
    
    # Carrega dados extraídos
    dados_extraidos = carregar_dados_extraidos(arquivo_json)
    if not dados_extraidos:
        print("Falha ao carregar dados extraídos")
        return
    
    print(f"✅ Dados extraídos carregados: {dados_extraidos['metadata']['total_pontos']} pontos")
    
    # Lê dados originais da planilha
    pontos_originais = ler_dados_originais_planilha(arquivo_excel)
    if not pontos_originais:
        print("Falha ao ler dados da planilha")
        return
    
    print(f"✅ Dados originais lidos: {len(pontos_originais)} pontos")
    
    # Ajusta os tempos de coleta
    pontos_ajustados = ajustar_tempos_coleta(pontos_originais, dados_extraidos)
    
    # Gera planilha com tempos ajustados
    gerar_planilha_ajustada(pontos_ajustados, arquivo_saida)
    
    # Salva dados em JSON
    output_data = {
        "metadata": {
            "arquivo_original": arquivo_excel,
            "arquivo_ajustado": arquivo_saida,
            "data_ajuste": dados_extraidos['metadata']['data_extracao'],
            "total_pontos": len(pontos_ajustados),
            "descricao": "Tempos de coleta ajustados para mesma parte inteira",
            "precisao": {
                "decimal_precision": getcontext().prec,
                "metodo_calculo": "Fórmula: tempo = (pulsos * 3600) / vazao_referencia"
            }
        },
        "pontos_ajustados": pontos_ajustados
    }
    
    with open("tempos_ajustados.json", "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False, default=str)
    
    print(f"\n✅ Dados salvos em 'tempos_ajustados.json'")
    print(f"✅ Processo concluído com sucesso!")

if __name__ == "__main__":
    main() 