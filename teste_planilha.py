import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal com 15 casas decimais
    """
    if valor is None:
        return Decimal('0.000000000000000')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        decimal_valor = Decimal(valor_limpo)
        # Força precisão de 15 casas decimais
        return decimal_valor.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    decimal_valor = Decimal(str(valor))
    # Força precisão de 15 casas decimais
    return decimal_valor.quantize(Decimal('0.000000000000000'), rounding=ROUND_HALF_UP)

def get_numeric_value(df, row, col):
    """Extrai valor numérico de uma célula específica"""
    try:
        value = df.iloc[row, col]
        if pd.notna(value):
            decimal_value = converter_para_decimal_padrao(value)
            # Verifica se o valor é zero usando comparação com tolerância
            if abs(decimal_value) < Decimal('0.000000000000001'):
                return Decimal('0.000000000000000')
            return decimal_value
        return Decimal('0.000000000000000')
    except:
        return Decimal('0.000000000000000')

def testar_planilha():
    """Testa a leitura da planilha para entender o problema"""
    arquivo_excel = "SAN-038-25-09.xlsx"
    
    print(f"🔍 TESTANDO PLANILHA: {arquivo_excel}")
    print("=" * 50)
    
    # Carrega a planilha
    wb = load_workbook(arquivo_excel, data_only=True)
    coleta_sheet = wb["Coleta de Dados"]
    
    print("✅ Aba 'Coleta de Dados' carregada")
    
    # Carrega com pandas
    coleta_df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
    
    print(f"📊 Dimensões da planilha: {coleta_df.shape}")
    
    # Testa algumas linhas específicas
    linha_inicial = 50
    
    print(f"\n🔍 TESTANDO LINHA {linha_inicial}:")
    for i in range(3):
        linha = linha_inicial + 3 + i
        print(f"   Linha {linha}:")
        
        # Testa diferentes colunas
        for col in [2, 3, 6, 9, 15, 18, 21]:
            valor_pandas = coleta_df.iloc[linha, col] if linha < len(coleta_df) else "N/A"
            valor_openpyxl = coleta_sheet.cell(row=linha+1, column=col+1).value if linha+1 <= coleta_sheet.max_row else "N/A"
            
            print(f"     Coluna {col}: Pandas={valor_pandas}, OpenPyXL={valor_openpyxl}")
    
    # Testa a função get_numeric_value
    print(f"\n🔍 TESTANDO get_numeric_value:")
    for i in range(3):
        linha = linha_inicial + 3 + i
        pulsos = get_numeric_value(coleta_df, linha, 2)
        print(f"   Linha {linha}, Coluna 2 (Pulsos): {pulsos}")
        print(f"     É zero? {abs(pulsos) < Decimal('0.000000000000001')}")
    
    # Testa a lógica de detecção de pontos
    print(f"\n🔍 TESTANDO DETECÇÃO DE PONTOS:")
    linha_inicial = 50
    avanca_linha = 9
    num_ponto = 1
    
    while linha_inicial < 200:  # Limite para evitar loop infinito
        valores_nulos = 0
        print(f"\n   Testando ponto {num_ponto} na linha {linha_inicial}:")
        
        for i in range(3): 
            linha = linha_inicial + 3 + i
            if linha < len(coleta_df):
                pulsos = get_numeric_value(coleta_df, linha, 2)
                print(f"     Linha {linha}: Pulsos = {pulsos}")
                # Verifica se o valor é zero ou nulo usando tolerância
                if abs(pulsos) < Decimal('0.000000000000001') or pd.isna(pulsos):
                    valores_nulos += 1
                    print(f"       → Valor nulo detectado")
            else:
                valores_nulos += 1
                print(f"     Linha {linha}: Fora dos limites da planilha")
        
        print(f"   Valores nulos encontrados: {valores_nulos}/3")
        
        if valores_nulos == 3:
            print(f"   → Ponto {num_ponto} não encontrado (3 valores nulos)")
            break
        
        print(f"   → Ponto {num_ponto} encontrado!")
        linha_inicial += avanca_linha
        num_ponto += 1
        
        if num_ponto > 10:  # Limite de segurança
            break

if __name__ == "__main__":
    testar_planilha() 