import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import threading
from openpyxl import load_workbook
from datetime import datetime
import shutil

class FormatadorExcelMassa:
    def __init__(self, root):
        self.root = root
        self.root.title("Formatador de Excel em Massa")
        self.root.geometry("800x600")
        self.root.resizable(True, True)
        
        # Variáveis
        self.arquivos_selecionados = []
        self.processando = False
        
        # Configurar estilo
        self.configurar_estilo()
        
        # Criar interface
        self.criar_interface()
        
    def configurar_estilo(self):
        """Configura o estilo visual da aplicação"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar cores
        style.configure('TFrame', background='#f0f0f0')
        style.configure('TButton', 
                       background='#4CAF50', 
                       foreground='white',
                       font=('Arial', 10, 'bold'),
                       padding=10)
        style.configure('TLabel', 
                       background='#f0f0f0',
                       font=('Arial', 10))
        style.configure('Title.TLabel', 
                       background='#f0f0f0',
                       font=('Arial', 14, 'bold'),
                       foreground='#2E86AB')
        
    def criar_interface(self):
        """Cria a interface gráfica principal"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Título
        titulo = ttk.Label(main_frame, 
                          text="Formatador de Excel em Massa", 
                          style='Title.TLabel')
        titulo.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Descrição
        descricao = ttk.Label(main_frame, 
                             text="Selecione múltiplos arquivos Excel para aplicar formatação 'General' e revelar a precisão total dos números.",
                             wraplength=700)
        descricao.grid(row=1, column=0, columnspan=3, pady=(0, 20))
        
        # Botão Selecionar Arquivos
        self.btn_selecionar = ttk.Button(main_frame, 
                                        text="📁 Selecionar Arquivos Excel", 
                                        command=self.selecionar_arquivos)
        self.btn_selecionar.grid(row=2, column=0, pady=(0, 10), sticky=tk.W)
        
        # Label para mostrar quantidade de arquivos
        self.lbl_quantidade = ttk.Label(main_frame, text="Nenhum arquivo selecionado")
        self.lbl_quantidade.grid(row=2, column=1, pady=(0, 10), sticky=tk.W, padx=(10, 0))
        
        # Frame para lista de arquivos
        lista_frame = ttk.LabelFrame(main_frame, text="Arquivos Selecionados", padding="10")
        lista_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 20))
        lista_frame.columnconfigure(0, weight=1)
        lista_frame.rowconfigure(0, weight=1)
        
        # Lista de arquivos
        self.lista_arquivos = tk.Listbox(lista_frame, 
                                        height=8, 
                                        selectmode=tk.EXTENDED,
                                        font=('Consolas', 9))
        self.lista_arquivos.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar para lista
        scrollbar = ttk.Scrollbar(lista_frame, orient=tk.VERTICAL, command=self.lista_arquivos.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.lista_arquivos.configure(yscrollcommand=scrollbar.set)
        
        # Botões de ação
        botoes_frame = ttk.Frame(main_frame)
        botoes_frame.grid(row=4, column=0, columnspan=3, pady=(0, 20))
        
        # Botão Limpar Lista
        self.btn_limpar = ttk.Button(botoes_frame, 
                                    text="🗑️ Limpar Lista", 
                                    command=self.limpar_lista)
        self.btn_limpar.pack(side=tk.LEFT, padx=(0, 10))
        
        # Botão Iniciar Processamento
        self.btn_processar = ttk.Button(botoes_frame, 
                                       text="⚡ Iniciar Processamento", 
                                       command=self.iniciar_processamento,
                                       style='TButton')
        self.btn_processar.pack(side=tk.LEFT, padx=(0, 10))
        
        # Botão Abrir Pasta de Saída
        self.btn_abrir_pasta = ttk.Button(botoes_frame, 
                                         text="📂 Abrir Pasta de Saída", 
                                         command=self.abrir_pasta_saida)
        self.btn_abrir_pasta.pack(side=tk.LEFT)
        
        # Frame para log
        log_frame = ttk.LabelFrame(main_frame, text="Log de Processamento", padding="10")
        log_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Área de log
        self.log_area = scrolledtext.ScrolledText(log_frame, 
                                                 height=12, 
                                                 font=('Consolas', 9),
                                                 wrap=tk.WORD)
        self.log_area.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Barra de progresso
        self.progresso = ttk.Progressbar(main_frame, 
                                        mode='determinate',
                                        length=400)
        self.progresso.grid(row=6, column=0, columnspan=3, pady=(10, 0), sticky=(tk.W, tk.E))
        
        # Status
        self.lbl_status = ttk.Label(main_frame, text="Pronto para processar arquivos")
        self.lbl_status.grid(row=7, column=0, columnspan=3, pady=(5, 0))
        
        # Log inicial
        self.log("=== Formatador de Excel em Massa ===")
        self.log("Aplicação iniciada. Selecione os arquivos Excel para começar.")
        self.log("")
        
    def selecionar_arquivos(self):
        """Abre diálogo para selecionar múltiplos arquivos Excel"""
        arquivos = filedialog.askopenfilenames(
            title="Selecionar Arquivos Excel",
            filetypes=[
                ("Arquivos Excel", "*.xlsx"),
                ("Todos os arquivos", "*.*")
            ]
        )
        
        if arquivos:
            self.arquivos_selecionados = list(arquivos)
            self.atualizar_lista_arquivos()
            self.log(f"✅ {len(self.arquivos_selecionados)} arquivo(s) selecionado(s)")
            
    def atualizar_lista_arquivos(self):
        """Atualiza a lista de arquivos na interface"""
        self.lista_arquivos.delete(0, tk.END)
        
        for arquivo in self.arquivos_selecionados:
            nome_arquivo = os.path.basename(arquivo)
            self.lista_arquivos.insert(tk.END, nome_arquivo)
            
        self.lbl_quantidade.config(text=f"{len(self.arquivos_selecionados)} arquivo(s) selecionado(s)")
        
    def limpar_lista(self):
        """Limpa a lista de arquivos selecionados"""
        self.arquivos_selecionados = []
        self.atualizar_lista_arquivos()
        self.log("🗑️ Lista de arquivos limpa")
        
    def log(self, mensagem):
        """Adiciona mensagem ao log com timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {mensagem}\n"
        
        self.log_area.insert(tk.END, log_entry)
        self.log_area.see(tk.END)
        self.root.update_idletasks()
        
    def iniciar_processamento(self):
        """Inicia o processamento dos arquivos em thread separada"""
        if not self.arquivos_selecionados:
            messagebox.showwarning("Aviso", "Nenhum arquivo selecionado!")
            return
            
        if self.processando:
            messagebox.showinfo("Info", "Processamento já em andamento!")
            return
            
        # Iniciar processamento em thread separada
        thread = threading.Thread(target=self.processar_arquivos)
        thread.daemon = True
        thread.start()
        
    def processar_arquivos(self):
        """Processa todos os arquivos selecionados"""
        self.processando = True
        self.btn_processar.config(state='disabled')
        self.btn_selecionar.config(state='disabled')
        
        try:
            # Criar pasta de saída
            pasta_saida = "output_formatado"
            if not os.path.exists(pasta_saida):
                os.makedirs(pasta_saida)
                self.log(f"📁 Pasta de saída criada: {pasta_saida}")
            
            total_arquivos = len(self.arquivos_selecionados)
            self.progresso['maximum'] = total_arquivos
            self.progresso['value'] = 0
            
            self.log(f"🚀 Iniciando processamento de {total_arquivos} arquivo(s)...")
            self.log("")
            
            arquivos_processados = 0
            arquivos_com_erro = 0
            
            for i, arquivo_original in enumerate(self.arquivos_selecionados):
                try:
                    nome_arquivo = os.path.basename(arquivo_original)
                    self.lbl_status.config(text=f"Processando: {nome_arquivo}")
                    self.log(f"📄 Processando: {nome_arquivo}")
                    
                    # Carregar workbook
                    workbook = load_workbook(arquivo_original)
                    
                    # Processar cada planilha
                    planilhas_processadas = 0
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        planilhas_processadas += self.processar_planilha(worksheet)
                    
                    # Salvar arquivo processado
                    arquivo_saida = os.path.join(pasta_saida, nome_arquivo)
                    workbook.save(arquivo_saida)
                    workbook.close()
                    
                    arquivos_processados += 1
                    self.log(f"✅ Concluído: {nome_arquivo} ({planilhas_processadas} planilhas)")
                    
                except Exception as e:
                    arquivos_com_erro += 1
                    self.log(f"❌ Erro ao processar {nome_arquivo}: {str(e)}")
                
                # Atualizar progresso
                self.progresso['value'] = i + 1
                self.root.update_idletasks()
            
            # Resumo final
            self.log("")
            self.log("=" * 50)
            self.log("📊 RESUMO DO PROCESSAMENTO:")
            self.log(f"   • Total de arquivos: {total_arquivos}")
            self.log(f"   • Processados com sucesso: {arquivos_processados}")
            self.log(f"   • Erros: {arquivos_com_erro}")
            self.log(f"   • Pasta de saída: {pasta_saida}")
            self.log("=" * 50)
            
            if arquivos_com_erro == 0:
                self.log("🎉 Processamento concluído com sucesso!")
                messagebox.showinfo("Sucesso", f"Todos os {arquivos_processados} arquivos foram processados com sucesso!\n\nArquivos salvos em: {pasta_saida}")
            else:
                self.log("⚠️ Processamento concluído com alguns erros.")
                messagebox.showwarning("Aviso", f"Processamento concluído!\n\nSucessos: {arquivos_processados}\nErros: {arquivos_com_erro}\n\nVerifique o log para detalhes.")
                
        except Exception as e:
            self.log(f"❌ Erro geral: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante o processamento:\n{str(e)}")
            
        finally:
            self.processando = False
            self.btn_processar.config(state='normal')
            self.btn_selecionar.config(state='normal')
            self.lbl_status.config(text="Processamento concluído")
            
    def processar_planilha(self, worksheet):
        """Processa uma planilha individual, aplicando formatação 'General'"""
        celulas_processadas = 0
        
        # Iterar por todas as células com dados
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Aplicar formatação 'General' para mostrar precisão total
                    cell.number_format = 'General'
                    celulas_processadas += 1
                    
        return celulas_processadas
        
    def abrir_pasta_saida(self):
        """Abre a pasta de saída no explorador de arquivos"""
        pasta_saida = "output_formatado"
        
        if os.path.exists(pasta_saida):
            if os.name == 'nt':  # Windows
                os.startfile(pasta_saida)
            elif os.name == 'posix':  # macOS e Linux
                os.system(f'open "{pasta_saida}"' if os.uname().sysname == 'Darwin' else f'xdg-open "{pasta_saida}"')
            self.log(f"📂 Pasta de saída aberta: {pasta_saida}")
        else:
            messagebox.showinfo("Info", "Pasta de saída ainda não foi criada. Processe alguns arquivos primeiro.")

def main():
    """Função principal"""
    root = tk.Tk()
    app = FormatadorExcelMassa(root)
    
    # Centralizar janela
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    root.mainloop()

if __name__ == "__main__":
    main() 