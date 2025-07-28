# -*- coding: utf-8 -*-
"""
SOLUÇÃO FINAL - PADRONIZAÇÃO DE TEMPOS DE COLETA
================================================

Este é o script final que implementa exatamente o que você pediu:
1. ✅ Padroniza todos os tempos de coleta como 360 segundos (inteiros)
2. ✅ Recalcula todos os outros valores proporcionalmente
3. ✅ Mantém precisão Decimal de 28 dígitos
4. ✅ Trata corretamente o formato brasileiro (vírgula como separador)
5. ✅ Gera relatórios detalhados

RESULTADO ESPERADO:
- Tempos de coleta: 360, 360, 360 (todos iguais e inteiros)
- Valores recalculados proporcionalmente
- Precisão máxima garantida
"""

import pandas as pd
from decimal import Decimal, getcontext
from openpyxl import load_workbook
import json
import os

# Configurar precisão máxima
getcontext().prec = 28

def converter_para_decimal_padrao(valor):
    """
    Função padronizada para converter valores para Decimal
    Trata corretamente formato brasileiro (vírgula como separador decimal)
    """
    if valor is None:
        return Decimal('0')
    
    if isinstance(valor, str):
        # Remove espaços e pontos de milhares, substitui vírgula por ponto
        valor_limpo = valor.replace(' ', '').replace('.', '').replace(',', '.')
        return Decimal(valor_limpo)
    
    # Para valores numéricos, converter para string primeiro para preservar precisão
    return Decimal(str(valor))

def ler_valor_exato(sheet, linha, coluna):
    """
    Lê valor exato da planilha sem qualquer modificação
    """
    valor = sheet.cell(row=linha, column=coluna).value
    return converter_para_decimal_padrao(valor)

def processar_planilha_completa():
    """
    Processa a planilha completa com padronização de tempos
    """
    print("🎯 SOLUÇÃO FINAL - PADRONIZAÇÃO DE TEMPOS DE COLETA")
    print("=" * 70)
    
    wb = load_workbook("SAN-038-25-09-1.xlsx", data_only=True)
    coleta = wb["Coleta de Dados"]
    
    # Linhas dos pontos de calibração
    linhas = [54, 55, 56]
    tempo_padrao = Decimal('360')  # Tempo padrão desejado
    
    print(f"\n📖 PROCESSANDO PLANILHA COMPLETA")
    print(f"   Tempo padrão: {tempo_padrao} segundos")
    print(f"   Linhas processadas: {linhas}")
    
    resultados = {}
    
    for linha in linhas:
        print(f"\n🔍 PROCESSANDO LINHA {linha}:")
        
        # Ler valores originais
        pulsos = ler_valor_exato(coleta, linha, 3)      # Coluna C
        tempo_original = ler_valor_exato(coleta, linha, 6)  # Coluna F
        volume_original = ler_valor_exato(coleta, linha, 9)  # Coluna I
        leitura_original = ler_valor_exato(coleta, linha, 15)  # Coluna O
        temperatura = ler_valor_exato(coleta, linha, 18)  # Coluna R
        
        print(f"   Valores originais:")
        print(f"     Pulsos: {pulsos}")
        print(f"     Tempo: {tempo_original}")
        print(f"     Volume: {volume_original}")
        print(f"     Leitura: {leitura_original}")
        print(f"     Temperatura: {temperatura}")
        
        # Calcular fator de correção
        fator = tempo_padrao / tempo_original
        
        # Recalcular valores proporcionalmente
        volume_corrigido = volume_original * fator
        leitura_corrigida = leitura_original * fator
        
        print(f"   Valores padronizados:")
        print(f"     Tempo: {tempo_padrao} ✅ (padronizado)")
        print(f"     Fator: {fator}")
        print(f"     Volume: {volume_corrigido}")
        print(f"     Leitura: {leitura_corrigida}")
        
        resultados[f"linha_{linha}"] = {
            "pulsos": pulsos,
            "tempo_original": tempo_original,
            "tempo_padronizado": tempo_padrao,
            "volume_original": volume_original,
            "volume_corrigido": volume_corrigido,
            "leitura_original": leitura_original,
            "leitura_corrigida": leitura_corrigida,
            "temperatura": temperatura,
            "fator_correcao": fator
        }
    
    return resultados

def calcular_medias_finais(resultados):
    """
    Calcula médias dos valores padronizados
    """
    print(f"\n🧮 CALCULANDO MÉDIAS FINAIS")
    
    volumes = [resultados[f"linha_{linha}"]["volume_corrigido"] for linha in [54, 55, 56]]
    leituras = [resultados[f"linha_{linha}"]["leitura_corrigida"] for linha in [54, 55, 56]]
    tempos = [resultados[f"linha_{linha}"]["tempo_padronizado"] for linha in [54, 55, 56]]
    
    media_volume = sum(volumes) / Decimal('3')
    media_leitura = sum(leituras) / Decimal('3')
    media_tempo = sum(tempos) / Decimal('3')
    
    print(f"   Média Volume: {media_volume}")
    print(f"   Média Leitura: {media_leitura}")
    print(f"   Média Tempo: {media_tempo} ✅ (deve ser 360)")
    
    return {
        "media_volume": media_volume,
        "media_leitura": media_leitura,
        "media_tempo": media_tempo
    }

def gerar_relatorio_final(resultados, medias):
    """
    Gera relatório final completo
    """
    print(f"\n📄 GERANDO RELATÓRIO FINAL...")
    
    # Salvar relatório detalhado
    with open("solucao_final_padronizacao.txt", "w", encoding="utf-8") as f:
        f.write("=== SOLUÇÃO FINAL - PADRONIZAÇÃO DE TEMPOS DE COLETA ===\n\n")
        
        f.write("🎯 OBJETIVO ALCANÇADO:\n")
        f.write("   • Padronizar todos os tempos de coleta como 360 segundos\n")
        f.write("   • Recalcular valores proporcionalmente\n")
        f.write("   • Manter precisão máxima\n\n")
        
        f.write("✅ CONFIGURAÇÕES:\n")
        f.write("   • Precisão: Decimal com 28 dígitos\n")
        f.write("   • Tempo padrão: 360 segundos (inteiros)\n")
        f.write("   • Formato brasileiro: Tratado corretamente\n\n")
        
        f.write("📊 RESULTADOS DETALHADOS:\n")
        for linha in [54, 55, 56]:
            dados = resultados[f"linha_{linha}"]
            f.write(f"   LINHA {linha}:\n")
            f.write(f"     Pulsos Padrão: {dados['pulsos']}\n")
            f.write(f"     Tempo Original: {dados['tempo_original']} → Padronizado: {dados['tempo_padronizado']} ✅\n")
            f.write(f"     Volume Original: {dados['volume_original']} → Corrigido: {dados['volume_corrigido']}\n")
            f.write(f"     Leitura Original: {dados['leitura_original']} → Corrigida: {dados['leitura_corrigida']}\n")
            f.write(f"     Temperatura: {dados['temperatura']}\n")
            f.write(f"     Fator de Correção: {dados['fator_correcao']}\n\n")
        
        f.write("🧮 MÉDIAS CALCULADAS:\n")
        f.write(f"   Volume Corrigido: {medias['media_volume']}\n")
        f.write(f"   Leitura Corrigida: {medias['media_leitura']}\n")
        f.write(f"   Tempo: {medias['media_tempo']} ✅\n\n")
        
        f.write("🎉 CONCLUSÃO:\n")
        f.write("   ✅ Todos os tempos de coleta padronizados como 360 segundos\n")
        f.write("   ✅ Valores recalculados proporcionalmente\n")
        f.write("   ✅ Precisão máxima mantida\n")
        f.write("   ✅ Formato brasileiro tratado corretamente\n")
    
    # Salvar dados em JSON para uso posterior
    dados_completos = {
        "resultados": resultados,
        "medias": medias,
        "configuracoes": {
            "tempo_padrao": "360 segundos",
            "precisao": "Decimal com 28 dígitos",
            "linhas_processadas": [54, 55, 56]
        },
        "timestamp": str(pd.Timestamp.now())
    }
    
    with open("solucao_final_padronizacao.json", "w", encoding="utf-8") as f:
        json.dump(dados_completos, f, indent=2, ensure_ascii=False, default=str)
    
    print("   ✅ Relatórios salvos:")
    print("      • solucao_final_padronizacao.txt")
    print("      • solucao_final_padronizacao.json")

def main():
    """
    Função principal
    """
    try:
        # 1. Processar planilha completa
        resultados = processar_planilha_completa()
        
        # 2. Calcular médias finais
        medias = calcular_medias_finais(resultados)
        
        # 3. Gerar relatório final
        gerar_relatorio_final(resultados, medias)
        
        print(f"\n🎯 SOLUÇÃO FINAL IMPLEMENTADA COM SUCESSO!")
        print(f"   ✅ Tempos de coleta padronizados como 360 segundos")
        print(f"   ✅ Valores recalculados proporcionalmente")
        print(f"   ✅ Precisão máxima mantida")
        print(f"   ✅ Relatórios gerados com sucesso")
        
        print(f"\n📋 RESUMO DOS RESULTADOS:")
        for linha in [54, 55, 56]:
            dados = resultados[f"linha_{linha}"]
            print(f"   Linha {linha}: {dados['tempo_original']}s → {dados['tempo_padronizado']}s ✅")
        
        print(f"\n🎉 MISSION ACCOMPLISHED!")
        print(f"   A planilha foi processada exatamente como você solicitou!")
        
    except Exception as e:
        print(f"❌ ERRO: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 