import pandas as pd
from openpyxl import load_workbook
from decimal import Decimal

def ler_valor_exato(sheet, linha, coluna):
    """Lê valor exato da planilha"""
    valor = sheet.cell(row=linha, column=coluna).value
    return valor

def main():
    arquivo_excel = "SAN-038-25-09-1.xlsx"
    
    print("=== TESTE DE COLUNAS ===")
    
    # Carregar planilha
    wb = load_workbook(arquivo_excel, data_only=True)
    coleta_sheet = wb["Coleta de Dados"]
    
    # Testar diferentes linhas e colunas
    linhas_teste = [54, 55, 56, 63, 64, 65]
    
    for linha in linhas_teste:
        print(f"\n--- Linha {linha} ---")
        for col in range(1, 11):  # Testar colunas A até J
            valor = ler_valor_exato(coleta_sheet, linha, col)
            if valor is not None and valor != 0:
                print(f"  Coluna {col}: {valor}")
    
    # Testar também com pandas
    print(f"\n=== TESTE COM PANDAS ===")
    df = pd.read_excel(arquivo_excel, sheet_name='Coleta de Dados', header=None)
    
    for linha in linhas_teste:
        print(f"\n--- Linha {linha} (Pandas) ---")
        for col in range(10):  # Testar colunas 0 até 9
            valor = df.iloc[linha-1, col]
            if pd.notna(valor) and valor != 0:
                print(f"  Coluna {col}: {valor}")

if __name__ == "__main__":
    main() 